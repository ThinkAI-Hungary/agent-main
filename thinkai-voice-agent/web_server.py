from typing import Optional
"""
ThinkAI Voice Agent — Web Server
Serves the voice widget, generates LiveKit tokens,
and provides a JWT-protected admin API with analytics.
"""

import json
import re
import logging
import os
import uuid
import csv
import io
from datetime import datetime, timedelta
from pathlib import Path

import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, Depends, HTTPException, Request, status, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
import jwt as pyjwt

from livekit.api import AccessToken, VideoGrants, RoomConfiguration, RoomAgentDispatch
import livekit.api as lk_api_module
import asyncio

import database as db
import email_processor
from classifier import classify_interaction
from anthropic import AsyncAnthropic

THIS_DIR = Path(__file__).resolve().parent
load_dotenv(THIS_DIR / ".env")
logger = logging.getLogger(__name__)

# ── JWT config ────────────────────────────────────────────────────────────────
JWT_SECRET  = os.getenv("JWT_SECRET", "thinkai-admin-secret-change-me")
JWT_ALGO    = "HS256"
JWT_EXPIRES = 60 * 60 * 8  # 8 hours

# ── Init DB on startup ────────────────────────────────────────────────────────
db.init_db()
db.seed_admin_from_env()
db.migrate_from_json()   # one-time migration from legacy JSON files

app = FastAPI(title="ThinkAI Voice Agent")

background_tasks = set()

# ── Health check (Docker / monitoring) ────────────────────────────────────────
_start_time = datetime.utcnow()

@app.get("/api/health")
async def health_check():
    return {"status": "ok", "uptime_seconds": int((datetime.utcnow() - _start_time).total_seconds())}


# ── Social Publisher Worker ─────────────────────────────────────────────────────
async def social_publisher_worker():
    """Háttér worker: percenként ellenőrzi az ütemezett tartalmakat és publikálja."""
    logger.info("Social publisher worker elindítva.")
    while True:
        try:
            await asyncio.sleep(60)  # 60 másodpercenként
            scheduled = db.get_scheduled_content()
            for item in scheduled:
                content_id = item.get("id")
                caption = item.get("body", "")
                image_url = item.get("image_url", "")
                platforms = item.get("target_platforms", ["instagram"])
                published_platforms = []

                logger.info(f"Ütemezett poszt publikálás: {content_id} -> {platforms}")

                for platform in platforms:
                    try:
                        if platform == "instagram" and image_url:
                            result = await social_media.publish_instagram_post(image_url, caption)
                            if result.get("success"):
                                published_platforms.append("instagram")
                                db.update_content_item(content_id, {"ig_media_id": result.get("media_id", "")})
                        elif platform == "facebook":
                            result = await social_media.publish_facebook_post(caption, image_url or None)
                            if result.get("success"):
                                published_platforms.append("facebook")
                                db.update_content_item(content_id, {"fb_post_id": result.get("post_id", "")})
                    except Exception as pub_err:
                        logger.error(f"Publish error for {content_id} on {platform}: {pub_err}")

                if published_platforms:
                    db.update_content_item(content_id, {
                        "status": "published",
                        "published_at": datetime.utcnow().isoformat() + "Z",
                        "published_platforms": published_platforms,
                    })
                    logger.info(f"Ütemezett poszt sikeresen publikálva: {content_id} -> {published_platforms}")
                else:
                    logger.warning(f"Ütemezett poszt nem sikerült: {content_id}")
        except Exception as e:
            logger.error(f"Social publisher worker error: {e}")

async def campaign_scheduler_worker():
    """Háttérfolyamat: 30 másodpercenként ellenőrzi az ütemezett kampányokat és
    elindítja őket — MINDEN aktív tenantnak (FÁZIS 6 multi-tenant)."""
    from datetime import datetime
    import zoneinfo
    local_tz = zoneinfo.ZoneInfo("Europe/Budapest")
    while True:
        await asyncio.sleep(30)
        try:
            # Tenant-iteráció: minden tenant kampányai a saját scope-jukban
            all_campaigns = []
            for tenant in db.get_active_tenants():
                try:
                    db.set_current_tenant(tenant["id"])
                    for c in db.get_campaigns():
                        all_campaigns.append((tenant, c))
                except Exception as t_err:
                    logger.error(f"[Scheduler] Tenant kampány-lekérdezés hiba ({tenant.get('slug')}): {t_err}")
            now = datetime.now(local_tz)
            for _tenant, c in all_campaigns:
                db.set_current_tenant(_tenant["id"])  # a kampány saját tenantja legyen a kontextus
                if c.get("status") != "Ütemezett":
                    continue
                # Read scheduled_at from SCHED: prefix in ai_instructions
                ai_inst = c.get("ai_instructions", "") or ""
                if not ai_inst.startswith("SCHED:"):
                    continue
                pipe_idx = ai_inst.find("|")
                if pipe_idx < 0:
                    continue
                sched = ai_inst[6:pipe_idx]  # Extract datetime string
                if not sched:
                    continue
                try:
                    dt = datetime.fromisoformat(sched)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=local_tz)
                    if dt <= now:
                        logger.info(f"[Scheduler] Kampány indítása: {c['name']} (id={c['id']}, scheduled_at={sched})")
                        campaign_id = c["id"]
                        # Restore original ai_instructions (remove SCHED: prefix)
                        original_instructions = ai_inst[pipe_idx + 1:]
                        try:
                            db.supabase.table("campaigns").update({"ai_instructions": original_instructions}).eq("id", campaign_id).execute()
                        except Exception:
                            pass
                        # Start the campaign
                        channels = c.get("channels", [c.get("channel", "email")])
                        supported = {"email", "messenger", "telefon"}
                        active_channels = [ch for ch in channels if ch in supported]
                        if not active_channels:
                            logger.warning(f"[Scheduler] Nem támogatott csatorna: {channels}")
                            continue
                        db.update_campaign_status(campaign_id, "Aktív")
                        text_channels = [ch for ch in active_channels if ch in {"email", "messenger"}]
                        if text_channels:
                            task = asyncio.create_task(_run_campaign(c, text_channels))
                            background_tasks.add(task)
                            task.add_done_callback(background_tasks.discard)
                        if "telefon" in active_channels:
                            phone_task = asyncio.create_task(_run_phone_campaign(c))
                            background_tasks.add(phone_task)
                            phone_task.add_done_callback(background_tasks.discard)
                        logger.info(f"[Scheduler] Kampány sikeresen elindítva: {c['name']}")
                except Exception as parse_err:
                    logger.error(f"[Scheduler] Dátum parse hiba: {sched} - {parse_err}")
        except Exception as e:
            logger.error(f"[Scheduler] Campaign scheduler error: {e}")


@app.on_event("startup")
async def startup_event():
    # Elindítjuk az email worker loopot a háttérben
    task = asyncio.create_task(email_processor.email_worker_loop())
    background_tasks.add(task)
    task.add_done_callback(background_tasks.discard)
    
    task2 = asyncio.create_task(email_processor.reminder_worker_loop())
    background_tasks.add(task2)
    task2.add_done_callback(background_tasks.discard)
    
    # Eseményvezérelt automatizációk worker
    task3 = asyncio.create_task(email_processor.automation_worker_loop())
    background_tasks.add(task3)
    task3.add_done_callback(background_tasks.discard)

    # Social média ütemezett posztolás worker
    task4 = asyncio.create_task(social_publisher_worker())
    background_tasks.add(task4)
    task4.add_done_callback(background_tasks.discard)

    # Kampány ütemezés worker
    task5 = asyncio.create_task(campaign_scheduler_worker())
    background_tasks.add(task5)
    task5.add_done_callback(background_tasks.discard)

    # Inbound SIP szoba monitor — KIKAPCSOLVA
    # A lk_trigger.py (Asterisk) mar kezeli a dispatch-et, nem kell dupla.
    # mon = asyncio.create_task(inbound_sip_room_monitor())
    # background_tasks.add(mon)
    # mon.add_done_callback(background_tasks.discard)

    # Inbound SIP dispatch: Telnyx → LiveKit dispatch rule (SDR_KjoiKH4icXeX)
    # automatikusan kezeli, nem kell külön monitor vagy trigger.

async def inbound_sip_room_monitor():
    """Figyeli a 'call-' prefix szobakat es dispatch-eli az agentet ha meg nem csatlakozott.
    KIKAPCSOLVA: A lk_trigger.py mar kezeli az inbound dispatch-et.
    Ez a monitor dupla dispatch-et okozott (ket agent csatlakozott egy szobaba).
    """
    return  # lk_trigger.py kezeli
    lk_url    = os.getenv("LIVEKIT_URL", "").replace("wss://", "https://")
    lk_key    = os.getenv("LIVEKIT_API_KEY", "")
    lk_secret = os.getenv("LIVEKIT_API_SECRET", "")
    dispatched = set()  # mar dispatch-elt szobak
    while True:
        try:
            await asyncio.sleep(3)
            if not lk_key or not lk_secret:
                continue
            lk = lk_api_module.LiveKitAPI(url=lk_url, api_key=lk_key, api_secret=lk_secret)
            rooms = await lk.room.list_rooms(lk_api_module.ListRoomsRequest())
            for room in rooms.rooms:
                if not room.name.startswith("call-"):
                    continue
                if room.name in dispatched:
                    continue
                # Van-e mar agent a szobaban?
                parts = await lk.room.list_participants(lk_api_module.ListParticipantsRequest(room=room.name))
                has_agent = any(p.identity.startswith("agent-")
                               for p in parts.participants)
                if not has_agent:
                    await lk.agent_dispatch.create_dispatch(
                        lk_api_module.CreateAgentDispatchRequest(
                            agent_name="dobozos-ai",
                            room=room.name,
                            metadata="inbound_sip",
                        )
                    )
                    dispatched.add(room.name)
                    print(f"[SIP Monitor] Agent dispatch -> {room.name}", flush=True)
            await lk.aclose()
        except Exception as e:
            pass  # csendben folytatjuk

# ── CORS ──────────────────────────────────────────────────────────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://thinkai.hu",
        "https://www.thinkai.hu",
        "http://localhost:3000",
        "http://localhost:5173",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Auth helpers ──────────────────────────────────────────────────────────────
bearer_scheme = HTTPBearer(auto_error=False)


def create_jwt(username: str, tenant_id: str = "", role: str = "") -> str:
    payload = {
        "sub": username,
        "tenant_id": tenant_id,
        "role": role,
        "exp": datetime.utcnow() + timedelta(seconds=JWT_EXPIRES),
        "iat": datetime.utcnow(),
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


def verify_jwt(credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    if not credentials:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Nincs token")
    try:
        payload = pyjwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        return payload["sub"]
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token lejárt")
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Érvénytelen token")


def _decode_jwt_claims(credentials: HTTPAuthorizationCredentials | None) -> dict | None:
    """JWT dekódolás — a teljes payload (sub, tenant_id, role), vagy None."""
    if not credentials:
        return None
    try:
        return pyjwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
    except Exception:
        return None


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)) -> dict:
    """Dependency: a teljes aktuális felhasználó (username, role, tenant_id).
    A verify_jwt-hez hasonlóan 401-et dob hibás token esetén, de a teljes
    user-objektumot adja (a ~5 endpointnak, amelyiknek tényleg kell a user)."""
    payload = _decode_jwt_claims(credentials)
    if payload is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Érvénytelen token")
    username = payload.get("sub")
    # A role-t a DB-ből frissen olvassuk (a JWT-ben lévő lehet régi), a tenantot a JWT-ből.
    user = db.get_admin_user_by_username(username) or {}
    return {
        "username": username,
        "role": user.get("role") or payload.get("role", ""),
        "tenant_id": payload.get("tenant_id") or user.get("tenant_id"),
    }


# ── Tenant middleware: minden requesthez beállítja a tenant-kontextust ────────
# A JWT-ben lévő tenant_id-t a database.py contextvar-jába tölti — így a
# ~86 "csak gate" endpoint változatlan marad, a db-lekérdezések pedig már a
# helyes tenant scope-ban futnak. Hitelesítetlen (publikus/webhook) requesteknél
# a contextvar a DEFAULT_TENANT_SLUG-ra esik vissza.
@app.middleware("http")
async def tenant_context_middleware(request: Request, call_next):
    auth = request.headers.get("Authorization", "")
    tenant_id = None
    if auth.startswith("Bearer "):
        try:
            payload = pyjwt.decode(auth[7:], JWT_SECRET, algorithms=[JWT_ALGO])
            tenant_id = payload.get("tenant_id")
            # Tegyük elérhetővé a teljes usert is a request.state-en
            request.state.jwt_payload = payload
        except Exception:
            tenant_id = None
    if tenant_id:
        db.set_current_tenant(tenant_id)
    # Ha nincs tenant (publikus), nem állítjuk be — a get_current_tenant()
    # a DEFAULT_TENANT_SLUG-ra esik vissza.
    response = await call_next(request)
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/")
async def index():
    """Főoldal → átirányítás a login page-re (/admin)."""
    return RedirectResponse(url="/admin/", status_code=302)

@app.get("/demo")
@app.get("/demo/")
async def demo_widget():
    """A voice widget demo oldal (korábban a főoldal volt)."""
    return FileResponse(THIS_DIR / "voice-widget.html")

@app.get("/widget")
async def widget():
    return FileResponse(THIS_DIR / "voice-widget.html")

# ── React frontend (SPA + statikus fájlok) ──────────────────────────────────────
FRONTEND_DIST = THIS_DIR / "frontend_dist"
if FRONTEND_DIST.exists():
    app.mount("/admin/assets", StaticFiles(directory=FRONTEND_DIST / "assets"), name="frontend-assets")

# NOTE: admin SPA catch-all moved to end of file to avoid intercepting API routes

@app.get("/marketing")
def marketing_page():
    return FileResponse(THIS_DIR / "marketing.html")

@app.get("/marketing/elemzes")
def marketing_elemzes():
    return FileResponse(THIS_DIR / "elemzes.html")


# ── Publikus adatvédelmi oldalak (Meta App Verification: auth nélkül elérhető) ─
# A hivatalos (kanonikus) Privacy Policy-k az eaisy.hu-n élnek:
#   - https://eaisy.hu/privacy              — vállalati (ThinkAI Kft., minden termék)
#   - https://eaisy.hu/eaisydesk/privacy    — eaisyDesk termékspecifikus (ezt adjuk
#                                             meg a Meta-nak, mert ez írja le a
#                                             vizsgált app adatkezelését pontosan)
# Ezek a backend-endpointok is szolgálják a markdown-tartalmat HTML-ben, hogy a
# Meta crawlerek a saját domainünkön is megtalálják (és ne csak az eaisy.hu-n).
PRIVACY_POLICY_URL = os.getenv("PRIVACY_POLICY_URL", "https://eaisy.hu/eaisydesk/privacy")
CORPORATE_PP_URL = os.getenv("CORPORATE_PP_URL", "https://eaisy.hu/privacy")
PRIVACY_POLICY_FILE = THIS_DIR / "legal" / "PRIVACY_POLICY_eaisydesk.md"
CORPORATE_PRIVACY_FILE = THIS_DIR / "legal" / "PRIVACY_POLICY.md"


def _render_md_as_html(md: str, title: str, description: str) -> str:
    """Markdown tájékoztató HTML-oldallá alakítása (egyszerű, escape-elt <pre>
    blokk — nem igényel külső markdown-függőséget)."""
    import html as _html
    body = _html.escape(md)
    return f"""<!doctype html>
<html lang="hu"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{_html.escape(title)}</title>
<meta name="description" content="{_html.escape(description)}">
<meta property="og:title" content="{_html.escape(title)}">
<meta property="og:type" content="article">
</head>
<body style="max-width:820px;margin:24px auto;padding:0 16px;font-family:-apple-system,Segoe UI,Arial,sans-serif;line-height:1.6;color:#1f2937">
<pre style="white-space:pre-wrap;word-wrap:break-word;font-family:inherit;font-size:15px;background:#f9fafb;padding:24px;border-radius:8px;border:1px solid #e5e7eb">{body}</pre>
</body></html>"""


@app.get("/privacy")
async def public_privacy_policy():
    """Vállalati szintű Privacy Policy (ThinkAI Kft., minden termék). A Meta-nak
    a termékspecifikus eaisyDesk-t adjuk meg (l. /eaisydesk/privacy), de ez a
    kanonikus vállalati dokumentum."""
    try:
        md = CORPORATE_PRIVACY_FILE.read_text(encoding="utf-8")
    except Exception:
        md = "# Adatvédelmi tájékoztató — ThinkAI Kft.\n\nLásd: " + CORPORATE_PP_URL
    return HTMLResponse(content=_render_md_as_html(
        md,
        "Adatvédelmi tájékoztató — ThinkAI Kft.",
        "ThinkAI Kft. vállalati adatvédelmi tájékoztató"
    ))


@app.get("/eaisydesk/privacy")
async def public_eaisydesk_privacy_policy():
    """eaisyDesk termékspecifikus Privacy Policy — EZT adjuk meg a Meta
    App Dashboardon, mert pontosan a vizsgált app (eaisyDesk) adatkezelését
    írja le (Meta permission API-k, adatforrások, törlés)."""
    try:
        md = PRIVACY_POLICY_FILE.read_text(encoding="utf-8")
    except Exception:
        md = "# eaisyDesk — Privacy Policy\n\nLásd: " + PRIVACY_POLICY_URL
    return HTMLResponse(content=_render_md_as_html(
        md,
        "eaisyDesk — Adatvédelmi tájékoztató",
        "eaisyDesk termékspecifikus adatvédelmi tájékoztató / Privacy Policy"
    ))


@app.get("/eaisydesk/aszf")
async def public_eaisydesk_aszf():
    """eaisyDesk Általános Szerződési Feltételek (ÁSZF) — publikus."""
    try:
        md = (THIS_DIR / "legal" / "TERMS_OF_SERVICE.md").read_text(encoding="utf-8")
    except Exception:
        md = "# ÁSZF\n\nLásd: https://eaisy.hu/eaisydesk/aszf"
    return HTMLResponse(content=_render_md_as_html(
        md,
        "eaisyDesk — Általános Szerződési Feltételek (ÁSZF)",
        "eaisyDesk Általános Szerződési Feltételek"
    ))


@app.get("/eaisydesk/dpa")
async def public_eaisydesk_dpa():
    """eaisyDesk Adatkezelési Megállapodás (DPA) — publikus."""
    try:
        md = (THIS_DIR / "legal" / "DATA_PROCESSING_AGREEMENT.md").read_text(encoding="utf-8")
    except Exception:
        md = "# Adatkezelési Megállapodás\n\nLásd: https://eaisy.hu/eaisydesk/dpa"
    return HTMLResponse(content=_render_md_as_html(
        md,
        "eaisyDesk — Adatkezelési Megállapodás (DPA)",
        "eaisyDesk Adatkezelési Megállapodás / Data Processing Agreement"
    ))


@app.get("/gdpr")
async def public_gdpr_info():
    """A GDPR-oldal is publikusan elérhető — átirányít a Privacy Policy-re,
    hogy egyetlen kanonikus forrás maradjon."""
    return RedirectResponse(url="/eaisydesk/privacy", status_code=301)


@app.get("/robots.txt")
async def robots_txt():
    """A Meta crawlerei számára elérhető (kötelező). Semmit sem blokkolunk —
    a Privacy Policy és a publikus oldalak indexelhetők."""
    content = (
        "User-agent: *\n"
        "Allow: /\n"
        "Allow: /privacy\n"
        "Allow: /eaisydesk/privacy\n"
        "Allow: /eaisydesk/aszf\n"
        "Allow: /eaisydesk/dpa\n"
        "Disallow: /admin/\n"
        "Disallow: /api/\n"
        f"\n# Corporate Privacy Policy: {CORPORATE_PP_URL}\n"
        f"# eaisyDesk Privacy Policy: {PRIVACY_POLICY_URL}\n"
    )
    return PlainTextResponse(content=content, media_type="text/plain")



# ═══════════════════════════════════════════════════════════════════════════════
# MARKETING API — Email Campaigns & Subscribers
# ═══════════════════════════════════════════════════════════════════════════════

import brevo_campaigns

@app.get("/marketing/api/campaigns")
async def marketing_get_campaigns():
    """Kampányok listázása."""
    return db.get_email_campaigns()

@app.post("/marketing/api/campaigns")
async def marketing_create_campaign(req: Request):
    """Új kampány létrehozása."""
    data = await req.json()
    campaign = db.create_email_campaign(data)
    if not campaign:
        return JSONResponse({"error": "Kampány létrehozása sikertelen"}, status_code=500)
    return campaign

@app.get("/marketing/api/campaigns/stats")
async def marketing_campaigns_stats():
    """Kampány KPI összesítés."""
    return db.get_email_campaign_stats_summary()

@app.get("/marketing/api/campaigns/{campaign_id}")
async def marketing_get_campaign(campaign_id: str):
    """Egy kampány részletei."""
    campaign = db.get_email_campaign(campaign_id)
    if not campaign:
        return JSONResponse({"error": "Kampány nem található"}, status_code=404)
    return campaign

@app.put("/marketing/api/campaigns/{campaign_id}")
async def marketing_update_campaign(campaign_id: str, req: Request):
    """Kampány szerkesztése."""
    data = await req.json()
    ok = db.update_email_campaign(campaign_id, data)
    if not ok:
        return JSONResponse({"error": "Frissítés sikertelen"}, status_code=500)
    return {"ok": True}

@app.delete("/marketing/api/campaigns/{campaign_id}")
async def marketing_delete_campaign(campaign_id: str):
    """Kampány törlése."""
    ok = db.delete_email_campaign(campaign_id)
    if not ok:
        return JSONResponse({"error": "Törlés sikertelen"}, status_code=500)
    return {"ok": True}

@app.post("/marketing/api/campaigns/{campaign_id}/send")
async def marketing_send_campaign(campaign_id: str):
    """Kampány küldése Brevo-n keresztül.

    Lépések:
    1. Kampány lekérése DB-ből
    2. Brevo lista biztosítása (auto-create)
    3. Feliratkozók szinkronizálása a listára
    4. Brevo kampány létrehozása
    5. Azonnali küldés
    6. DB frissítése (brevo_campaign_id, status, sent_at)
    """
    campaign = db.get_email_campaign(campaign_id)
    if not campaign:
        return JSONResponse({"error": "Kampány nem található"}, status_code=404)

    if campaign.get("status") == "sent":
        return JSONResponse({"error": "Ez a kampány már el lett küldve"}, status_code=400)

    # 1. Ensure Brevo marketing list exists
    list_id = await brevo_campaigns.ensure_marketing_list()
    if not list_id:
        return JSONResponse({"error": "Brevo lista létrehozása sikertelen"}, status_code=500)

    # 2. Sync subscribers to Brevo list
    subscribers = db.get_email_subscribers()
    active_subs = [s for s in subscribers if s.get("status") == "active"]
    if not active_subs:
        return JSONResponse({"error": "Nincsenek aktív feliratkozók"}, status_code=400)

    synced = await brevo_campaigns.sync_contacts_batch(active_subs, list_id)
    logger.info(f"Synced {synced} contacts to Brevo list {list_id}")

    # 3. Create campaign in Brevo
    html = campaign.get("template_html") or f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2>{campaign.get('subject_line', 'Hírlevél')}</h2>
        <p>Tartalom hamarosan...</p>
    </div>
    """
    brevo_id = await brevo_campaigns.create_campaign(
        name=campaign.get("name", "Névtelen kampány"),
        subject=campaign.get("subject_line", "Hírlevél"),
        html_content=html,
        list_id=list_id,
        subject_b=campaign.get("subject_line_b")
    )
    if not brevo_id:
        return JSONResponse({"error": "Brevo kampány létrehozása sikertelen"}, status_code=500)

    # 4. Send now
    from datetime import datetime, timezone
    sent = await brevo_campaigns.send_campaign_now(brevo_id)
    if not sent:
        return JSONResponse({"error": "Brevo kampány küldése sikertelen"}, status_code=500)

    # 5. Update DB
    db.update_email_campaign(campaign_id, {
        "brevo_campaign_id": brevo_id,
        "status": "sent",
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "recipients_count": len(active_subs)
    })

    return {"ok": True, "brevo_campaign_id": brevo_id, "recipients": len(active_subs)}

@app.post("/marketing/api/campaigns/{campaign_id}/schedule")
async def marketing_schedule_campaign(campaign_id: str, request: Request):
    """Kampány ütemezése jövőbeli időpontra.

    Body: { "scheduled_at": "2026-06-03T10:00:00" }  (ISO 8601, lokális vagy UTC)
    """
    body = await request.json()
    scheduled_at = body.get("scheduled_at")
    if not scheduled_at:
        return JSONResponse({"error": "Hiányzó scheduled_at mező"}, status_code=400)

    campaign = db.get_email_campaign(campaign_id)
    if not campaign:
        return JSONResponse({"error": "Kampány nem található"}, status_code=404)

    if campaign.get("status") == "sent":
        return JSONResponse({"error": "Ez a kampány már el lett küldve"}, status_code=400)

    # 1. Ensure Brevo marketing list
    list_id = await brevo_campaigns.ensure_marketing_list()
    if not list_id:
        return JSONResponse({"error": "Brevo lista létrehozása sikertelen"}, status_code=500)

    # 2. Sync subscribers
    subscribers = db.get_email_subscribers()
    active_subs = [s for s in subscribers if s.get("status") == "active"]
    if not active_subs:
        return JSONResponse({"error": "Nincsenek aktív feliratkozók"}, status_code=400)

    await brevo_campaigns.sync_contacts_batch(active_subs, list_id)

    # 3. Create campaign in Brevo (if not already created)
    brevo_id = campaign.get("brevo_campaign_id")
    if not brevo_id:
        html = campaign.get("template_html") or f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2>{campaign.get('subject_line', 'Hírlevél')}</h2>
            <p>Tartalom hamarosan...</p>
        </div>
        """
        brevo_id = await brevo_campaigns.create_campaign(
            name=campaign.get("name", "Névtelen kampány"),
            subject=campaign.get("subject_line", "Hírlevél"),
            html_content=html,
            list_id=list_id,
            subject_b=campaign.get("subject_line_b")
        )
        if not brevo_id:
            return JSONResponse({"error": "Brevo kampány létrehozása sikertelen"}, status_code=500)

    # 4. Schedule in Brevo
    # Ensure UTC ISO format
    from datetime import datetime, timezone
    try:
        dt = datetime.fromisoformat(scheduled_at.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            # Assume local time, convert to UTC (CET = UTC+2)
            import zoneinfo
            local_tz = zoneinfo.ZoneInfo("Europe/Budapest")
            dt = dt.replace(tzinfo=local_tz).astimezone(timezone.utc)
        utc_iso = dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    except Exception:
        utc_iso = scheduled_at

    scheduled = await brevo_campaigns.schedule_campaign(brevo_id, utc_iso)
    if not scheduled:
        return JSONResponse({"error": "Brevo ütemezés sikertelen"}, status_code=500)

    # 5. Update DB
    db.update_email_campaign(campaign_id, {
        "brevo_campaign_id": brevo_id,
        "status": "scheduled",
        "scheduled_at": scheduled_at,
        "recipients_count": len(active_subs)
    })

    return {"ok": True, "brevo_campaign_id": brevo_id, "scheduled_at": scheduled_at, "recipients": len(active_subs)}

@app.post("/marketing/api/campaigns/{campaign_id}/refresh-stats")
async def marketing_refresh_stats(campaign_id: str):
    """Kampány statisztikák frissítése Brevo-ból."""
    campaign = db.get_email_campaign(campaign_id)
    if not campaign or not campaign.get("brevo_campaign_id"):
        return JSONResponse({"error": "Nincs Brevo kampány ID"}, status_code=400)

    stats = await brevo_campaigns.get_campaign_stats(campaign.get("brevo_campaign_id"))
    db.update_email_campaign(campaign_id, {"stats": stats})
    return stats

@app.post("/marketing/api/campaigns/refresh-all-stats")
async def marketing_refresh_all_stats():
    """Összes elküldött kampány statisztikáinak frissítése Brevo-ból."""
    campaigns = db.get_email_campaigns()
    sent_campaigns = [c for c in campaigns if c.get("status") == "sent" and c.get("brevo_campaign_id")]
    refreshed = 0
    for c in sent_campaigns:
        try:
            stats = await brevo_campaigns.get_campaign_stats(c["brevo_campaign_id"])
            db.update_email_campaign(c["id"], {"stats": stats})
            refreshed += 1
        except Exception as e:
            logger.error(f"Stats refresh error for campaign {c['id']}: {e}")
    return {"ok": True, "refreshed": refreshed, "total_sent": len(sent_campaigns)}

# ── Subscribers ──

@app.get("/marketing/api/subscribers")
async def marketing_get_subscribers():
    """Feliratkozók listázása."""
    return db.get_email_subscribers()

@app.get("/marketing/api/subscribers/count")
async def marketing_subscriber_count():
    """Feliratkozók száma."""
    return {"count": db.get_subscriber_count()}

@app.post("/marketing/api/subscribers")
async def marketing_add_subscriber(req: Request):
    """Új feliratkozó hozzáadása."""
    data = await req.json()
    email = data.get("email", "").strip()
    if not email:
        return JSONResponse({"error": "Email cím szükséges"}, status_code=400)
    sub = db.add_email_subscriber(
        email=email,
        name=data.get("name", ""),
        tags=data.get("tags", []),
        consent_source=data.get("consent_source", "manual")
    )
    if not sub:
        return JSONResponse({"error": "Feliratkozó hozzáadása sikertelen"}, status_code=500)
    return sub
# ── CRM Sync Import ──

@app.post("/marketing/api/subscribers/import-crm")
async def marketing_import_crm_subscribers():
    """DigiDesk ügyféladatbázisból importálja az email címmel rendelkező ügyfeleket feliratkozóként."""
    try:
        clients = db.get_clients(limit=1000)
        imported = 0
        for client in clients:
            email = client.get("email", "").strip()
            if not email or email == "-" or "@" not in email:
                continue
            name = client.get("name", "").strip()
            if name in ("Névtelen", "-", ""):
                name = email.split("@")[0]
            sub = db.add_email_subscriber(
                email=email,
                name=name,
                tags=["crm-import"],
                consent_source="import"
            )
            if sub:
                imported += 1
        logger.info(f"CRM import: {imported} subscribers imported from {len(clients)} clients")
        return {"ok": True, "imported": imported, "total_clients": len(clients)}
    except Exception as e:
        logger.error(f"CRM import error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

# ── AI Campaign Generation ──

@app.post("/marketing/api/ai/generate-campaign")
async def marketing_ai_generate(req: Request):
    """AI kampány tartalom generálás Gemini 2.5 Flash-sel.
    Input: { "instruction": "...", "campaign_type": "newsletter", "tone": "professional" }
    Output: { "subject": "...", "body": "..." }
    """
    from google import genai
    from google.genai import types

    google_key = os.getenv("GOOGLE_API_KEY")
    if not google_key:
        return JSONResponse({"error": "GOOGLE_API_KEY nincs beállítva"}, status_code=500)

    data = await req.json()
    instruction = data.get("instruction", "").strip()
    if not instruction:
        return JSONResponse({"error": "Add meg az utasítást az AI-nak"}, status_code=400)

    campaign_type = data.get("campaign_type", "newsletter")
    tone = data.get("tone", "professzionális")

    type_labels = {
        "newsletter": "hírlevél",
        "promotion": "promóciós/akciós e-mail",
        "drip": "automatizált sorozat e-mail",
        "transactional": "tranzakciós e-mail"
    }

    system_prompt = f"""Te egy professzionális e-mail marketing copywriter vagy.
A feladatod: a felhasználó utasítása alapján írj egy {type_labels.get(campaign_type, 'hírlevél')} szöveget.

SZABÁLYOK:
- A hangnem legyen: {tone}
- Magyar nyelven írj
- A válaszod KIZÁRÓLAG egyetlen valid JSON objektum legyen, minden egyéb szöveg nélkül
- Ne használj HTML tageket a body-ban, csak sima szöveget sortörésekkel
- A tárgysor legyen figyelemfelkeltő, rövid (max 60 karakter)
- A szöveg legyen célratörő, emberi hangú, ne legyen sablonos
- Ha akcióról/kedvezményről van szó, emeld ki a számokat
- Feladó cég: EAISY Marketing

JSON STRUKTÚRA:
{{
    "subject": "Az e-mail tárgysora",
    "body": "Az e-mail teljes szövege\\n\\nTöbb bekezdéssel\\n\\nSortörésekkel elválasztva"
}}"""

    try:
        client = genai.Client(api_key=google_key)
        response = await client.aio.models.generate_content(
            model="gemini-2.5-flash",
            contents=f"UTASÍTÁS: {instruction}",
            config=types.GenerateContentConfig(
                system_instruction=system_prompt,
                temperature=0.7,
                response_mime_type="application/json"
            )
        )
        ai_text = response.text.strip()

        # Clean markdown blocks
        if ai_text.startswith("```json"):
            ai_text = ai_text[7:]
        if ai_text.startswith("```"):
            ai_text = ai_text[3:]
        if ai_text.endswith("```"):
            ai_text = ai_text[:-3]
        ai_text = ai_text.strip()

        import json
        result = json.loads(ai_text)
        return {"subject": result.get("subject", ""), "body": result.get("body", "")}
    except Exception as e:
        logger.error(f"AI campaign generation error: {e}")
        return JSONResponse({"error": f"AI generálási hiba: {str(e)}"}, status_code=500)

# ── AI Content & Social Media ──

import social_media

@app.get("/marketing/api/content")
async def marketing_get_content(status: str = None):
    """AI tartalmak listázása."""
    return db.get_content_items(status_filter=status)

@app.get("/marketing/api/content/stats")
async def marketing_content_stats():
    """AI tartalom statisztikák."""
    return db.get_content_stats()

@app.post("/marketing/api/content")
async def marketing_create_content(req: Request):
    """Új AI tartalom létrehozás."""
    data = await req.json()
    item = db.create_content_item(data)
    if not item:
        return JSONResponse({"error": "Tartalom létrehozása sikertelen"}, status_code=500)
    return item

@app.put("/marketing/api/content/{item_id}")
async def marketing_update_content(item_id: str, req: Request):
    """AI tartalom frissítés."""
    data = await req.json()
    item = db.update_content_item(item_id, data)
    if not item:
        return JSONResponse({"error": "Frissítés sikertelen"}, status_code=500)
    return item

@app.delete("/marketing/api/content/{item_id}")
async def marketing_delete_content(item_id: str):
    """AI tartalom törlés."""
    ok = db.delete_content_item(item_id)
    if not ok:
        return JSONResponse({"error": "Törlés sikertelen"}, status_code=500)
    return {"success": True}

@app.post("/marketing/api/ai/generate-social")
async def marketing_ai_generate_social(req: Request):
    """AI social media poszt generálás Gemini 2.5 Flash-sel.
    Input: { "instruction": "...", "platform": "instagram", "tone": "professional" }
    Output: { "title": "...", "caption": "...", "hashtags": [...], "image_description": "..." }
    """
    from google import genai
    from google.genai import types

    google_key = os.getenv("GOOGLE_API_KEY")
    if not google_key:
        return JSONResponse({"error": "GOOGLE_API_KEY nincs beállítva"}, status_code=500)

    data = await req.json()
    instruction = data.get("instruction", "").strip()
    if not instruction:
        return JSONResponse({"error": "Add meg az utasítást az AI-nak"}, status_code=400)

    platform = data.get("platform", "instagram")
    tone = data.get("tone", "professzionális")

    system_prompt = f"""Te egy profi social media marketing szakértő vagy.
A feladatod: a felhasználó utasítása alapján írj egy {platform} posztot.

SZABÁLYOK:
- Magyar nyelven írj
- A válaszod KIZÁRÓLAG egyetlen valid JSON objektum legyen
- A caption legyen figyelemfelkeltő, emberi, emoji-kkal
- Instagram-ra max 2200 karakter caption
- Adj releváns hashtag javaslatokat (5-15 db)
- Adj kép leírást (milyen képet kellene hozzá használni)
- A hangnem legyen: {tone}
- Feladó: A cég/márka nevében

JSON STRUKTÚRA:
{{
    "title": "Rövid belső cím a tartalomnak (max 40 kar)",
    "caption": "A teljes poszt szövege emoji-kkal\\n\\nTöbb bekezdéssel",
    "hashtags": ["hashtag1", "hashtag2", "hashtag3"],
    "image_description": "Milyen képet kellene használni ehhez a poszthoz",
    "image_prompt": "Angol nyelvű, részletes képgenerálási prompt DALL-E/Midjourney számára, ami illeszkedik a poszt témájához. Legyen vizuálisan vonzó, modern, professzionális."
}}"""

    try:
        client = genai.Client(api_key=google_key)
        response = await client.aio.models.generate_content(
            model="gemini-2.5-flash",
            contents=f"UTASÍTÁS: {instruction}",
            config=types.GenerateContentConfig(
                system_instruction=system_prompt,
                temperature=0.7,
                response_mime_type="application/json"
            )
        )
        ai_text = response.text.strip()
        if ai_text.startswith("```json"): ai_text = ai_text[7:]
        if ai_text.startswith("```"): ai_text = ai_text[3:]
        if ai_text.endswith("```"): ai_text = ai_text[:-3]
        ai_text = ai_text.strip()

        import json
        result = json.loads(ai_text)
        return {
            "title": result.get("title", ""),
            "caption": result.get("caption", ""),
            "hashtags": result.get("hashtags", []),
            "image_description": result.get("image_description", ""),
            "image_prompt": result.get("image_prompt", "")
        }
    except Exception as e:
        logger.error(f"AI social generation error: {e}")
        return JSONResponse({"error": f"AI generálási hiba: {str(e)}"}, status_code=500)

# ── AI Image Generation ──

GENERATED_IMAGES_DIR = THIS_DIR / "generated_images"
GENERATED_IMAGES_DIR.mkdir(exist_ok=True)

@app.post("/marketing/api/ai/generate-image")
async def marketing_ai_generate_image(req: Request):
    """AI képgenerálás Gemini-vel.
    Input: { "prompt": "English image generation prompt..." }
    Output: { "image_url": "/generated-images/abc123.png" }
    """
    from google import genai
    from google.genai import types

    google_key = os.getenv("GOOGLE_API_KEY")
    if not google_key:
        return JSONResponse({"error": "GOOGLE_API_KEY nincs beállítva"}, status_code=500)

    data = await req.json()
    prompt = data.get("prompt", "").strip()
    if not prompt:
        return JSONResponse({"error": "Adj meg egy kép promptot!"}, status_code=400)

    try:
        client = genai.Client(api_key=google_key)
        response = await client.aio.models.generate_images(
            model='imagen-4.0-generate-001',
            prompt=prompt,
            config=types.GenerateImagesConfig(
                number_of_images=1,
            ),
        )

        # Extract image from response
        if not response.generated_images or len(response.generated_images) == 0:
            return JSONResponse({"error": "Az AI nem tudott képet generálni ehhez a prompthoz. Próbálj más megfogalmazást!"}, status_code=422)

        image_obj = response.generated_images[0].image

        # Save to file
        filename = f"{uuid.uuid4().hex[:12]}.png"
        filepath = GENERATED_IMAGES_DIR / filename
        image_obj.save(str(filepath))

        image_url = f"/generated-images/{filename}"
        logger.info(f"AI image generated: {filename}")
        return {"image_url": image_url, "filename": filename}

    except Exception as e:
        logger.error(f"AI image generation error: {e}")
        return JSONResponse({"error": f"Képgenerálási hiba: {str(e)}"}, status_code=500)

@app.get("/generated-images/{filename}")
async def serve_generated_image(filename: str):
    """Generált képek kiszolgálása."""
    import re
    if not re.match(r'^[a-zA-Z0-9_\-]+\.(png|jpg|jpeg|webp)$', filename):
        raise HTTPException(status_code=400, detail="Érvénytelen fájlnév")
    filepath = GENERATED_IMAGES_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Kép nem található")
    return FileResponse(filepath)

@app.post("/marketing/api/social/publish")
async def marketing_social_publish(req: Request):
    """Social média poszt publikálása (Instagram / Facebook).
    Input: { "content_id": "uuid", "image_url": "https://...", "caption": "...", "platform": "instagram"|"facebook" }
    """
    data = await req.json()
    content_id = data.get("content_id")
    image_url = data.get("image_url", "").strip()
    caption = data.get("caption", "").strip()
    platform = data.get("platform", "instagram")

    if not caption:
        return JSONResponse({"error": "Caption megadása kötelező"}, status_code=400)

    if platform == "instagram":
        if not image_url:
            return JSONResponse({"error": "Kép URL megadása kötelező az Instagramhoz"}, status_code=400)
        result = await social_media.publish_instagram_post(image_url, caption)
        if result.get("success"):
            if content_id:
                from datetime import datetime, timezone
                db.update_content_item(content_id, {
                    "status": "published",
                    "published_at": datetime.now(timezone.utc).isoformat(),
                    "published_platforms": ["instagram"],
                    "ig_media_id": result.get("media_id", ""),
                })
            return result
        else:
            return JSONResponse({"error": result.get("error", "Ismeretlen hiba")}, status_code=500)

    elif platform == "facebook":
        result = await social_media.publish_facebook_post(caption, image_url or None)
        if result.get("success"):
            if content_id:
                from datetime import datetime, timezone
                existing = db.get_content_item(content_id)
                platforms = ["facebook"]
                if existing and existing.get("published_platforms"):
                    platforms = list(set(existing["published_platforms"] + ["facebook"]))
                db.update_content_item(content_id, {
                    "status": "published",
                    "published_at": datetime.now(timezone.utc).isoformat(),
                    "published_platforms": platforms,
                    "fb_post_id": result.get("post_id", ""),
                })
            return result
        else:
            return JSONResponse({"error": result.get("error", "Ismeretlen hiba")}, status_code=500)

    elif platform == "all":
        # Multi-platform: IG + FB egyszerre
        from datetime import datetime, timezone
        results = {"success": True, "platforms": {}}
        all_platforms = []
        # Instagram
        if image_url:
            ig_result = await social_media.publish_instagram_post(image_url, caption)
            results["platforms"]["instagram"] = ig_result
            if ig_result.get("success"):
                all_platforms.append("instagram")
                if content_id:
                    db.update_content_item(content_id, {"ig_media_id": ig_result.get("media_id", "")})
        # Facebook
        fb_result = await social_media.publish_facebook_post(caption, image_url or None)
        results["platforms"]["facebook"] = fb_result
        if fb_result.get("success"):
            all_platforms.append("facebook")
            if content_id:
                db.update_content_item(content_id, {"fb_post_id": fb_result.get("post_id", "")})
        # Update content item
        if content_id and all_platforms:
            existing = db.get_content_item(content_id)
            existing_platforms = existing.get("published_platforms", []) if existing else []
            merged = list(set(existing_platforms + all_platforms))
            db.update_content_item(content_id, {
                "status": "published",
                "published_at": datetime.now(timezone.utc).isoformat(),
                "published_platforms": merged,
            })
        if not all_platforms:
            results["success"] = False
        return results

    else:
        return JSONResponse({"error": f"Platform '{platform}' nem támogatott"}, status_code=400)

@app.get("/marketing/api/social/instagram/media")
async def marketing_ig_media():
    """Instagram legutóbbi posztok."""
    return await social_media.get_instagram_media(limit=12)

@app.get("/marketing/api/social/instagram/quota")
async def marketing_ig_quota():
    """Instagram publikálási kvóta."""
    return await social_media.get_publishing_limit()

@app.get("/marketing/api/social/facebook/posts")
async def marketing_fb_posts():
    """Facebook legutóbbi posztok."""
    return await social_media.get_facebook_posts(limit=12)

@app.post("/marketing/api/content/{item_id}/schedule")
async def marketing_schedule_content(item_id: str, req: Request):
    """Tartalom ütemezése adott időpontra.
    Input: { "scheduled_at": "2026-06-03T09:00:00Z", "platforms": ["instagram", "facebook"] }
    """
    data = await req.json()
    scheduled_at = data.get("scheduled_at")
    platforms = data.get("platforms", ["instagram"])
    if not scheduled_at:
        return JSONResponse({"error": "scheduled_at megadása kötelező"}, status_code=400)
    item = db.update_content_item(item_id, {
        "status": "scheduled",
        "scheduled_at": scheduled_at,
        "target_platforms": platforms,
    })
    if not item:
        return JSONResponse({"error": "Ütemezés sikertelen"}, status_code=500)
    return {"success": True, "scheduled_at": scheduled_at, "platforms": platforms}

@app.get("/marketing/api/social/analytics")
async def marketing_social_analytics():
    """Social média összesített analytics."""
    return await social_media.get_social_overview()

@app.get("/thinkai-logo.png")
async def logo():
    return FileResponse(THIS_DIR / "thinkai-logo.png", media_type="image/png")

@app.get("/login-bg.jpg")
async def bg():
    return FileResponse(THIS_DIR / "login-bg.jpg", media_type="image/jpeg")

@app.get("/api/token")
async def get_token(tenant: str = ""):
    """Generate a LiveKit room token for a new user.

    FÁZIS 6 (multi-tenant): a widget a tenant slugját adhatja meg a `tenant`
    query-paraméterben — a room neve `<slug>-<uuid>` lesz, így a voice agent
    entrypoint a tenantot a room-name-ből feloldja. Visszafelé kompatibilis:
    paraméter nélkül a régi `thinkai-<uuid>` név (default tenant)."""
    api_key    = os.getenv("LIVEKIT_API_KEY")
    api_secret = os.getenv("LIVEKIT_API_SECRET")
    if not api_key or not api_secret:
        return JSONResponse({"error": "LiveKit credentials not configured"}, status_code=500)

    prefix = f"{tenant}-" if tenant else "thinkai-"
    room_name        = f"{prefix}{uuid.uuid4().hex[:8]}"
    participant_name = f"user-{uuid.uuid4().hex[:6]}"

    token = (
        AccessToken(api_key, api_secret)
        .with_identity(participant_name)
        .with_name("Visitor")
        .with_grants(VideoGrants(room_join=True, room=room_name))
        .with_room_config(
            RoomConfiguration(
                agents=[RoomAgentDispatch(agent_name="dobozos-ai")]
            )
        )
    )
    return JSONResponse({
        "token": token.to_jwt(),
        "url": os.getenv("LIVEKIT_URL"),
        "room": room_name,
    })


@app.post("/api/session/end")
async def session_end(request: Request):
    """Called by the widget on disconnect to record session duration."""
    try:
        body = await request.json()
        session_id = body.get("session_id", "")
        if session_id:
            db.close_session(session_id)
    except Exception:
        pass
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════════
# META WEBHOOK
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/webhook/meta")
async def meta_webhook_verify(request: Request):
    """Meta Webhook verification challenge."""
    verify_token = os.getenv("META_VERIFY_TOKEN", "")
    mode = request.query_params.get("hub.mode")
    token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")

    if mode and token:
        if mode == "subscribe" and token == verify_token:
            print("META_WEBHOOK_VERIFIED")
            return PlainTextResponse(content=challenge)
        raise HTTPException(status_code=403, detail="Verification token mismatch")
    raise HTTPException(status_code=400, detail="Missing parameters")


# ── Meta Data Deletion Callback (kötelező az App Verification-hez) ──────────
# A Meta akkor hívja, amikor egy felhasználó törli az app-hozzáférését a
# Facebook/Instagram/WhatsApp fiókjából. A válasznak JSON { url }-t kell
# visszaadnia, ami egy visszaigazoló link (ahol az érintett láthatja a státuszt).
# A tényleges törlést a database.delete_client_by_meta_id() végzi: az összes
# olyan ügyfelet és interakciót törli, amely a megadott Meta-azonosítóhoz
# (PSID/IGSID/WhatsApp telefonszám) kapcsolódik.


@app.get("/api/webhook/meta-data-deletion")
async def meta_data_deletion_verify(request: Request):
    """A Data Deletion Callback URL-t a Meta a normál webhook-szerkezettel is
    elérheti subscribe-kéréssel — háttér-kompat. (A hivatalos spec szerint a
    data-deletion csak POST signed_request-et használ, de ez az endpoint nem árt.)"""
    verify_token = os.getenv("META_VERIFY_TOKEN", "")
    mode = request.query_params.get("hub.mode")
    token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")
    if mode == "subscribe" and token == verify_token:
        return PlainTextResponse(content=challenge)
    raise HTTPException(status_code=403, detail="Verification token mismatch")


def _decode_meta_signed_request(signed_request: str, app_secret: str) -> dict | None:
    """Meta signed_request dekódolás (base64url + HMAC-SHA256).
    Formátum: '<signature_base64>.<payload_base64>'
    A signature-t az App Secret-tel HMAC-SHA256-cal aláírt payload-lal hasonlítjuk össze."""
    import base64
    import hmac
    import hashlib
    try:
        sig_b64, payload_b64 = signed_request.split(".", 1)
        # base64url → base64 padding
        def _pad(s):
            return s + "=" * (-len(s) % 4)
        sig = base64.urlsafe_b64decode(_pad(sig_b64))
        expected_sig = hmac.new(app_secret.encode(), payload_b64.encode(), hashlib.sha256).digest()
        if not hmac.compare_digest(sig, expected_sig):
            return None
        payload_json = base64.urlsafe_b64decode(_pad(payload_b64)).decode("utf-8")
        return json.loads(payload_json)
    except Exception:
        return None


@app.post("/api/webhook/meta-data-deletion")
async def meta_data_deletion_callback(request: Request):
    """Meta Data Deletion Request callback (kötelező az App Verification-hez).

    A Meta hivatalos specifikációja szerint: amikor egy felhasználó törli az
    app-hozzáférését a Facebook/Instagram fiókjából, a Meta egy SIGNED REQUEST-et
    POST-ol ide:
        { "signed_request": "<base64url signature>.<base64url payload>" }
    A payload (App Secret-tel HMAC-SHA256-aláírt) tartalmazza a user_id-t (PSID/IGSID)
    és az algorithm-mezőt.

    A válasznak JSON { url, confirmation_code } formátumúnak kell lennie:
      - url: publikus oldal, ahol az érintett láthatja a törlés státuszát
      - confirmation_code: alphanumeric kód, amit NEM a Meta küld, hanem MI generáljuk
        és visszaadjuk — ezt mutatja a felhasználónak a Meta a törlés visszaigazolásakor.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    app_secret = os.getenv("META_APP_SECRET", "")
    if not app_secret:
        # Teszt-üzem (App Secret nélkül, pl. App Review előtti próba): elfogadjuk
        # a nyers user_id-t a korábbi formátumban is, hogy a webhook hand-shake
        # tesztelhető maradjon. Élesben az App Secret KÖTELEZŐ az aláírás-ellenőrzéshez.
        user_id = str(body.get("user_id") or body.get("psid") or "")
        confirmation_code = body.get("confirmation_code", "")
    else:
        signed_request = body.get("signed_request", "")
        if not signed_request:
            raise HTTPException(status_code=400, detail="Missing signed_request")
        payload = _decode_meta_signed_request(signed_request, app_secret)
        if payload is None:
            raise HTTPException(status_code=400, detail="Invalid signature")
        user_id = str(payload.get("user_id") or "")
        confirmation_code = ""

    # Tényleges törlés: az összes a Meta-azonosítóhoz kapcsolódó ügyfél + interakció
    deleted_count = 0
    if user_id:
        try:
            deleted_count = db.delete_client_by_meta_id(user_id)
        except Exception as e:
            print(f"[Data Deletion] Törlési hiba (user_id={user_id}): {e}")

    # Mi generáljuk a confirmation_code-ot (alphanumeric), ha a Meta nem adta meg.
    # A Meta ezt a kódot jeleníti meg a felhasználónak a törlés visszaigazolásaként.
    if not confirmation_code:
        import secrets as _secrets
        confirmation_code = _secrets.token_hex(8).upper()

    print(f"[Data Deletion] Meta adattörlési kérelem: user_id={user_id} "
          f"confirmation={confirmation_code} törölt_sorok={deleted_count}")

    # A Meta kötelező válasza: publikus URL + confirmation_code.
    # Az URL a Privacy Policy (leírja a törlési folyamatot; a tényleges törlés azonnal megtörténik).
    return {"url": PRIVACY_POLICY_URL, "confirmation_code": confirmation_code}


async def analyze_alert_tags(message_text: str) -> list:
    """Gyors AI elemzés a bejövő üzenet címkézéséhez."""
    import os
    from google import genai
    from google.genai import types
    import json
    
    google_key = os.getenv("GOOGLE_API_KEY")
    if not google_key: return []
    
    prompt = """Elemezd a következő ügyfélüzenetet, és add vissza egy valid JSON listában az alábbi címkéket, ha relevánsak:
- "urgent": nagyon sürgős ügy
- "complaint": panasz, elégedetlenség
- "callback": telefonos visszahívást kér
- "recurring": gyakran ismétlődő probléma
Csak a címkéket tartalmazó JSON listát (pl. ["urgent", "complaint"]) add vissza, semmi mást, markdown nélkül! Ha egy sem illik, üres listát adj vissza: []."""

    try:
        client = genai.Client(api_key=google_key)
        resp = await client.aio.models.generate_content(
            model="gemini-2.5-flash",
            contents=f"Üzenet: {message_text}",
            config=types.GenerateContentConfig(system_instruction=prompt, temperature=0.1)
        )
        t = resp.text.strip()
        if t.startswith("```json"): t = t[7:-3]
        if t.startswith("```"): t = t[3:-3]
        return json.loads(t.strip())
    except Exception:
        return []

async def fetch_meta_user_profile(sender_id: str, source_channel: str, tenant_id: str = None) -> Optional[dict]:
    """Fetch the user's name and profile picture from Meta Graph API using their PSID/IGSID."""
    if source_channel not in ("Messenger", "Instagram"):
        return None

    # Instagram DM uses Page Access Token (Messenger platform), not IG API token
    # Per-tenant token (env fallback — _meta_cred a fájl későbbi részében definiált)
    token = _meta_cred(tenant_id or db.get_current_tenant(), "meta_page_token")
    if not token:
        return None
    
    # Instagram IGSID supports 'name' field directly.
    # Messenger PSID only supports 'first_name', 'last_name', 'profile_pic' — NOT 'name'.
    # Requesting 'name' on a PSID causes a 400 error, breaking the entire request.
    if source_channel == "Instagram":
        fields = "name,username,profile_pic"
    else:
        fields = "first_name,last_name,profile_pic"
        
    url = f"https://graph.facebook.com/v21.0/{sender_id}?fields={fields}&access_token={token}"
    try:
        import httpx
        async with httpx.AsyncClient() as client:
            resp = await client.get(url, timeout=5.0)
            if resp.status_code == 200:
                data = resp.json()
                name = data.get("name")
                username = data.get("username")
                if source_channel == "Instagram" and not name:
                    name = username
                if not name:
                    first = data.get("first_name", "")
                    last = data.get("last_name", "")
                    name = f"{first} {last}".strip()
                profile_pic = data.get("profile_pic")
                res_dict = {"name": name, "profile_pic": profile_pic}
                if username:
                    res_dict["username"] = username
                return res_dict if name else None
            else:
                print(f"[Meta API] Error fetching profile for {sender_id} ({source_channel}): {resp.text}")
                return None
    except Exception as e:
        print(f"[Meta API] Exception fetching profile: {e}")
        return None

# ═══════════════════════════════════════════════════════════════════════════════
# SPAM FILTER — Heuristic message spam detection (Messenger/Instagram/WhatsApp)
# ═══════════════════════════════════════════════════════════════════════════════

import re as _re_spam

_SPAM_MSG_URL_RE = _re_spam.compile(r"https?://\S+", _re_spam.IGNORECASE)

# Tipikus spam/bot szöveg minták
_SPAM_MSG_PATTERNS = [
    _re_spam.compile(p, _re_spam.IGNORECASE) for p in [
        r"\bcrypto\b", r"\bbitcoin\b", r"\bethereum\b",
        r"\bearn\s+money\b", r"\bmake\s+money\b",
        r"\b(click|tap)\s+here\b.*\bfree\b",
        r"\bcongratulations.*won\b", r"\byou\s+(have\s+)?won\b",
        r"\bfree\s+gift\b", r"\bclaim\s+(your|now)\b",
        r"\binvest\s+now\b", r"\bget\s+rich\b",
        r"\b\d+%\s*(profit|return|cashback)\b",
        r"\bcasino\b", r"\bbet(ting)?\s+online\b",
        r"\badult\s+content\b", r"\b18\+\b.*\b(click|link)\b",
        r"\bviagra\b", r"\bcialis\b",
        r"\bunsubscribe\b",
        r"\bwork\s+from\s+home\b.*\$([\d,]+)",
    ]
]


def is_spam_message(message_text: str) -> bool:
    """
    Heuristic spam detection for Messenger/Instagram/WhatsApp messages.
    Conservative — only catches obvious spam/bot messages.
    """
    if not message_text or not message_text.strip():
        return True  # Üres üzenet = nem kell feldolgozni

    text = message_text.strip()

    # 1. Link-only message (no real text, just URL)
    urls = _SPAM_MSG_URL_RE.findall(text)
    text_without_urls = _SPAM_MSG_URL_RE.sub("", text).strip()
    if urls and len(text_without_urls) < 5:
        # Csak URL(ek), szinte semmi szöveg — spam
        return True

    # 2. Excessive links (5+ in a single message)
    if len(urls) >= 5:
        return True

    # 3. Known spam/bot patterns
    spam_hits = sum(1 for rx in _SPAM_MSG_PATTERNS if rx.search(text))
    if spam_hits >= 2:
        return True

    # 4. Excessive length with many links (likely copypasted spam)
    if len(text) > 1000 and len(urls) >= 3:
        return True

    return False


# ═══════════════════════════════════════════════════════════════════════════════
# META MULTI-TENANT — tenant feloldás + per-tenant credentialök (FÁZIS 5)
# ═══════════════════════════════════════════════════════════════════════════════
# Egy közös ThinkAI Meta app szolgál ki minden tenantot; a webhook payload
# Meta-azonosítóiból (page_id / ig_user_id / phone_number_id) oldjuk fel, hogy
# melyik tenanthoz tartozik az üzenet. A leképezést a tenant_credentials tábla
# tárolja ('meta_page_id', 'instagram_user_id', 'whatsapp_phone_id' kulcsok).
# Ha nincs találat (pl. Rivergate — nincs credential sor), None-t adunk, és
# minden a régi módon, az env tokenekkel + default tenanttal működik.

import time as _time

# meta_id → tenant_id cache (TTL ~60 mp, hogy ne legyen N+1 lekérdezés üzenetenként)
_META_TENANT_CACHE_TTL = 60
_meta_tenant_cache: dict = {"ts": 0.0, "map": {}}


def _rebuild_meta_tenant_cache():
    """Újraépíti a meta_id → tenant_id leképezést az aktív tenantok
    tenant_credentials soraiból (a tárolt ÉRTÉKRE illesztünk, nem a kulcsra)."""
    mapping = {}
    try:
        for tenant in db.get_active_tenants():
            tid = tenant.get("id")
            if not tid:
                continue
            for key in ("meta_page_id", "instagram_user_id", "whatsapp_phone_id"):
                val = db.get_credential(tid, key, default=None)
                if val:
                    mapping[str(val)] = tid
    except Exception as e:
        print(f"[Meta tenant] Cache rebuild hiba: {e}")
    _meta_tenant_cache["map"] = mapping
    _meta_tenant_cache["ts"] = _time.monotonic()


def resolve_meta_tenant(page_id: str = "", ig_user_id: str = "", phone_number_id: str = "") -> str | None:
    """Megkeresi, melyik tenant birtokolja a megadott Meta page ID-t /
    Instagram user ID-t / WhatsApp phone_number_id-t. Visszatér a tenant_id-val
    (uuid string) vagy None-nal (ilyenkor a hívó a default tenantot használja —
    visszafelé kompatibilis, egycéges működés)."""
    if _time.monotonic() - _meta_tenant_cache["ts"] > _META_TENANT_CACHE_TTL:
        _rebuild_meta_tenant_cache()
    mapping = _meta_tenant_cache["map"]
    for meta_id in (page_id, ig_user_id, phone_number_id):
        if meta_id and str(meta_id) in mapping:
            return mapping[str(meta_id)]
    return None


def _meta_cred(tenant_id: str | None, key: str) -> str:
    """Per-tenant Meta credential feloldása env fallbackkel.
    Ha a tenantnak nincs credential sora, pontosan a régi env értéket adja —
    így a meglévő egycéges (Rivergate) működés változatlan marad."""
    if key == "meta_page_token":
        fallback = os.getenv("META_PAGE_ACCESS_TOKEN", "")
    elif key == "whatsapp_token":
        # A tenant saját page tokenje is jó fallback (mint az env-ben a META_PAGE_ACCESS_TOKEN)
        fallback = db.get_credential(
            tenant_id, "meta_page_token",
            default=os.getenv("WHATSAPP_TOKEN", os.getenv("META_PAGE_ACCESS_TOKEN", "")),
        )
    elif key == "whatsapp_phone_id":
        fallback = os.getenv("WHATSAPP_PHONE_ID", "")
    elif key == "instagram_token":
        fallback = os.getenv("META_INSTAGRAM_TOKEN", "")
    elif key == "instagram_user_id":
        fallback = os.getenv("META_INSTAGRAM_USER_ID", "26530155976686869")
    else:
        fallback = ""
    return db.get_credential(tenant_id, key, default=fallback) or ""


async def _send_channel_message(source_channel: str, sender_id: str, text: str, phone_number_id: str = None, tenant_id: str = None) -> bool:
    """
    EAISY-241 §1.1.2 — Meta csatornára (Messenger/Instagram/WhatsApp) küldő helper.
    Az autonóm válasz (restriction == 'none') azonnali kiküldéséhez használjuk.
    Visszatér True-val ha sikeres, False ha hiba.
    A logika az approve_approval_api-ból van extractálva (web_server.py:~3110).
    A tenant_id alapján per-tenant tokeneket használ (env fallbackkel).
    """
    import httpx
    # Tenant: paraméter, különben a hívó által beállított contextvar
    tid = tenant_id or db.get_current_tenant()
    ch = (source_channel or "").lower()
    try:
        async with httpx.AsyncClient() as http_client:
            if ch == "whatsapp":
                wa_token = _meta_cred(tid, "whatsapp_token")
                wa_phone_id = phone_number_id or _meta_cred(tid, "whatsapp_phone_id")
                if not wa_token or not wa_phone_id:
                    return False
                resp = await http_client.post(
                    f"https://graph.facebook.com/v21.0/{wa_phone_id}/messages",
                    headers={"Authorization": f"Bearer {wa_token}"},
                    json={"messaging_product": "whatsapp", "to": sender_id, "type": "text", "text": {"body": text}},
                    timeout=20,
                )
                resp.raise_for_status()
                return True
            elif ch == "instagram":
                ig_token = _meta_cred(tid, "instagram_token")
                ig_user_id = _meta_cred(tid, "instagram_user_id")
                if not ig_token:
                    return False
                resp = await http_client.post(
                    f"https://graph.instagram.com/v21.0/{ig_user_id}/messages",
                    headers={"Authorization": f"Bearer {ig_token}", "Content-Type": "application/json"},
                    json={"recipient": {"id": sender_id}, "message": {"text": text}},
                    timeout=20,
                )
                resp.raise_for_status()
                return True
            elif ch == "messenger":
                page_access_token = _meta_cred(tid, "meta_page_token")
                if not page_access_token:
                    return False
                resp = await http_client.post(
                    "https://graph.facebook.com/v21.0/me/messages",
                    headers={"Authorization": f"Bearer {page_access_token}", "Content-Type": "application/json"},
                    json={"recipient": {"id": sender_id}, "message": {"text": text}},
                    timeout=20,
                )
                resp.raise_for_status()
                return True
            # email: a Brevo küldés külön logika (az autonóm email-válasz ritka;
            # biztonság kedvéért itt nem küldünk auto-emailt, csak Messenger/IG/WA).
            return False
    except Exception as e:
        print(f"[Auto-send] {source_channel} küldési hiba: {e}")
        return False


async def process_meta_message(sender_id: str, message_text: str, source_channel: str = "Messenger", phone_number_id: str = None, sender_name: str = None, tenant_id: str = None):
    """Aszinkron háttérfeladat a Meta Messenger / Instagram / WhatsApp üzenetek feldolgozására."""
    import asyncio
    import json
    from datetime import datetime, timedelta
    from google import genai
    from google.genai import types
    from prompt_utils import get_system_prompt
    import database as db
    import email_processor

    # ── Multi-tenant: a webhook által feloldott tenant kontextus beállítása ──
    # (az asyncio.create_task amúgy is örökli a contextvart, de így biztos,
    #  ha más hívó közvetlenül indítaná). Nincs tenant → default (Rivergate).
    if tenant_id:
        db.set_current_tenant(tenant_id)

    # ── SPAM CHECK — before any AI call ──────────────────────────────
    if is_spam_message(message_text):
        print(f"[SPAM] {source_channel} spam szűrve (silent drop): sender={sender_id}, msg={message_text[:100]}")
        db.log_interaction(
            type=source_channel.lower(),
            topic=f"[SPAM] {source_channel} üzenet: {message_text[:200]}",
            summary=f"Spam üzenet automatikusan szűrve ({source_channel})",
            result="Automatikusan szűrve",
            tool_name="spam_filter",
            session_id=f"spam_{source_channel.lower()}_{sender_id}",
            funnel_stage="spam",
            approval_status="spam"
        )
        return

    alert_tags_task = asyncio.create_task(analyze_alert_tags(message_text))

    try:
        # 1. Beolvassuk a rendszer promptot
        system_prompt = get_system_prompt(channel=source_channel)
        today = datetime.now().strftime("%Y-%m-%d (%A)")
        
        # Először is elmentjük a bejövő üzenetet a Kanbanba
        client_data = {"messenger_id": sender_id, "forras_csatorna": source_channel}
        meta_name = None

        # ── WhatsApp: a nevet a webhook contacts tömbből kapjuk (sender_name param) ──
        if source_channel == "WhatsApp":
            # WhatsApp sender_id = telefonszám → mentsük phone-ként is
            client_data["phone"] = sender_id
            if sender_name:
                meta_name = sender_name
                client_data["name"] = sender_name
                print(f"[WhatsApp] Feladó neve a webhook contacts-ból: {sender_name}")
            else:
                print(f"[WhatsApp] Feladó neve nem elérhető a webhook-ból, sender_id: {sender_id}")
        else:
            # ── Messenger / Instagram: profil lekérdezés a Graph API-ból (változatlan) ──
            meta_profile = await fetch_meta_user_profile(sender_id, source_channel, tenant_id=tenant_id)
            if meta_profile:
                if source_channel == "Instagram":
                    meta_name = meta_profile.get("username") or meta_profile.get("name")
                else:
                    meta_name = meta_profile.get("name")
                if meta_name:
                    client_data["name"] = meta_name
                profile_pic = meta_profile.get("profile_pic")
                if profile_pic:
                    client_data["profile_pic_url"] = profile_pic
                username = meta_profile.get("username")
                if username:
                    client_data["instagram_username"] = username

        if not meta_name:
            # Fallback: keressük az adatbázisban a nevet
            print(f"[Meta API] Név feloldás sikertelen ({source_channel} {sender_id}), DB fallback...")
            existing = db.find_client_by_contact(messenger_id=sender_id)
            if existing:
                cd_fb = existing.get("custom_data")
                if isinstance(cd_fb, str):
                    try: cd_fb = json.loads(cd_fb)
                    except: cd_fb = {}
                if isinstance(cd_fb, dict):
                    db_name = cd_fb.get("nev") or cd_fb.get("name") or cd_fb.get("instagram_username") or existing.get("name")
                    if db_name and db_name not in ("Névtelen", "-", ""):
                        meta_name = db_name
                        print(f"[Meta API] DB fallback név: {meta_name}")
            
        db.upsert_client(client_data, additional_log=f"Ügyfél ({source_channel}): {message_text}")

        # Előzmények beolvasása
        client_record = db.find_client_by_contact(messenger_id=sender_id)
        if client_record:
            try:
                cd = client_record.get("custom_data")
                if isinstance(cd, str):
                    c_data = json.loads(cd or "{}")
                elif isinstance(cd, dict):
                    c_data = cd
                else:
                    c_data = {}
                
                chat_history = c_data.get("beszelgetes_naplo", "")
                if chat_history:
                    if len(chat_history) > 3000:
                        chat_history = "... " + chat_history[-3000:]
                    system_prompt += f"\n\n--- Eddigi beszélgetés előzménye a felhasználóval ---\n{chat_history}\n----------------------------------------------------"
            except Exception as e:
                print(f"[Meta AI Process] Hiba a napló beolvasásakor: {e}")

        # Szabályok
        triage_rules = db.get_triage_rules()
        if triage_rules:
            rules_text = "\n".join([f"- Szabály ID: {r['id']}, Helyzet: {r['situation']}, Prioritás: {r['priority']}" for r in triage_rules])
            system_prompt += f"\n\n--- TRIÁZS SZABÁLYOK ---\nKérlek értékeld a páciens problémáját az alábbi szabályok alapján is. Ha egyezik egy 'Sürgős' prioritású szabállyal, KÖTELEZŐ felvenned az 'urgent' tag-et az alert_tags listába!\n{rules_text}\n----------------------------------------------------"

        # (A telephelyeket a prompt_utils.py már beletette a system_prompt-ba!)

        # JSON UTASÍTÁS
        json_instruction = f"""
FONTOS INSTRUKCIÓ: A mai dátum: {today}. Minden dátumot ehhez a dátumhoz viszonyíts!
TE FELADATOD:
Értékeld a beérkezett üzenetet és a beszélgetés előzményeit. Formázz röviden, mint egy Messenger üzenetet. Válaszolj közvetlenül az ügyfélnek.
A kimeneted KIZÁRÓLAG egyetlen valid JSON objektum legyen, minden további markdown formázás (pl. ```json) NÉLKÜL.

JSON STRUKTÚRA:
{{
    "reply_text": "A válaszüzenet szövege. Ez fog kimenni a Messengerre.",
    "summary": "Tömör, 1 mondatos összefoglaló magyarul az ügyfél panaszáról, kérdéséről vagy kéréséről (összegzés).",
    "kanban_data": {{
        "name": "Ügyfél neve (ha megadta vagy tudod)",
        "email": "Ügyfél e-mailje (ha megadta)",
        "phone": "Telefonszám (ha megadta)",
        "clinic_id": 0, // A telephely ID-ja, ha kiválasztotta, különben 0
        "priority": "Normál" // vagy 'Sürgős', 'Kiemelt' stb. a triázs alapján
    }},
    "meeting": {{
        "title": "Találkozó címe (ha VÉGLEGESÍTVE időpontot foglal, különben null)",
        "date": "YYYY-MM-DD",
        "time": "HH:MM",
        "duration_minutes": 30
    }},
    "action_modify_meeting": {{
        "event_title_to_modify": "A módosítandó esemény címe vagy része (csak ha kéri)",
        "new_date": "YYYY-MM-DD",
        "new_time": "HH:MM"
    }},
    "action_delete_meeting": {{
        "event_title_to_delete": "A törlendő esemény címe vagy része (csak ha lemondja)"
    }},
    "alert_tags": ["urgent", "callback", "kiemelt"] // Válaszd ki, ha releváns, különben üres lista []
}}
FIGYELEM: Ha az eset Sürgős vagy Kiemelt prioritású, VAGY a kérés szerepel a Kivételek (Exceptions) listájában, a "meeting" értéke KÖTELEZŐEN null kell legyen (SZIGORÚAN TILOS időpontot foglalni!).
KIVÉTEL A TILTÁS ALÓL: Ha az ügyfél egyértelműen időpontot kér, de NEM adja meg a panaszát, AKKOR IS FOGLALD LE az időpontot!
"""
        system_prompt += f"\n\n--- JSON UTASÍTÁS ---\n{json_instruction}"

        # 3. Gemini hívás
        client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))
        user_content = f"Ügyfél neve: {meta_name if meta_name else 'Ismeretlen'}\nÚj üzenet: {message_text}"

        try:
            response = await client.aio.models.generate_content(
                model="gemini-2.5-flash",
                contents=user_content,
                config=types.GenerateContentConfig(
                    system_instruction=system_prompt,
                    temperature=0.2,
                    response_mime_type="application/json"
                )
            )
            ai_text = response.text.strip()
        except Exception as e:
            print(f"[Meta AI Process] Kritikus Gemini API Hiba: {e}")
            db.upsert_client({"messenger_id": sender_id}, additional_log=f"[Rendszer] Gemini API hiba: {e}")
            return

        if ai_text.startswith("```json"): ai_text = ai_text[7:]
        if ai_text.startswith("```"): ai_text = ai_text[3:]
        if ai_text.endswith("```"): ai_text = ai_text[:-3]
        ai_text = ai_text.strip()

        try:
            data = json.loads(ai_text)
            print("AI JSON Output:", json.dumps(data, indent=2))
        except json.JSONDecodeError as e:
            print(f"[Meta AI Process] Hibás JSON válasz: {e}")
            db.upsert_client({"messenger_id": sender_id}, additional_log=f"[Rendszer] Hibás JSON válasz az AI-tól: {e}")
            return

        final_text = data.get("reply_text", "")
        summary = data.get("summary", "")
        kanban = data.get("kanban_data") or {}
        meeting = data.get("meeting")
        modify_action = data.get("action_modify_meeting")
        delete_action = data.get("action_delete_meeting")
        alert_tags = data.get("alert_tags", [])
        
        booked_meeting = False
        chosen_clinic_id = None
        client_id = None

        # --- ACTION: KANBAN ADATOK MENTÉSE ---
        if kanban:
            custom_data = {
                "messenger_id": sender_id,
                "forras_csatorna": source_channel
            }
            if summary:
                custom_data["problem_description"] = summary
            kanban_name = kanban.get("name", "").strip()
            # Don't accept placeholder names from AI — prefer the real name from Meta API
            if kanban_name and kanban_name not in ("Ismeretlen", "Névtelen", "-", "ismeretlen", "névtelen"):
                custom_data["name"] = kanban_name
            elif meta_name:
                custom_data["name"] = meta_name
            if kanban.get("email"): custom_data["email"] = kanban["email"]
            if kanban.get("phone"): custom_data["phone"] = kanban["phone"]
            if kanban.get("clinic_id"):
                try:
                    custom_data["clinic_id"] = int(kanban["clinic_id"])
                    chosen_clinic_id = int(kanban["clinic_id"])
                except:
                    pass
                    
            if kanban.get("priority"):
                custom_data["prioritas"] = kanban["priority"]
            if "urgent" in alert_tags and custom_data.get("prioritas") != "Sürgős":
                custom_data["prioritas"] = "Sürgős"
            elif "kiemelt" in alert_tags and custom_data.get("prioritas") != "Kiemelt":
                custom_data["prioritas"] = "Kiemelt"

            custom_data = {k: v for k, v in custom_data.items() if v}
            
            # Fetch existing client to keep current status, or default to "uj"
            existing_client = db.find_client_by_contact(messenger_id=sender_id)
            current_status = existing_client.get("status", "uj") if existing_client else "uj"
            
            # --- AUTOMATIKUS TAG HOZZÁRENDELÉS ---
            # Az üzenet és AI válasz alapján releváns tag-eket rendelünk az ügyfélhez
            existing_tags = []
            if existing_client:
                ec_data = existing_client.get("custom_data", {}) or {}
                if isinstance(ec_data, str):
                    try: ec_data = json.loads(ec_data)
                    except: ec_data = {}
                existing_tags = ec_data.get("tags", []) if isinstance(ec_data, dict) else []
            
            msg_lower = message_text.lower() if message_text else ""
            reply_lower = final_text.lower() if final_text else ""
            combined_text = msg_lower + " " + reply_lower
            
            new_tags = list(existing_tags)  # copy
            
            # Árkérdés: csak ha tényleg árakról van szó (szóhatárokkal!)
            # "ár" túl rövid → "korábban", "járt", "már" hamis pozitív

            ar_pattern = r'\b(árak|árat|árazás|árlista|mennyibe|költség|fizetés|kedvezmény|részletfizetés|bruttó|nettó|mennyi.*kerül|ár\b)'
            if re.search(ar_pattern, combined_text):
                if "árkérdés" not in new_tags:
                    new_tags.append("árkérdés")
            
            # Ajánlatkérés: ha ajánlatot, árajánlatot kér
            if any(kw in combined_text for kw in ["ajánlat", "árajánlat", "kérek ajánlatot", "ajánlatot kér"]):
                if "ajánlatkérés" not in new_tags:
                    new_tags.append("ajánlatkérés")
            
            # Kampány lead: ha kampányból érkezett vagy marketing tartalom
            if any(kw in combined_text for kw in ["kampány", "promóció", "hírlevél", "newsletter"]):
                if "kampány lead" not in new_tags:
                    new_tags.append("kampány lead")
            
            if new_tags != existing_tags:
                custom_data["tags"] = new_tags
            
            client_id = db.upsert_client(custom_data, status=current_status)
            
            # Kiemelt eszkaláció
            priority = custom_data.get("prioritas", "Normál")
            if priority == "Kiemelt" or "kiemelt" in alert_tags:
                email_to_send = None
                t_rules = db.get_triage_rules()
                for r in t_rules:
                    if r.get("priority") == "Kiemelt" and r.get("escalation_email"):
                        email_to_send = r["escalation_email"]
                        break
                if email_to_send:
                    name_val = kanban.get("name") or meta_name or "Ismeretlen"
                    contact_val = f"Email: {kanban.get('email', '-')} | Telefon: {kanban.get('phone', '-')}"
                    asyncio.create_task(email_processor.send_escalation_email_to_staff(
                        to_email=email_to_send,
                        patient_name=name_val,
                        patient_contact=contact_val,
                        problem_description=message_text,
                        priority="Kiemelt"
                    ))

        # --- ACTION: NAPTÁR FOGLALÁS ---
        if meeting and meeting.get("title") and meeting.get("date") and meeting.get("time"):
            start_dt_val = f"{meeting['date']}T{meeting['time']}:00"
            dur = int(meeting.get("duration_minutes", 30))
            try:
                import zoneinfo
                tz = zoneinfo.ZoneInfo("Europe/Budapest")
                start_dt_obj = datetime.fromisoformat(start_dt_val).replace(tzinfo=tz)
                start_dt_val = start_dt_obj.isoformat()
                end_dt_val = (start_dt_obj + timedelta(minutes=dur)).isoformat()
            except:
                end_dt_val = None
                
            existing = db.get_calendar_events()
            if any(ev.get("start_dt") == start_dt_val for ev in existing):
                db.upsert_client({"messenger_id": sender_id}, additional_log="[Rendszer] Figyelmeztetés: Ebbe az időpontba már van foglalás, nem rögzítve.")
            else:
                created_event_id = db.add_calendar_event(
                    title=meeting.get("title", "Konzultáció"),
                    start_dt=start_dt_val,
                    end_dt=end_dt_val,
                    duration_minutes=dur,
                    attendee=kanban.get("name") or meta_name or "Ismeretlen Ügyfél",
                    attendee_email=kanban.get("email", "-")
                )
                
                # Visszaállítjuk a státuszt "uj"-ra, hogy kikerüljön a "lemondott" oszlopból
                client_to_reset = db.find_client_by_contact(messenger_id=sender_id)
                if client_to_reset:
                    c_data = client_to_reset.get("custom_data", {})
                    if "cancelled_viewed" in c_data:
                        del c_data["cancelled_viewed"]
                    db.edit_client_details(client_to_reset["id"], c_data)
                    db.update_client_status(client_to_reset["id"], "uj")
                    
                db.upsert_client({"messenger_id": sender_id}, additional_log=f"[Rendszer] Naptár bejegyzés létrehozva: {start_dt_val}")
                booked_meeting = True
                
                attendee_email = kanban.get("email")
                if attendee_email and created_event_id:
                    asyncio.create_task(
                        email_processor.send_booking_confirmation_email(
                            event_id=created_event_id,
                            title=meeting.get("title", "Konzultáció"),
                            date=meeting.get("date"),
                            time=meeting.get("time"),
                            attendee=kanban.get("name") or meta_name or "Ismeretlen Ügyfél",
                            attendee_email=attendee_email
                        )
                    )

        # --- ACTION: NAPTÁR MÓDOSÍTÁS ---
        if modify_action and modify_action.get("event_title_to_modify"):
            ev_title = modify_action["event_title_to_modify"]
            found = db.find_calendar_event_by_title(ev_title)
            if found:
                updates = {}
                old_dt = datetime.fromisoformat(found["start_dt"].replace("Z", "+00:00"))
                d = modify_action.get("new_date") or old_dt.strftime("%Y-%m-%d")
                t = modify_action.get("new_time") or old_dt.strftime("%H:%M")
                try:
                    import zoneinfo
                    tz = zoneinfo.ZoneInfo("Europe/Budapest")
                    new_start = datetime.fromisoformat(f"{d}T{t}:00").replace(tzinfo=tz)
                except:
                    new_start = datetime.fromisoformat(f"{d}T{t}:00")
                dur = found.get("duration_minutes", 30)
                updates["start_dt"] = new_start.isoformat()
                updates["end_dt"] = (new_start + timedelta(minutes=dur)).isoformat()
                db.update_calendar_event(found["id"], **updates)
                db.upsert_client({"messenger_id": sender_id}, additional_log=f"[Rendszer] Naptár bejegyzés módosítva: {found['title']}")
                # Módosítás visszaigazolás email küldése
                attendee_email = found.get("attendee_email")
                if attendee_email and attendee_email != "-":
                    asyncio.create_task(
                        email_processor.send_modification_confirmation_email(
                            attendee=found.get("attendee", "Ügyfél"),
                            attendee_email=attendee_email,
                            title=found.get("title", "Konzultáció"),
                            old_datetime=found["start_dt"],
                            new_datetime=updates.get("start_dt", found["start_dt"])
                        )
                    )
            else:
                db.upsert_client({"messenger_id": sender_id}, additional_log=f"[Rendszer] Módosítás sikertelen, nem található: {ev_title}")

        # --- ACTION: NAPTÁR TÖRLÉS ---
        if delete_action:
            found = None
            client_to_cancel = db.find_client_by_contact(messenger_id=sender_id)
            if client_to_cancel:
                c_email = client_to_cancel.get("custom_data", {}).get("email")
                c_name = client_to_cancel.get("name")
                if c_email and c_email != "-":
                    found = db.find_upcoming_event_by_attendee(email=c_email)
                if not found and c_name and c_name != "-":
                    found = db.find_upcoming_event_by_attendee(name=c_name)
                    
            if found:
                db.delete_calendar_event(found["id"])
                if client_to_cancel:
                    c_data = client_to_cancel.get("custom_data", {})
                    if isinstance(c_data, str):
                        try: c_data = json.loads(c_data)
                        except: c_data = {}
                    if not isinstance(c_data, dict): c_data = {}
                    
                    c_data["cancelled_viewed"] = False
                    # Automatikus 'törölt időpont' tag hozzáadása
                    existing_tags = c_data.get("tags", [])
                    if "törölt időpont" not in existing_tags:
                        existing_tags.append("törölt időpont")
                        c_data["tags"] = existing_tags
                    db.edit_client_details(client_to_cancel["id"], c_data)
                    db.update_client_status(client_to_cancel["id"], "lemondott")
                
                db.upsert_client({"messenger_id": sender_id}, additional_log=f"[Rendszer] Naptár bejegyzés törölve: {found['title']}")
            else:
                db.upsert_client({"messenger_id": sender_id}, additional_log="[Rendszer] Törlés sikertelen, nem található az ügyfélhez tartozó esemény a naptárban.")

        # 4. Válasz rögzítése a Kanbanba
        if final_text:
            existing_client = db.find_client_by_contact(messenger_id=sender_id)
            # Név feloldás: meta API → DB → sender_id
            display_name = meta_name
            if not display_name and existing_client:
                cd_for_name = existing_client.get("custom_data")
                if isinstance(cd_for_name, str):
                    try: cd_for_name = json.loads(cd_for_name)
                    except: cd_for_name = {}
                if isinstance(cd_for_name, dict):
                    display_name = cd_for_name.get("nev") or cd_for_name.get("name") or cd_for_name.get("instagram_username") or existing_client.get("name")
            if not display_name or display_name in ("Névtelen", "Ismeretlen", "-", ""):
                display_name = None

            current_status = existing_client.get("status", "uj") if existing_client else "uj"
            db.upsert_client({"messenger_id": sender_id, "forras_csatorna": source_channel}, additional_log=f"AI Válasz: {final_text}", status=current_status)
            
            f_stage = "foglalt" if booked_meeting else "valaszolt"
            
            # Piszkozat készítése
            draft_payload = {
                "channel": source_channel,
                "sender_id": sender_id,
                "to_name": display_name if display_name else sender_id,
                "phone_number_id": phone_number_id,
                "body": final_text
            }
            draft_json = json.dumps(draft_payload)
            
            session_id = f"{source_channel.lower()}_{sender_id}"
            db.create_session(session_id=session_id, room_name=f"{source_channel} Chat", participant=display_name if display_name else "Ismeretlen")
            
            # alert tags beolvasása az aszinkron feladatból, ha az AI nem adott
            tags_from_ai = alert_tags if isinstance(alert_tags, list) else []
            try:
                tags_from_task = await alert_tags_task
                if not tags_from_task: tags_from_task = []
            except:
                tags_from_task = []
                
            combined_tags = list(set(tags_from_ai + tags_from_task))
            
            # ── KLASSZIFIKÁCIÓ ──
            # Ha az AI visszahívást/sürgőset jelzett (alert_tags), azt
            # handover_reason-ként adjuk át — ne lehessen autonóm válasz emberi ügyre.
            meta_handover = ""
            if "urgent" in combined_tags:
                meta_handover = "sürgős jelzés (alert_tags)"
            elif "callback" in combined_tags or "complaint" in combined_tags:
                meta_handover = "visszahívás/panasz jelzés (alert_tags)"
            classification = await classify_interaction(
                message_text=message_text,
                channel=source_channel.lower(),
                tool_calls=["book_meeting"] if booked_meeting else [],
                handover_reason=meta_handover
            )

            # EAISY-241 §1.1.2 — Ha az ügytípus eljárása „Önállóan kezelhető"
            # (restriction == 'none' / autonomous), a válasz AZONNAL kiküldésre
            # kerül, nem pedig pending draft-ként. Egyébként pending (jóváhagyásra vár).
            is_autonomous = bool(classification.get("autonomous")) and classification.get("restriction") == "none"
            send_ok = False
            if is_autonomous:
                try:
                    send_ok = await _send_channel_message(
                        source_channel, sender_id, final_text,
                        phone_number_id=phone_number_id,
                        tenant_id=tenant_id,
                    )
                except Exception as send_err:
                    print(f"[Meta AI Process] Auto-send hiba ({source_channel}): {send_err}")
                    send_ok = False

            approval_status = "approved" if (is_autonomous and send_ok) else "pending"
            funnel_stage_final = "valaszolt" if (is_autonomous and send_ok) else f_stage

            # Logolás az interactions táblába + approval
            db.log_interaction(
                type=source_channel.lower(),
                topic=f"{source_channel} AI válasz - {message_text[:200]}",
                summary=classification.get("osszefoglalas") or summary or message_text[:200],
                result=classification.get("eredmeny", "Várakozik jóváhagyásra"),
                tool_name="process_meta_message",
                session_id=session_id,
                direction="inbound",
                funnel_stage=funnel_stage_final,
                alert_tags=combined_tags,
                handover_reason=meta_handover or None,
                approval_status=approval_status,
                ai_draft_response=draft_json,
                clinic_id=str(chosen_clinic_id) if chosen_clinic_id else None,
                client_id=client_id if client_id else None,
                classification=classification
            )

    except Exception as e:
        import traceback
        print(f"[Meta AI Process] Hiba: {e}")
        traceback.print_exc()


@app.post("/api/webhook/meta")
async def meta_webhook_receive(request: Request):
    """Receive messages from Meta Messenger, Instagram, and WhatsApp."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    
    obj_type = body.get("object")

    # ── Multi-tenant routing (FÁZIS 5): a payload Meta-azonosítóiból tenant feloldás ──
    # Messenger/IG: entry.id a page/IG id; WhatsApp: metadata.phone_number_id.
    # Nincs találat → None → a contextvar a default tenantra (Rivergate) esik vissza.
    payload_tenant = None
    for _entry in body.get("entry", []):
        payload_tenant = resolve_meta_tenant(page_id=_entry.get("id", ""), ig_user_id=_entry.get("id", ""))
        if payload_tenant:
            break
        for _change in _entry.get("changes", []):
            _pnid = (_change.get("value", {}).get("metadata") or {}).get("phone_number_id", "")
            payload_tenant = resolve_meta_tenant(phone_number_id=_pnid)
            if payload_tenant:
                break
        if payload_tenant:
            break
    if payload_tenant:
        db.set_current_tenant(payload_tenant)
        print(f"[Meta Webhook] Tenant feloldva: {payload_tenant}")

    # 1) Messenger / Instagram
    if obj_type in ("page", "instagram"):
        is_instagram = (obj_type == "instagram")

        for entry in body.get("entry", []):
            # Get our own page/IG ID to detect echoes
            page_id = entry.get("id", "")

            # Entry-szintű tenant finomítás (egy payloadban több page is lehet)
            entry_tenant = resolve_meta_tenant(page_id=page_id, ig_user_id=page_id) or payload_tenant
            if entry_tenant:
                db.set_current_tenant(entry_tenant)

            for webhook_event in entry.get("messaging", []):
                sender_id = webhook_event.get("sender", {}).get("id")

                # ── Echo detection: skip messages sent BY our page/bot ──
                message_obj = webhook_event.get("message", {})
                if message_obj.get("is_echo"):
                    print(f"[Meta Webhook] Echo üzenet kiszűrve (sender: {sender_id})")
                    continue
                # Also skip if sender is our own page ID or IG business account
                ig_user_id = _meta_cred(entry_tenant, "instagram_user_id")
                if sender_id and (sender_id == page_id or sender_id == ig_user_id):
                    print(f"[Meta Webhook] Saját üzenet kiszűrve (sender: {sender_id})")
                    continue

                if "message" in webhook_event and "text" in message_obj:
                    message_text = message_obj["text"]
                    source_channel = "Instagram" if is_instagram else "Messenger"
                    print(f"[Meta Webhook] Új üzenet feladótól (ID: {sender_id}, Csatorna: {source_channel}): {message_text}")

                    # AI aszinkron feldolgozás (a create_task örökli a tenant contextvart)
                    task = asyncio.create_task(process_meta_message(sender_id, message_text, source_channel, tenant_id=entry_tenant))
                    background_tasks.add(task)
                    task.add_done_callback(background_tasks.discard)

        return PlainTextResponse(content="EVENT_RECEIVED", status_code=200)

    # 2) WhatsApp
    elif obj_type == "whatsapp_business_account":
        for entry in body.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value", {})
                
                # Check if it's a message event
                if "messages" in value:
                    phone_number_id = value.get("metadata", {}).get("phone_number_id")

                    # ── Multi-tenant: tenant feloldás a phone_number_id alapján ──
                    wa_tenant = resolve_meta_tenant(phone_number_id=phone_number_id) or payload_tenant
                    if wa_tenant:
                        db.set_current_tenant(wa_tenant)

                    # ── Extract sender name from contacts array ──
                    # WhatsApp webhook provides contacts[].profile.name
                    wa_contacts = value.get("contacts", [])
                    wa_name_map = {}
                    for contact in wa_contacts:
                        wa_id = contact.get("wa_id", "")
                        profile_name = (contact.get("profile") or {}).get("name", "")
                        if wa_id and profile_name:
                            wa_name_map[wa_id] = profile_name
                    
                    for message in value.get("messages", []):
                        sender_id = message.get("from")
                        sender_name = wa_name_map.get(sender_id)  # Resolve name from contacts
                        
                        if message.get("type") == "text" and "text" in message:
                            message_text = message["text"].get("body", "")
                            print(f"[Meta Webhook] Új üzenet feladótól (ID: {sender_id}, Név: {sender_name or 'N/A'}, Csatorna: WhatsApp): {message_text}")
                            
                            task = asyncio.create_task(process_meta_message(sender_id, message_text, "WhatsApp", phone_number_id, sender_name=sender_name, tenant_id=wa_tenant))
                            background_tasks.add(task)
                            task.add_done_callback(background_tasks.discard)
                            
        return PlainTextResponse(content="EVENT_RECEIVED", status_code=200)
    
    raise HTTPException(status_code=404, detail="Not Found")


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN AUTH
# ═══════════════════════════════════════════════════════════════════════════════

class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/admin/login")
def admin_login(req: LoginRequest):
    """Admin login — returns JWT token + role + tenant."""
    user = db.verify_admin_user(req.username, req.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Hibás felhasználónév vagy jelszó"
        )
    tenant_id = user.get("tenant_id") or ""
    role = user.get("role", "admin")
    token = create_jwt(user["username"], tenant_id=tenant_id, role=role)
    # A tenant nevét is visszaadjuk a UI-nak (header/sidebar)
    tenant_name = ""
    if tenant_id:
        try:
            t = db.supabase.table("tenants").select("name").eq("id", tenant_id).limit(1).execute()
            tenant_name = (t.data[0].get("name") if t.data else "") or ""
        except Exception:
            tenant_name = ""
    return {
        "token": token,
        "username": user["username"],
        "role": role,
        "full_name": user.get("full_name", ""),
        "tenant_id": tenant_id,
        "tenant_name": tenant_name,
    }


def require_admin(credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    """Dependency: only admin role can access."""
    if not credentials:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Nincs token")
    try:
        payload = pyjwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        username = payload["sub"]
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token lejárt")
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Érvénytelen token")
    
    user = db.get_admin_user_by_username(username)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Csak admin jogosultsággal elérhető")
    return user


def require_admin_or_manager(credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    """Dependency: admin or manager role can access."""
    if not credentials:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Nincs token")
    try:
        payload = pyjwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        username = payload["sub"]
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token lejárt")
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Érvénytelen token")
    
    user = db.get_admin_user_by_username(username)
    if not user or user.get("role") not in ("admin", "manager"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Csak admin/manager jogosultsággal elérhető")
    return user


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN USER MANAGEMENT — admin/manager role required (with restrictions)
# ═══════════════════════════════════════════════════════════════════════════════

class CreateUserRequest(BaseModel):
    username: str
    password: str
    email: str = ""
    role: str = "member"
    full_name: str = ""

class RoleUpdateRequest(BaseModel):
    role: str
    full_name: Optional[str] = None

class ChangePasswordRequest(BaseModel):
    current_password: str = ""
    new_password: str
    user_id: Optional[int] = None

@app.get("/admin/api/users")
def api_get_users(caller: dict = Depends(require_admin_or_manager)):
    """List admin users. Manager only sees members."""
    all_users = db.get_admin_users()
    if caller.get("role") == "manager":
        all_users = [u for u in all_users if u.get("role") == "member"]
    return {"status": "success", "data": all_users}

@app.post("/admin/api/users")
def api_create_user(req: CreateUserRequest, caller: dict = Depends(require_admin_or_manager)):
    """Create a new user. Manager can only create members."""
    if req.role not in ("admin", "manager", "member"):
        raise HTTPException(400, "Érvénytelen szerepkör. Lehetséges: admin, manager, member")
    if caller.get("role") == "manager" and req.role != "member":
        raise HTTPException(403, "Manager csak member jogosultságú felhasználót hozhat létre")
    success = db.create_admin_user(req.username, req.password, req.email, req.role, caller["username"], req.full_name)
    if not success:
        raise HTTPException(400, "A felhasználónév már foglalt")
    return {"status": "success", "message": f"Felhasználó létrehozva: {req.username}"}

@app.put("/admin/api/users/{user_id}/role")
def api_update_user_role(user_id: int, req: RoleUpdateRequest, admin: dict = Depends(require_admin)):
    """Update user role (admin only). Cannot demote self."""
    if admin["id"] == user_id and req.role != "admin":
        raise HTTPException(400, "Nem módosíthatod a saját szerepkörödet")
    success = db.update_admin_role(user_id, req.role)
    if not success:
        raise HTTPException(400, "Szerepkör módosítása sikertelen")
    # Update full_name if provided
    if req.full_name is not None:
        try:
            db.supabase.table("admin_users").update({"full_name": req.full_name}).eq("id", user_id).execute()
        except Exception:
            pass  # non-critical
    return {"status": "success"}

class ProfileUpdateRequest(BaseModel):
    full_name: str = ""

@app.put("/admin/api/users/profile")
def api_update_own_profile(req: ProfileUpdateRequest, caller: str = Depends(verify_jwt)):
    """Update own profile (full_name). Any authenticated user can call this."""
    try:
        db.supabase.table("admin_users").update({"full_name": req.full_name}).eq("username", caller).execute()
    except Exception as ex:
        raise HTTPException(500, f"Profil mentés hiba: {ex}")
    return {"status": "success"}

class AvatarUploadRequest(BaseModel):
    avatar_data: str  # base64 data URL (e.g. "data:image/jpeg;base64,...")

@app.post("/admin/api/users/avatar")
async def api_upload_avatar(req: AvatarUploadRequest, username: str = Depends(verify_jwt)):
    """Upload profile avatar to Supabase Storage."""
    import base64 as b64mod
    data = req.avatar_data
    if not data.startswith("data:image/"):
        raise HTTPException(400, "Érvénytelen képformátum")
    if len(data) > 700_000:
        raise HTTPException(400, "A kép túl nagy (max ~500KB)")
    try:
        # Parse base64 data URL
        header, b64_content = data.split(",", 1)
        image_bytes = b64mod.b64decode(b64_content)
        content_type = header.split(":")[1].split(";")[0]  # e.g. "image/jpeg"
        ext = content_type.split("/")[1]  # e.g. "jpeg"
        file_path = f"{username}.{ext}"
        
        storage = db.supabase.storage.from_("avatars")
        # Try to remove existing file first (ignore errors)
        try:
            storage.remove([f"{username}.jpeg", f"{username}.jpg", f"{username}.png", f"{username}.webp"])
        except Exception:
            pass
        # Upload new file
        storage.upload(file_path, image_bytes, file_options={"content-type": content_type, "upsert": "true"})
        # Get public URL
        public_url = storage.get_public_url(file_path)
        # Add cache-busting parameter
        import time
        public_url = f"{public_url}?t={int(time.time())}"
        return {"status": "success", "avatar_url": public_url}
    except Exception as e:
        print(f"[Avatar Upload Error] {e}")
        raise HTTPException(500, f"Hiba: {e}")

@app.delete("/admin/api/users/avatar")
async def api_delete_avatar(username: str = Depends(verify_jwt)):
    """Remove profile avatar from Supabase Storage."""
    try:
        storage = db.supabase.storage.from_("avatars")
        storage.remove([f"{username}.jpeg", f"{username}.jpg", f"{username}.png", f"{username}.webp"])
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(500, f"Hiba: {e}")

@app.get("/admin/api/users/{username_or_id}/avatar")
async def api_get_avatar(username_or_id: str, _: str = Depends(verify_jwt)):
    """Get avatar URL for a user from Supabase Storage."""
    try:
        storage = db.supabase.storage.from_("avatars")
        # Try common extensions
        for ext in ["jpeg", "jpg", "png", "webp"]:
            file_path = f"{username_or_id}.{ext}"
            try:
                public_url = storage.get_public_url(file_path)
                # Verify it actually exists by listing
                files = storage.list()
                exists = any(f.get("name", "") == file_path for f in files)
                if exists:
                    import time
                    return {"avatar_url": f"{public_url}?t={int(time.time())}"}
            except Exception:
                continue
        return {"avatar_url": None}
    except Exception:
        return {"avatar_url": None}

@app.delete("/admin/api/users/{user_id}")
def api_delete_user(user_id: int, caller: dict = Depends(require_admin_or_manager)):
    """Delete a user. Manager can only delete members. Cannot delete self."""
    if caller["id"] == user_id:
        raise HTTPException(400, "Nem törölheted saját magadat")
    if caller.get("role") == "manager":
        target_user = db.get_admin_user_by_id(user_id) if hasattr(db, 'get_admin_user_by_id') else None
        if target_user is None:
            # Fallback: look up from all users
            all_users = db.get_admin_users()
            target_user = next((u for u in all_users if u.get("id") == user_id), None)
        if target_user and target_user.get("role") != "member":
            raise HTTPException(403, "Manager csak member felhasználót törölhet")
    success = db.delete_admin_user(user_id)
    if not success:
        raise HTTPException(400, "Törlés sikertelen")
    return {"status": "success"}

@app.post("/admin/api/users/change-password")
def api_change_password(req: ChangePasswordRequest, username: str = Depends(verify_jwt)):
    """Change password. Admins can change any user's password; members only their own."""
    caller = db.get_admin_user_by_username(username)
    if not caller:
        raise HTTPException(401, "Felhasználó nem található")
    
    if req.user_id and req.user_id != caller["id"]:
        # Changing someone else's password — admin only
        if caller.get("role") != "admin":
            raise HTTPException(403, "Csak admin módosíthat más felhasználó jelszavát")
        success = db.update_admin_password(req.user_id, req.new_password)
    else:
        # Changing own password — verify current password
        user = db.verify_admin_user(username, req.current_password)
        if not user:
            raise HTTPException(400, "A jelenlegi jelszó helytelen")
        success = db.update_admin_password(caller["id"], req.new_password)
    
    if not success:
        raise HTTPException(400, "Jelszó módosítása sikertelen")
    return {"status": "success", "message": "Jelszó sikeresen módosítva"}


@app.get("/admin/api/members")
def api_get_members(username: str = Depends(verify_jwt)):
    """List all member+manager users (for Felelős dropdown). Any logged-in user can access."""
    users = db.get_admin_users()
    members = [{"id": u["id"], "username": u["username"], "full_name": u.get("full_name", ""), "role": u.get("role", "member")} for u in users if u.get("role") in ("member", "manager")]
    return {"status": "success", "data": members}


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN API — protected routes
# ═══════════════════════════════════════════════════════════════════════════════


@app.get("/admin/api/stats")
def admin_stats(period: str = "month", channel: str = "mind", clinic_id: str = "mind", date_from: str = "", date_to: str = "", username: str = Depends(verify_jwt)):
    """Analytics summary stats."""
    return db.get_stats(period=period, channel=channel, clinic_id=clinic_id, date_from=date_from, date_to=date_to)

@app.get("/admin/api/analytics/funnel")
def admin_funnel(period: str = "month", channel: str = "mind", clinic_id: str = "mind", date_from: str = "", date_to: str = "", username: str = Depends(verify_jwt)):
    """Funnel stats based on interaction stages."""
    return db.get_funnel_stats(period=period, channel=channel, clinic_id=clinic_id, date_from=date_from, date_to=date_to)

@app.get("/admin/api/analytics/alerts")
def admin_alerts(period: str = "month", channel: str = "mind", clinic_id: str = "mind", date_from: str = "", date_to: str = "", username: str = Depends(verify_jwt)):
    """Operational alerts and tasks stats."""
    return db.get_alerts_stats(period=period, channel=channel, clinic_id=clinic_id, date_from=date_from, date_to=date_to)

@app.get("/admin/api/analytics/alerts/details")
def admin_alerts_details(type: str, username: str = Depends(verify_jwt)):
    """Get specific alert details."""
    details = db.get_alert_details(type)
    return {"status": "success", "data": details}

@app.get("/admin/api/analytics/insights")
def admin_get_insights(username: str = Depends(verify_jwt)):
    """Get latest AI insights."""
    insights = db.get_latest_ai_insights()
    if not insights:
        insights = [
            "Az árkérdések aránya 28%-kal nőtt – érdemes bővíteni az árakkal kapcsolatos tudásbázist.",
            "Az angol nyelvű megkeresések aránya nő – megfontolható az angol nyelvű válaszok fejlesztése.",
            "A 16:00–18:00 közti sávban emelkedett az átadási arány – érdemes vizsgálni az okot.",
            "Az időpontmódosítási kérdésekre érdemes új szabályt rögzíteni a Szabályok menüben."
        ]
    return {"status": "success", "insights": insights}

@app.post("/admin/api/analytics/insights/generate")
async def admin_generate_insights(_auth = Depends(require_admin_or_manager)):
    """Generate new AI insights based on stats."""
    stats = db.get_stats(period="month")
    google_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    insights = []
    
    if google_key:
        try:
            from google import genai
            from google.genai import types
            import json
            
            client = genai.Client(api_key=google_key)
            prompt = f"""
Te egy ügyfélszolgálati adatelemző AI vagy. Itt vannak az elmúlt időszak statisztikái:
- Összes megkeresés: {stats.get('total_interactions', 0)}
- Elakadt (nyitott) feladatok: {stats.get('open_tasks', 0)}
- Csatornák: {json.dumps(stats.get('interactions_by_type', []))}
- Témák: {json.dumps(stats.get('interactions_by_topic', []))}

Adj 4 darab, egy-egy mondatos, releváns finomhangolási javaslatot a folyamatok javítására ezen adatok alapján.
A válaszod kizárólag egy valid JSON lista legyen (pl. ["javaslat 1", "javaslat 2", ...]), markdown nélkül!
"""
            resp = await client.aio.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt,
                config=types.GenerateContentConfig(temperature=0.7)
            )
            t = resp.text.strip()
            if t.startswith("```json"): t = t[7:-3]
            if t.startswith("```"): t = t[3:-3]
            insights = json.loads(t.strip())
        except Exception as e:
            print(f"[Insights Generation Error] {e}")
            insights = []
            
    if not insights or len(insights) < 4:
        insights = [
            "Az árkérdések aránya 28%-kal nőtt – érdemes bővíteni az árakkal kapcsolatos tudásbázist.",
            "A telefonos csatorna továbbra is a legaktívabb, érdemes optimalizálni a hangasszisztenst.",
            "Sok a megválaszolatlan ügy, javasolt a belső értesítési rendszer felülvizsgálata.",
            "A 16:00–18:00 közti sávban emelkedett az átadási arány – érdemes vizsgálni az okot."
        ]
        
    db.save_ai_insights(insights)
    return {"status": "success", "insights": insights}

@app.get("/admin/api/analytics/outbound/summary")
def admin_outbound_summary(period: str = "month", channel: str = "mind", clinic_id: str = "mind", date_from: str = "", date_to: str = "", username: str = Depends(verify_jwt)):
    """Get outbound summary metrics."""
    return db.get_outbound_stats(period, channel, clinic_id, date_from=date_from, date_to=date_to)


@app.get("/admin/api/interactions")
def admin_interactions(
    limit: int = 100,
    type: str = "",
    username: str = Depends(verify_jwt)
):
    """Interaction list, newest first."""
    return {"interactions": db.get_interactions(limit=limit, type_filter=type)}


@app.get("/admin/api/interactions/grouped")
def admin_interactions_grouped(
    limit: int = 100,
    offset: int = 0,
    username: str = Depends(verify_jwt)
):
    """Session-aggregált interakciós lista (1 session = 1 sor, szerver-oldalon).
    Visszatérés: {sessions: [{session_id, interaction_count, last_created_at,
    session_statusz, representative}], total}."""
    return db.get_grouped_interactions(limit=limit, offset=offset)


@app.get("/admin/api/classification-labels")
def admin_classification_labels(username: str = Depends(verify_jwt)):
    """Kanonikus klasszifikációs címkék — a frontend ezzel validálhatja a
    RELATION_MATRIX-ét (backend↔frontend drift-megelőzés)."""
    import classifier as _clf
    return {
        "type_priority": _clf.TYPE_PRIORITY,
        "autonomous_automations": list(_clf.AUTONOMOUS_AUTOMATIONS),
        "ugytipusok": _clf.TYPE_PRIORITY,
        "statuszok": ["Lezárt", "Nyitott", "Sürgős"],
        "eredmenyek": [
            "Megválaszolt kérdés", "Válasz előkészítve", "Kérdés rögzítve",
            "Új időpont", "Módosított időpont", "Törölt időpont",
            "Foglalási szándék rögzítve", "Módosítási szándék rögzítve", "Lemondási szándék rögzítve",
            "Panasz rögzítve", "Igény rögzítve",
        ],
        "teendok": [
            "Nincs további teendő", "Jóváhagyás szükséges", "Időpont véglegesítése",
            "Válasz/visszahívás szükséges", "Intézkedés", "Azonnali beavatkozás szükséges",
        ],
    }


@app.get("/admin/api/calendar")
def admin_calendar(username: str = Depends(verify_jwt)):
    """Calendar events, sorted by start time."""
    return {"events": db.get_calendar_events()}


class ManualEventRequest(BaseModel):
    title: str
    attendee: str
    attendee_email: str = ""
    attendee_phone: str = ""
    start_dt: str  # ISO format datetime
    duration_minutes: int = 30

@app.post("/admin/api/calendar")
def admin_create_event(req: ManualEventRequest, _auth = Depends(require_admin_or_manager)):
    """Create a manual calendar event and auto-create client if needed."""
    from datetime import datetime, timedelta
    try:
        start = datetime.fromisoformat(req.start_dt.replace("Z", "+00:00"))
    except Exception:
        raise HTTPException(400, "Érvénytelen dátum formátum")
    
    end = start + timedelta(minutes=req.duration_minutes)
    
    event_id = db.add_calendar_event(
        title=req.title,
        start_dt=start.isoformat(),
        end_dt=end.isoformat(),
        duration_minutes=req.duration_minutes,
        attendee=req.attendee,
        attendee_email=req.attendee_email
    )
    
    # Auto-create client if not exists
    if req.attendee:
        existing = db.get_clients()
        found = False
        for c in existing:
            cd = c.get("custom_data")
            if isinstance(cd, str):
                try:
                    import json
                    cd = json.loads(cd)
                except:
                    cd = {}
            elif cd is None:
                cd = {}
            c_name = (cd.get("nev") or cd.get("name") or c.get("name") or "").lower().strip()
            c_email = (cd.get("email") or c.get("email") or "").lower().strip()
            if (req.attendee.lower().strip() == c_name) or (req.attendee_email and req.attendee_email.lower().strip() == c_email):
                found = True
                break
        
        if not found:
            # Get user's full_name for felelos
            felelos = _auth.get("full_name", "") or _auth.get("username", "")
            
            custom_data = {
                "name": req.attendee,
                "nev": req.attendee,
                "email": req.attendee_email,
                "phone": req.attendee_phone,
                "telefonszam": req.attendee_phone,
                "felelos": felelos
            }
            db.add_client(custom_data, "uj")
    
    return {"status": "success", "event_id": event_id, "message": "Időpont sikeresen létrehozva"}


@app.get("/admin/api/emails")
def admin_emails(limit: int = 100, username: str = Depends(verify_jwt)):
    """Email logs, newest first."""
    return {"emails": db.get_email_logs(limit=limit)}


@app.get("/admin/api/tasks")
def admin_tasks(completed: str = "all", username: str = Depends(verify_jwt)):
    """Task list."""
    comp = None if completed == "all" else (completed == "true")
    return {"tasks": db.get_tasks(completed=comp)}


@app.patch("/admin/api/tasks/{task_id}/complete")
def admin_task_complete(task_id: int, username: str = Depends(verify_jwt)):
    """Toggle task completed status."""
    res = db.update_task_complete(task_id)
    if not res.get("ok"):
        raise HTTPException(status_code=404, detail="Task not found or update failed")
    return {"ok": True, "completed": res.get("completed", False)}


@app.delete("/admin/api/tasks/{task_id}")
def admin_task_delete(task_id: int, _auth = Depends(require_admin_or_manager)):
    """Delete a task."""
    res = db.delete_task(task_id)
    if not res:
        raise HTTPException(status_code=404, detail="Delete failed")
    return {"ok": True}


@app.get("/admin/api/sessions")
def admin_sessions(limit: int = 50, username: str = Depends(verify_jwt)):
    """Recent sessions."""
    return {"sessions": db.get_sessions(limit=limit)}


@app.get("/admin/api/sessions/summary")
def admin_sessions_summary(limit: int = 50, username: str = Depends(verify_jwt)):
    """Sessions enriched with interaction summaries."""
    return {"sessions": db.get_sessions_with_summary(limit=limit)}


class BulkDeleteInteractionsRequest(BaseModel):
    interaction_ids: list[int] = []
    session_ids: list[str] = []

@app.post("/admin/api/interactions/delete")
def admin_delete_interactions(req: BulkDeleteInteractionsRequest, _auth = Depends(require_admin_or_manager)):
    """Delete interactions and/or sessions by ID."""
    deleted_interactions = 0
    deleted_sessions = 0
    try:
        if req.interaction_ids and db.supabase:
            db.supabase.table("interactions").delete().in_("id", req.interaction_ids).execute()
            deleted_interactions = len(req.interaction_ids)
        if req.session_ids and db.supabase:
            # Also delete interactions belonging to these sessions
            db.supabase.table("interactions").delete().in_("session_id", req.session_ids).execute()
            db.supabase.table("sessions").delete().in_("session_id", req.session_ids).execute()
            deleted_sessions = len(req.session_ids)
    except Exception as e:
        logger.error(f"Delete interactions error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    return {"deleted_interactions": deleted_interactions, "deleted_sessions": deleted_sessions}


class InteractionStatusUpdateRequest(BaseModel):
    status: str

@app.patch("/admin/api/interactions/{id}/status")
def update_interaction_status(id: int, req: InteractionStatusUpdateRequest, _auth = Depends(require_admin_or_manager)):
    """Safe status update that handles both approval_status and classification override."""
    if req.status != "lezárt":
        raise HTTPException(status_code=400, detail="Érvénytelen státusz érték. Csak a 'lezárt' támogatott.")
    
    try:
        if not db.supabase:
            raise Exception("Database connection not available")

        # FONTOS: az approval_status-t NEM írjuk — az az approval-flow domainje
        # (pending/approved/rejected/spam). Korábban 'lezárt'-ot írt bele, ami a
        # statusz-értékkészlet szennyezése volt. A lezárást a classification
        # statusz mezője hordozza (a frontend detectStatusz ezt olvassa elsőként).
        updates = {}

        # Fetch existing classification to avoid overriding other fields (ugytipus, eredmeny, etc.)
        res = db.supabase.table("interactions").select("classification").eq("id", id).execute()
        if res.data and len(res.data) > 0:
            cls = res.data[0].get("classification")
            if isinstance(cls, dict):
                cls["statusz"] = "Lezárt"
                cls["teendo"] = "Nincs további teendő"
                updates["classification"] = cls
            else:
                updates["classification"] = {"statusz": "Lezárt", "teendo": "Nincs további teendő"}
        else:
            updates["classification"] = {"statusz": "Lezárt", "teendo": "Nincs további teendő"}

        db.supabase.table("interactions").update(updates).eq("id", id).execute()
        logger.info(f"Interaction {id} marked as lezárt by {_auth.get('username')}")
        return {"status": "success", "message": "Interakció lezárva"}
    except Exception as e:
        logger.error(f"Status update error for interaction {id}: {e}")
        raise HTTPException(status_code=500, detail="Hiba a státusz frissítésekor az adatbázisban")


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENTS (KANBAN) API
# ═══════════════════════════════════════════════════════════════════════════════

class ClientCreateRequest(BaseModel):
    custom_data: dict

class ClientStatusUpdateRequest(BaseModel):
    status: str

@app.get("/admin/api/alerts/urgent")
def admin_alerts_urgent(username: str = Depends(verify_jwt)):
    """Fetch unhandled urgent alerts for notification."""
    clients = db.get_clients()
    urgent_clients = []
    
    for c in clients:
        status = c.get("status", "uj")
        if status in ["szerzodott", "siker", "sikeres", "lezart", "sikertelen"]:
            continue
            
        custom_data = c.get("custom_data")
        if isinstance(custom_data, str):
            try:
                import json
                custom_data = json.loads(custom_data)
            except:
                custom_data = {}
        if not isinstance(custom_data, dict):
            custom_data = {}
            
        if (custom_data.get("prioritas") == "Sürgős" or custom_data.get("priority") == "Sürgős") and not custom_data.get("urgent_viewed"):
            # Extract basic info for the toast
            name = custom_data.get("nev") or custom_data.get("name") or custom_data.get("név") or c.get("name", "Ismeretlen")
            channel = custom_data.get("forras_csatorna") or ("Messenger" if custom_data.get("messenger_id") else "Manuális")
            problem = custom_data.get("problem_description") or "Sürgős megkeresés beérkezett."
            
            urgent_clients.append({
                "id": c.get("id"),
                "name": name,
                "channel": channel,
                "problem": problem,
                "created_at": c.get("created_at")
            })
            
    return {"urgent_clients": urgent_clients}

@app.post("/admin/api/alerts/urgent/{client_id}/view")
def mark_urgent_alert_viewed(client_id: int, username: str = Depends(verify_jwt)):
    """Mark an urgent alert as viewed so it disappears from the notification bell."""
    clients = db.get_clients()
    for c in clients:
        if c.get("id") == client_id:
            custom_data = c.get("custom_data")
            if isinstance(custom_data, str):
                try:
                    import json
                    custom_data = json.loads(custom_data)
                except:
                    custom_data = {}
            if not isinstance(custom_data, dict):
                custom_data = {}
            
            custom_data["urgent_viewed"] = True
            db.edit_client_details(client_id, custom_data)
            return {"status": "success"}
    
    raise HTTPException(status_code=404, detail="Client not found")

@app.get("/admin/api/alerts/cancelled")
def admin_alerts_cancelled(username: str = Depends(verify_jwt)):
    """Fetch unhandled cancellation alerts for notification."""
    clients = db.get_clients()
    cancelled_clients = []
    
    for c in clients:
        status = c.get("status", "uj")
        if status != "lemondott":
            continue
            
        custom_data = c.get("custom_data")
        if isinstance(custom_data, str):
            try:
                import json
                custom_data = json.loads(custom_data)
            except:
                custom_data = {}
        if not isinstance(custom_data, dict):
            custom_data = {}
            
        if not custom_data.get("cancelled_viewed"):
            name = custom_data.get("nev") or custom_data.get("name") or custom_data.get("név") or c.get("name", "Ismeretlen")
            channel = custom_data.get("forras_csatorna") or ("Messenger" if custom_data.get("messenger_id") else "Manuális")
            email = custom_data.get("email") or c.get("email", "")
            phone = custom_data.get("phone") or c.get("phone", "")
            
            cancelled_clients.append({
                "id": c.get("id"),
                "name": name,
                "email": email,
                "phone": phone,
                "channel": channel,
                "created_at": c.get("created_at")
            })
            
    return {"cancelled_clients": cancelled_clients}

@app.post("/admin/api/alerts/cancelled/{client_id}/view")
def mark_cancelled_alert_viewed(client_id: int, username: str = Depends(verify_jwt)):
    """Mark a cancelled alert as viewed."""
    clients = db.get_clients()
    for c in clients:
        if c.get("id") == client_id:
            custom_data = c.get("custom_data")
            if isinstance(custom_data, str):
                try:
                    import json
                    custom_data = json.loads(custom_data)
                except:
                    custom_data = {}
            if not isinstance(custom_data, dict):
                custom_data = {}
            
            custom_data["cancelled_viewed"] = True
            db.edit_client_details(client_id, custom_data)
            return {"status": "success"}
    
    raise HTTPException(status_code=404, detail="Client not found")

@app.get("/admin/api/clients")
def admin_clients(username: str = Depends(verify_jwt)):
    """List all clients for Kanban."""
    clients = db.get_clients()
    for c in clients:
        if "custom_data" in c and isinstance(c["custom_data"], dict):
            c["custom_data"] = json.dumps(c["custom_data"])
    return {"clients": clients}

@app.post("/admin/api/clients")
def admin_add_client(req: ClientCreateRequest, _auth = Depends(require_admin_or_manager)):
    """Add a new client."""
    client_id = db.add_client(req.custom_data, "uj")
    return {"ok": True, "id": client_id}

@app.patch("/admin/api/clients/{client_id}/status")
def admin_update_client_status(client_id: int, req: ClientStatusUpdateRequest, _auth = Depends(require_admin_or_manager)):
    """Update client status (drag & drop)."""
    db.update_client_status(client_id, req.status)
    return {"ok": True}

@app.delete("/admin/api/clients/{client_id}")
def admin_delete_client(client_id: int, _auth = Depends(require_admin_or_manager)):
    """Delete client."""
    db.delete_client(client_id)
    return {"ok": True}

class BulkDeleteClientsRequest(BaseModel):
    client_ids: list[int]

@app.post("/admin/api/clients/bulk_delete")
def admin_bulk_delete_clients(req: BulkDeleteClientsRequest, _auth = Depends(require_admin_or_manager)):
    """Delete multiple clients."""
    for cid in req.client_ids:
        db.delete_client(cid)
    return {"ok": True}

@app.put("/admin/api/clients/{client_id}")
def admin_update_client_details(client_id: int, req: ClientCreateRequest, _auth = Depends(require_admin_or_manager)):
    """Update client basic details."""
    db.edit_client_details(client_id, req.custom_data)
    return {"ok": True}

class ClientFieldCreateRequest(BaseModel):
    id: str
    name: str
    order_index: int

class ClientFieldUpdateRequest(BaseModel):
    name: str

@app.get("/admin/api/client_fields")
def admin_get_client_fields(username: str = Depends(verify_jwt)):
    return {"fields": db.get_client_fields()}

@app.post("/admin/api/client_fields")
def admin_add_client_field(req: ClientFieldCreateRequest, _admin = Depends(require_admin)):
    success = db.add_client_field(req.id, req.name, req.order_index)
    if not success:
        raise HTTPException(status_code=400, detail="Field ID already exists")
    return {"ok": True}

@app.put("/admin/api/client_fields/{field_id}")
def admin_update_client_field(field_id: str, req: ClientFieldUpdateRequest, _admin = Depends(require_admin)):
    db.update_client_field(field_id, req.name)
    return {"ok": True}

@app.delete("/admin/api/client_fields/{field_id}")
def admin_delete_client_field(field_id: str, _admin = Depends(require_admin)):
    db.delete_client_field(field_id)
    return {"ok": True}

class KanbanColumnCreateRequest(BaseModel):
    id: str
    name: str
    order_index: int

class KanbanColumnUpdateRequest(BaseModel):
    name: str

@app.get("/admin/api/kanban_columns")
def admin_get_kanban_columns(username: str = Depends(verify_jwt)):
    return {"columns": db.get_kanban_columns()}

@app.post("/admin/api/kanban_columns")
def admin_add_kanban_column(req: KanbanColumnCreateRequest, _admin = Depends(require_admin)):
    success = db.add_kanban_column(req.id, req.name, req.order_index)
    if not success:
        raise HTTPException(status_code=400, detail="Column ID already exists")
    return {"ok": True}

@app.put("/admin/api/kanban_columns/{col_id}")
def admin_update_kanban_column(col_id: str, req: KanbanColumnUpdateRequest, _admin = Depends(require_admin)):
    db.update_kanban_column(col_id, req.name)
    return {"ok": True}

@app.delete("/admin/api/kanban_columns/{col_id}")
def admin_delete_kanban_column(col_id: str, _admin = Depends(require_admin)):
    try:
        db.delete_kanban_column(col_id)
        return {"ok": True}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ── Legacy public API (for backward compat with voice-widget.html) ────────────
@app.get("/api/calendar")
def get_calendar():
    events = db.get_calendar_events()
    return JSONResponse({"events": events})

@app.get("/api/emails")
def get_emails():
    return JSONResponse({"emails": db.get_email_logs()})


# ═══════════════════════════════════════════════════════════════════════════════
# Settings — now backed by Supabase tables (agent_settings, knowledge_base,
#            business_info, text_configs).  File paths kept only for legacy
#            fallback / initial seeding.
# ═══════════════════════════════════════════════════════════════════════════════

# Legacy file paths — used only for initial seeding into Supabase on first boot
SYSTEM_PROMPT_FILE = THIS_DIR / "system_prompt.md"
WORKFLOW_FILE      = THIS_DIR / "workflow.md"

DEFAULT_SETTINGS = {
    "voice_id": os.getenv("CARTESIA_VOICE_ID", "93896c4f-aa00-4c17-a360-fec55579d7fa"),
    "tone": "professional_friendly",
    "tone_custom": "",
    "knowledge_format": "json",
    "greeting": "",
    "language": "hu",
    "business_hours": {
        "monday":    {"open": "09:00", "close": "18:00", "enabled": True},
        "tuesday":   {"open": "09:00", "close": "18:00", "enabled": True},
        "wednesday": {"open": "09:00", "close": "18:00", "enabled": True},
        "thursday":  {"open": "09:00", "close": "18:00", "enabled": True},
        "friday":    {"open": "09:00", "close": "16:00", "enabled": True},
        "saturday":  {"open": None,    "close": None,    "enabled": False},
        "sunday":    {"open": None,    "close": None,    "enabled": False},
    },
}


def _read_settings() -> dict:
    """Read agent settings from Supabase, fall back to DEFAULT_SETTINGS."""
    s = db.get_agent_settings()
    if s:
        # Merge with defaults so any missing keys get filled
        merged = dict(DEFAULT_SETTINGS)
        merged.update({k: v for k, v in s.items() if v is not None})
        return merged
    return dict(DEFAULT_SETTINGS)


def _read_knowledge() -> dict:
    """Read knowledge content from Supabase."""
    k = db.get_knowledge_base()
    return {"format": k.get("format", "json"), "content": k.get("content", "{}")}


@app.get("/admin/api/settings")
async def get_settings(username: str = Depends(verify_jwt)):
    """Return current agent settings + knowledge base content."""
    s = _read_settings()
    k = _read_knowledge()
    return {**s, "knowledge_content": k["content"]}


class SettingsSaveRequest(BaseModel):
    voice_id: str = ""
    tone: str = "professional_friendly"
    tone_custom: str = ""
    knowledge_format: str = "json"
    knowledge_content: str = ""
    greeting: str = ""
    language: str = "hu"
    business_hours: dict = {}


class TextFileRequest(BaseModel):
    content: str = ""


@app.post("/admin/api/settings")
async def save_settings(payload: SettingsSaveRequest, _admin = Depends(require_admin)):
    """Save agent settings and knowledge base to Supabase."""
    print(f"[DEBUG] save_settings received payload: {payload.dict()}", flush=True)
    settings = {
        "voice_id":         payload.voice_id,
        "tone":             payload.tone,
        "tone_custom":      payload.tone_custom,
        "knowledge_format": payload.knowledge_format,
        "greeting":         payload.greeting,
        "language":         payload.language,
        "business_hours":   payload.business_hours,
    }
    ok = db.update_agent_settings(settings)
    if not ok:
        raise HTTPException(status_code=500, detail="Nem sikerült menteni a beállításokat az adatbázisba.")

    # Save knowledge content
    if payload.knowledge_content:
        if payload.knowledge_format != "md":
            try:
                json.loads(payload.knowledge_content)  # validate JSON
            except json.JSONDecodeError as e:
                print(f"[DEBUG] JSON decode error: {e} | Content: {payload.knowledge_content}", flush=True)
                raise HTTPException(status_code=400, detail=f"Hibás JSON formátum: {e}")
        db.update_knowledge_base(payload.knowledge_format, payload.knowledge_content)

    return {"ok": True, "message": "Beállítások elmentve. A változtatások a következő hívásnál már érvényesek."}


# ── System Prompt ─────────────────────────────────────────────────────────────

@app.get("/admin/api/system-prompt")
async def get_system_prompt(username: str = Depends(verify_jwt)):
    """Return the current system prompt from Supabase (falls back to local file for seeding)."""
    content = db.get_text_config("system_prompt")
    if not content and SYSTEM_PROMPT_FILE.exists():
        content = SYSTEM_PROMPT_FILE.read_text(encoding="utf-8")
        db.update_text_config("system_prompt", content)
    return {"content": content}


@app.post("/admin/api/system-prompt")
async def save_system_prompt(payload: TextFileRequest, _admin = Depends(require_admin)):
    """Save system prompt to Supabase."""
    db.update_text_config("system_prompt", payload.content)
    return {"ok": True, "message": "System prompt elmentve."}


# ── Workflow ───────────────────────────────────────────────────────────────────

@app.get("/admin/api/workflow")
async def get_workflow(username: str = Depends(verify_jwt)):
    """Return the current workflow from Supabase (falls back to local file for seeding)."""
    content = db.get_text_config("workflow")
    if not content and WORKFLOW_FILE.exists():
        content = WORKFLOW_FILE.read_text(encoding="utf-8")
        db.update_text_config("workflow", content)
    return {"content": content}


@app.post("/admin/api/workflow")
async def save_workflow(payload: TextFileRequest, _admin = Depends(require_admin)):
    """Save workflow to Supabase."""
    db.update_text_config("workflow", payload.content)
    return {"ok": True, "message": "Workflow elmentve."}


# ── Written behavior (íráos kommunikáció beállítás: autonomous/approval) ──────

@app.get("/admin/api/written-behavior")
async def get_written_behavior(username: str = Depends(verify_jwt)):
    """EAISY-241 — Return the written communication behavior setting."""
    content = db.get_text_config("written_behavior") or "autonomous"
    return {"content": content}


@app.post("/admin/api/written-behavior")
async def save_written_behavior(payload: TextFileRequest, _admin = Depends(require_admin)):
    """EAISY-241 — Save the written communication behavior setting."""
    db.update_text_config("written_behavior", payload.content)
    import classifier as _clf
    _clf.invalidate_classifier_cache()
    return {"ok": True, "message": "Írásos kommunikáció beállítása elmentve."}


# ── Issue-handling (ügykezelési szabályok) — teljes állapot a backendben ──────
# Korábban a written_behavior kivételével minden (defaultRequestNotify,
# defaultComplaintNotify, customRules) csak localStorage-ban élt — böngészőnként
# eltért, és a classifier nem látta. Most a text_configs az igaz forrás.

class IssueHandlingRequest(BaseModel):
    writtenBehavior: str = "autonomous"
    defaultRequestNotify: str = "email"
    defaultComplaintNotify: str = "email"
    customRules: list = []

@app.get("/admin/api/issue-handling")
async def get_issue_handling(username: str = Depends(verify_jwt)):
    """Ügykezelési szabályok teljes állapota (JSON a text_configs-ban)."""
    raw = db.get_text_config("issue_handling")
    if raw:
        try:
            return json.loads(raw)
        except Exception:
            pass
    # Back-compat: ha még nincs issue_handling, a written_behavior-ból seedelünk
    return {
        "writtenBehavior": db.get_text_config("written_behavior") or "autonomous",
        "defaultRequestNotify": "email",
        "defaultComplaintNotify": "email",
        "customRules": [],
    }

@app.post("/admin/api/issue-handling")
async def save_issue_handling(payload: IssueHandlingRequest, _admin = Depends(require_admin)):
    """Ügykezelési szabályok mentése — a written_behavior-t külön kulcsra is
    kiírjuk (a classifier azt olvassa)."""
    data = payload.model_dump()
    db.update_text_config("issue_handling", json.dumps(data, ensure_ascii=False))
    db.update_text_config("written_behavior", payload.writtenBehavior)
    import classifier as _clf
    _clf.invalidate_classifier_cache()
    return {"ok": True, "message": "Ügykezelési szabályok elmentve."}



# ── Céginformációk ────────────────────────────────────────────────────────────────

_HU_TO_EN = {
    "hetfo": "monday", "kedd": "tuesday", "szerda": "wednesday",
    "csutortok": "thursday", "pentek": "friday",
    "szombat": "saturday", "vasarnap": "sunday",
}

class BusinessInfoSaveRequest(BaseModel):
    practice_name: str = ""
    description:   str = ""
    address:       str = ""
    markanev:      str = ""
    szakterulet:   str = ""
    kulcsszavak:   str = ""
    megkozelites:  str = ""
    price_list:    str = ""
    price_list_file_meta: dict = {}
    campaigns: list = []
    exceptions: list = []
    faq: list = []
    modositas_eng: str = "igen"
    modositas_szoveg: str = ""
    lemondas_24h: str = "figyelmeztetoSzoveggel"
    figyelmezteto_szoveg: str = ""
    pacient_id_question: str = "Korábban járt már a rendelőnkben?"
    new_patient_required: str = "Születési dátum, teljes név"
    new_patient_auto_visit: bool = True
    returning_patient_required: str = "Páciens azonosító vagy telefonszám"
    service_description: str = ""
    sender_name: str = ""
    sender_email: str = ""

@app.get("/admin/api/triage_rules")
def api_get_triage_rules(admin: dict = Depends(verify_jwt)):
    return db.get_triage_rules()

class TriageRuleCreate(BaseModel):
    situation: str
    priority: str
    escalation_email: str = ""

@app.post("/admin/api/triage_rules")
def api_post_triage_rules(rule: TriageRuleCreate, _admin = Depends(require_admin)):
    # Upsert situation alapján — a unique index mellett case-eltérésnél sem lesz
    # 500-as (korábban sima insert volt, ami duplikációt/23505-öt okozhatott)
    new_id = db.upsert_triage_rule(rule.situation, rule.priority, rule.escalation_email)
    if new_id:
        import classifier as _clf
        _clf.invalidate_classifier_cache()
        return {"ok": True, "id": new_id}
    raise HTTPException(status_code=500, detail="Hiba a létrehozáskor")

@app.put("/admin/api/triage_rules/{rule_id}")
def api_put_triage_rules(rule_id: int, rule: TriageRuleCreate, _admin = Depends(require_admin)):
    if db.update_triage_rule(rule_id, rule.situation, rule.priority, rule.escalation_email):
        import classifier as _clf
        _clf.invalidate_classifier_cache()
        return {"ok": True}
    raise HTTPException(status_code=400, detail="Hiba a frissítéskor")

@app.delete("/admin/api/triage_rules/{rule_id}")
def api_delete_triage_rules(rule_id: int, _admin = Depends(require_admin)):
    if db.delete_triage_rule(rule_id):
        import classifier as _clf
        _clf.invalidate_classifier_cache()
        return {"ok": True}
    raise HTTPException(status_code=400, detail="Hiba a törléskor")


# ── Szolgáltatások ────────────────────────────────────────────────────────────

class ServiceCreate(BaseModel):
    service_name: str
    duration_minutes: int
    description: str = ""
    assigned_to: str = ""
    note: str = ""

@app.get("/admin/api/services")
def api_get_services(admin: dict = Depends(verify_jwt)):
    return db.get_services()

@app.post("/admin/api/services")
def api_post_services(svc: ServiceCreate, _auth = Depends(require_admin_or_manager)):
    new_id = db.add_service(svc.service_name, svc.duration_minutes, svc.description, svc.assigned_to, svc.note)
    if new_id:
        return {"ok": True, "id": new_id}
    raise HTTPException(status_code=500, detail="Hiba a létrehozáskor")

@app.put("/admin/api/services/{srv_id}")
def api_put_services(srv_id: int, svc: ServiceCreate, _auth = Depends(require_admin_or_manager)):
    if db.update_service(srv_id, svc.service_name, svc.duration_minutes, svc.description, svc.assigned_to, svc.note):
        return {"ok": True}
    raise HTTPException(status_code=400, detail="Hiba a frissítéskor")

@app.delete("/admin/api/services/{srv_id}")
def api_delete_services(srv_id: int, _auth = Depends(require_admin_or_manager)):
    if db.delete_service(srv_id):
        return {"ok": True}
    raise HTTPException(status_code=400, detail="Hiba a törléskor")


@app.get("/admin/api/business-info")
async def get_business_info(username: str = Depends(verify_jwt)):
    """Return saved practice info from Supabase."""
    return db.get_business_info()

@app.post("/admin/api/business-info")
async def save_business_info(payload: BusinessInfoSaveRequest, _admin = Depends(require_admin)):
    """Save practice info to Supabase. Business hours are managed separately via /admin/api/settings."""
    data = {
        "practice_name": payload.practice_name,
        "description":   payload.description,
        "address":       payload.address,
        "markanev":      payload.markanev,
        "szakterulet":   payload.szakterulet,
        "kulcsszavak":   payload.kulcsszavak,
        "megkozelites":  payload.megkozelites,
        "price_list":    payload.price_list,
        "price_list_file_meta": payload.price_list_file_meta,
        "campaigns":     payload.campaigns,
        "exceptions":    payload.exceptions,
        "faq":           payload.faq,
        "modositas_eng": payload.modositas_eng,
        "modositas_szoveg": payload.modositas_szoveg,
        "lemondas_24h":  payload.lemondas_24h,
        "figyelmezteto_szoveg": payload.figyelmezteto_szoveg,
        "pacient_id_question": payload.pacient_id_question,
        "new_patient_required": payload.new_patient_required,
        "new_patient_auto_visit": payload.new_patient_auto_visit,
        "returning_patient_required": payload.returning_patient_required,
        "service_description": payload.service_description,
        "sender_name": payload.sender_name,
        "sender_email": payload.sender_email,
    }
    ok = db.update_business_info(data)
    if not ok:
        raise HTTPException(status_code=500, detail="Nem sikerült menteni a céginformációt.")
    return {"ok": True, "message": "Céginformáció elmentve."}

@app.get("/admin/api/prices/template/download")
async def download_price_template(_admin = Depends(require_admin)):
    """Generate and return an Excel template for price list upload."""
    import openpyxl
    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Árlista Minta"
    
    # Headers
    headers = ["Kategória", "Szolgáltatás megnevezése", "Ár", "Pénznem", "Megjegyzés / Extra infó"]
    ws.append(headers)
    
    # Premium Sample data
    sample_data = [
        ["Konzultáció", "Szakorvosi állapotfelmérés és kezelési terv", 15000, "HUF", "Tartalmazza a szájüregi szűrővizsgálatot."],
        ["Diagnosztika", "Digitális Panoráma röntgen (OPG)", 12000, "HUF", "Kiadható digitális formátumban (CD / E-mail)."],
        ["Szájhigiénia", "Ultrahangos fogkőeltávolítás (Air-Flow)", 28000, "HUF", "Mindkét állcsontra, polírozással együtt."],
        ["Konzerváló fogászat", "Esztétikus kompozit tömés (1 felszínű)", 25000, "HUF", "Fényre kötő prémium tömőanyag."],
        ["Szájsebészet", "Bölcsességfog műtéti eltávolítása", 45000, "HUF", "A bonyolultságtól függően az ár változhat."],
        ["Implantológia", "Prémium implantátum beültetése", 250000, "HUF", "Az ár a felépítményt és koronát nem tartalmazza."]
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # Styling
    from openpyxl.styles import PatternFill, Font
    header_fill = PatternFill(start_color="14B8A6", end_color="14B8A6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    # Freeze panes
    ws.freeze_panes = "A2"
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="arlista_minta.xlsx"'}
    )

@app.post("/admin/api/upload_prices")
async def upload_prices(file: UploadFile = File(...), _admin = Depends(require_admin)):
    content = await file.read()
    prices_text = ""
    fname = (file.filename or "").lower()
    print(f"[DEBUG upload_prices] filename={file.filename}, size={len(content)} bytes", flush=True)
    
    if fname.endswith(".csv"):
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("iso-8859-2", errors="replace")
        
        reader = csv.reader(io.StringIO(text))
        lines = []
        for row in reader:
            if any(row):
                lines.append(" - ".join(str(item).strip() for item in row if str(item).strip()))
        prices_text = "\n".join(lines)
        
    elif fname.endswith(".xlsx"):
        try:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(content))
            lines = []
            for _, row in df.iterrows():
                row_items = [str(item).strip() for item in row if pd.notna(item) and str(item).strip()]
                if row_items:
                    lines.append(" - ".join(row_items))
            prices_text = "\n".join(lines)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Nem sikerült beolvasni az Excel fájlt: {e}")
    else:
        raise HTTPException(status_code=400, detail=f"Kizárólag .csv vagy .xlsx fájl tölthető fel! (kapott: {file.filename})")
        
    if not prices_text.strip():
        raise HTTPException(status_code=400, detail="A fájl üres vagy nem tartalmaz értelmezhető adatot.")
        
    data = db.get_business_info()
            
    data["price_list"] = prices_text
    data["price_list_file_meta"] = {
        "filename": file.filename,
        "uploaded_at": datetime.now().strftime("%Y. %m. %d.")
    }
    
    db.update_business_info(data)
    
    return {
        "ok": True, 
        "message": "Árlista sikeresen feltöltve és feldolgozva.", 
        "price_list": prices_text,
        "price_list_file_meta": data["price_list_file_meta"]
    }



@app.get("/admin/api/cartesia/voices")
async def cartesia_voices(username: str = Depends(verify_jwt)):
    """Proxy: list available Cartesia voices."""
    api_key = os.getenv("CARTESIA_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="CARTESIA_API_KEY nincs beállítva")
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(
                "https://api.cartesia.ai/voices",
                headers={"X-API-Key": api_key, "Cartesia-Version": "2024-06-10"}
            )
        return resp.json()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Cartesia API hiba: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# SIP OUTBOUND CALL API
# ═══════════════════════════════════════════════════════════════════════════════

class SipCallRequest(BaseModel):
    phone_number: str   # E.164 format pl. +36301234567
    note: str = ""      # Megjegyzés
    script: str = ""    # Szöveg amit az AI-nak el kell mondania a hívás során
    client_name: str = ""  # Ügyfél neve (személyre szabáshoz)

@app.post("/admin/api/sip/call")
async def sip_outbound_call(req: SipCallRequest, _auth = Depends(require_admin_or_manager)):
    """Kimenő SIP hívás indítása az AI agenttel — opcionális scripttel."""
    from livekit import api as lk_api_module

    lk_url    = os.getenv("LIVEKIT_URL")
    lk_key    = os.getenv("LIVEKIT_API_KEY")
    lk_secret = os.getenv("LIVEKIT_API_SECRET")
    trunk_id  = os.getenv("SIP_OUTBOUND_TRUNK_ID", "ST_jgPctgJYZcAf")  # Telnyx HD Voice outbound

    phone = req.phone_number.strip()
    if not phone.startswith("+"):
        phone = "+" + phone

    # Egyedi szoba a híváshoz
    room_name = f"call-out-{uuid.uuid4().hex[:8]}"

    # Ha van script, metadata-ba tesszük hogy az agent megkapja
    call_metadata = None
    if req.script and req.script.strip():
        call_metadata = json.dumps({
            "type": "outbound_script_call",
            "script": req.script.strip(),
            "client_name": req.client_name.strip() if req.client_name else "",
            "call_note": req.note.strip() if req.note else "",
            # FÁZIS 6: a tenantot is átadjuk a metadata-ban — a voice agent
            # entrypoint így a helyes tenant scope-ban dolgozik
            "tenant_id": db.get_current_tenant() or "",
        })

    try:
        lk = lk_api_module.LiveKitAPI(url=lk_url, api_key=lk_key, api_secret=lk_secret)

        # 1. Szoba létrehozása (metadata-val ha van script)
        await lk.room.create_room(
            lk_api_module.CreateRoomRequest(
                name=room_name,
                empty_timeout=120,
                metadata=call_metadata or "",
            )
        )

        # 2. Először SIP hívás -- megvárjuk amíg felveszik
        participant = await lk.sip.create_sip_participant(
            lk_api_module.CreateSIPParticipantRequest(
                sip_trunk_id=trunk_id,
                sip_call_to=phone,
                room_name=room_name,
                participant_identity="phone-caller",
                participant_name=phone,
                wait_until_answered=True,
                krisp_enabled=True,
            )
        )

        # 3. Csak ha felvették: agent dispatch (metadata-val ha van script)
        await lk.agent_dispatch.create_dispatch(
            lk_api_module.CreateAgentDispatchRequest(
                agent_name="dobozos-ai",
                room=room_name,
                metadata=call_metadata or "outbound_call",
            )
        )

        await lk.aclose()

        return {
            "ok": True,
            "room": room_name,
            "phone": phone,
            "participant": participant.participant_identity,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hívás sikertelen: {str(e)}")




# ═══════════════════════════════════════════════════════════════════════════════
# JÓVÁHAGYÓ RENDSZER (HUMAN-IN-THE-LOOP) API
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/api/approvals")
def get_approvals_api(status: str = "pending", username: str = Depends(verify_jwt)):
    approvals = db.get_approvals(status)
    return {"approvals": approvals}

@app.post("/admin/api/approvals/{id}/reject")
def reject_approval_api(id: int, _auth = Depends(require_admin_or_manager)):
    success = db.update_approval_status(id, "rejected")
    if success: return {"status": "success"}
    raise HTTPException(status_code=500, detail="Hiba az elutasítás során")

class DeleteApprovalsRequest(BaseModel):
    ids: list[int]

@app.delete("/admin/api/approvals")
def delete_approvals_api(req: DeleteApprovalsRequest, _auth = Depends(require_admin_or_manager)):
    success = db.delete_approvals(req.ids)
    if success: return {"status": "success"}
    raise HTTPException(status_code=500, detail="Hiba a törlés során")

class ApproveRequest(BaseModel):
    modified_draft: str = ""
    # Multi-channel draftnál a frontend csatornánként szerkeszt: {channel: szöveg}
    modified_drafts: dict | None = None

@app.post("/admin/api/approvals/{id}/approve")
async def approve_approval_api(id: int, req: ApproveRequest, _auth = Depends(require_admin_or_manager)):
    import json
    import httpx
    import base64 as b64module
    
    # 1. Keresés a pending és rejected listában (hátha egy rejected-et hagynak jóvá utólag)
    approvals = db.get_approvals("pending") + db.get_approvals("rejected")
    target = next((a for a in approvals if a.get("id") == id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Piszkozat nem található")
        
    draft_json = target.get("ai_draft_response")
    if not draft_json:
        raise HTTPException(status_code=400, detail="Nincs érvényes piszkozat")
        
    try:
        draft = json.loads(draft_json)
    except:
        raise HTTPException(status_code=400, detail="Érvénytelen JSON piszkozat")
        
    channel = draft.get("channel", "").lower()
    final_text = req.modified_draft or draft.get("body", "")

    # Multi-channel kampány piszkozat kezelése
    drafts_to_send = []
    if draft.get("multi_channel") and draft.get("drafts"):
        for sub_draft in draft["drafts"]:
            drafts_to_send.append(sub_draft)
    else:
        drafts_to_send.append(draft)

    try:
        async with httpx.AsyncClient() as http_client:
            for send_draft in drafts_to_send:
                ch = send_draft.get("channel", "").lower()
                if draft.get("multi_channel"):
                    # Csatornánként SZERKESZTETT szöveg, ha a frontend küldte
                    # (korábban az összefűzött szerkesztés csendben elveszett —
                    # minden csatornán az eredeti body ment ki)
                    ch_name = send_draft.get("channel", "")
                    send_text = (req.modified_drafts or {}).get(ch_name) or send_draft.get("body", "")
                else:
                    send_text = final_text

                if ch == "email":
                    api_key = email_processor._get_brevo_api_key()

                    html_body = f'<div style="font-family: Arial, sans-serif;">{send_text.replace(chr(10), "<br>")}</div>'
                    if send_draft.get("event_id"):
                        html_body += email_processor.get_cancellation_html(send_draft.get("event_id"))

                    _sender = email_processor._get_sender()
                    email_payload = {
                        "sender": _sender,
                        "to": [{"email": send_draft.get("to_email"), "name": send_draft.get("to_name", "")}],
                        "subject": send_draft.get("subject", "Re:"),
                        "htmlContent": html_body,
                    }
                    # Threading: ha a draft hordozza a bejövő levél Message-ID-jét,
                    # a válasz a meglévő szálhoz fűződik az ügyfél levelezőjében
                    if send_draft.get("in_reply_to"):
                        email_payload["headers"] = {
                            "In-Reply-To": send_draft["in_reply_to"],
                            "References": send_draft["in_reply_to"],
                        }
                    resp = await http_client.post(
                        "https://api.brevo.com/v3/smtp/email",
                        headers={"api-key": api_key, "Content-Type": "application/json"},
                        json=email_payload,
                        timeout=20,
                    )
                    resp.raise_for_status()
                    print(f"[Approval] Email elküldve: {send_draft.get('to_email')}")
                    
                elif ch == "whatsapp":
                    # Per-tenant tokenek (a middleware a JWT-ből állítja a contextvart; env fallback)
                    approval_tenant = db.get_current_tenant()
                    wa_token = _meta_cred(approval_tenant, "whatsapp_token")
                    wa_phone_id = send_draft.get("phone_number_id") or _meta_cred(approval_tenant, "whatsapp_phone_id")
                    if not wa_token or not wa_phone_id:
                        raise Exception("Hiányzó WhatsApp token vagy Phone ID")
                    
                    resp = await http_client.post(
                        f"https://graph.facebook.com/v21.0/{wa_phone_id}/messages",
                        headers={"Authorization": f"Bearer {wa_token}"},
                        json={
                            "messaging_product": "whatsapp",
                            "to": send_draft.get("sender_id"),
                            "type": "text",
                            "text": {"body": send_text}
                        }
                    )
                    resp.raise_for_status()
                    
                elif ch in ["messenger", "instagram"]:
                    # Per-tenant tokenek (a middleware a JWT-ből állítja a contextvart; env fallback)
                    approval_tenant = db.get_current_tenant()
                    page_access_token = _meta_cred(approval_tenant, "meta_page_token")
                    if not page_access_token:
                        raise Exception("Hiányzó Meta oldal token")

                    if ch == "instagram":
                        # Instagram DM: use META_INSTAGRAM_TOKEN + graph.instagram.com/{ig_user_id}/messages
                        ig_token = _meta_cred(approval_tenant, "instagram_token")
                        ig_user_id = _meta_cred(approval_tenant, "instagram_user_id")
                        if not ig_token:
                            raise Exception("Hiányzó META_INSTAGRAM_TOKEN")
                        resp = await http_client.post(
                            f"https://graph.instagram.com/v21.0/{ig_user_id}/messages",
                            headers={
                                "Authorization": f"Bearer {ig_token}",
                                "Content-Type": "application/json"
                            },
                            json={
                                "recipient": {"id": send_draft.get("sender_id")},
                                "message": {"text": send_text}
                            }
                        )
                    else:
                        # Messenger: use graph.facebook.com/me/messages
                        resp = await http_client.post(
                            "https://graph.facebook.com/v21.0/me/messages",
                            headers={
                                "Authorization": f"Bearer {page_access_token}",
                                "Content-Type": "application/json"
                            },
                            json={
                                "recipient": {"id": send_draft.get("sender_id")},
                                "message": {"text": send_text}
                            }
                        )
                    
                    if resp.status_code != 200:
                        error_body = resp.text
                        print(f"[Approval] {ch.capitalize()} API hiba ({resp.status_code}): {error_body}")
                        resp.raise_for_status()
                    print(f"[Approval] {ch.capitalize()} elküldve: {send_draft.get('sender_id', '')[:10]}...")

                elif ch == "telefon":
                    # Telefonhívás indítása a jóváhagyott szöveggel mint AI script
                    from livekit import api as lk_api_module

                    lk_url    = os.getenv("LIVEKIT_URL")
                    lk_key    = os.getenv("LIVEKIT_API_KEY")
                    lk_secret = os.getenv("LIVEKIT_API_SECRET")
                    trunk_id  = os.getenv("SIP_OUTBOUND_TRUNK_ID", "ST_jgPctgJYZcAf")  # Telnyx HD Voice outbound

                    call_phone = send_draft.get("phone_number", "")
                    if not call_phone:
                        raise Exception("Hiányzó telefonszám a piszkozatban")
                    if not call_phone.startswith("+"):
                        call_phone = "+" + call_phone

                    call_room = f"call-out-{uuid.uuid4().hex[:8]}"
                    client_name = send_draft.get("to_name", "")
                    call_metadata = json.dumps({
                        "type": "outbound_script_call",
                        "script": send_text,
                        "client_name": client_name,
                        "call_note": send_draft.get("campaign_name", "Jóváhagyott kimenő hívás"),
                        "tenant_id": db.get_current_tenant() or "",
                    })

                    lk = lk_api_module.LiveKitAPI(url=lk_url, api_key=lk_key, api_secret=lk_secret)
                    await lk.room.create_room(
                        lk_api_module.CreateRoomRequest(
                            name=call_room,
                            empty_timeout=120,
                            metadata=call_metadata,
                        )
                    )
                    await lk.sip.create_sip_participant(
                        lk_api_module.CreateSIPParticipantRequest(
                            sip_trunk_id=trunk_id,
                            sip_call_to=call_phone,
                            room_name=call_room,
                            participant_identity="phone-caller",
                            participant_name=call_phone,
                            wait_until_answered=True,
                            krisp_enabled=True,
                        )
                    )
                    await lk.agent_dispatch.create_dispatch(
                        lk_api_module.CreateAgentDispatchRequest(
                            agent_name="dobozos-ai",
                            room=call_room,
                            metadata=call_metadata,
                        )
                    )
                    await lk.aclose()
                    print(f"[Approval] Telefon hívás indítva: {call_phone} (script: {send_text[:50]}...)")
                
    except Exception as e:
        print(f"[Approval Error] Hiba a kiküldéskor: {e}")
        # Küldés sikertelen, de az approve-ot azért mentsük el
        draft["body"] = final_text
        new_draft_json = json.dumps(draft)
        db.update_approval_status(id, "approved", new_draft=new_draft_json)
        return {"status": "warning", "message": f"Jóváhagyva, de a küldés sikertelen: {str(e)[:150]}"}
        
    # 2. Adatbázis frissítése — a szerkesztett szöveg(ek) elmentése
    if draft.get("multi_channel") and req.modified_drafts:
        for sd in draft.get("drafts", []):
            ch_name = sd.get("channel", "")
            if ch_name in req.modified_drafts:
                sd["body"] = req.modified_drafts[ch_name]
    draft["body"] = final_text
    new_draft_json = json.dumps(draft)
    success = db.update_approval_status(id, "approved", new_draft=new_draft_json)
    
    if success: return {"status": "success"}
    raise HTTPException(status_code=500, detail="Sikeres küldés, de adatbázis frissítés hibás")


@app.get("/admin/api/clinics")
def get_clinics_api(admin: dict = Depends(verify_jwt)):
    return db.get_clinics()

@app.post("/admin/api/clinics")
def save_clinics_api(clinics: list[dict], _auth = Depends(require_admin_or_manager)):
    success = db.save_clinics(clinics)
    if success: return {"status": "ok"}
    raise HTTPException(status_code=500, detail="Failed to save clinics")

@app.delete("/admin/api/clinics/{clinic_id}")
def delete_clinic_api(clinic_id: int, _auth = Depends(require_admin_or_manager)):
    """Delete a single clinic by ID."""
    try:
        if db.supabase:
            db.supabase.table("clinics").delete().eq("id", clinic_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/api/clients/{client_id}/profile-pic")
async def fetch_client_profile_pic(client_id: int, username: str = Depends(verify_jwt)):
    """Fetch and cache profile picture for a client with a messenger_id."""
    import json
    try:
        res = db.supabase.table("clients").select("custom_data").eq("id", client_id).single().execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Ügyfél nem található")
        
        cd = res.data.get("custom_data")
        if isinstance(cd, str):
            try: cd = json.loads(cd)
            except: cd = {}
        if not isinstance(cd, dict):
            cd = {}
        
        # Already has profile pic
        if cd.get("profile_pic_url"):
            return {"profile_pic_url": cd["profile_pic_url"]}
        
        messenger_id = cd.get("messenger_id")
        if not messenger_id:
            return {"profile_pic_url": None}
        
        # Determine channel
        channel = cd.get("forras_csatorna", "Messenger")
        profile = await fetch_meta_user_profile(messenger_id, channel)
        
        if profile and profile.get("profile_pic"):
            pic_url = profile["profile_pic"]
            # Save to DB
            cd["profile_pic_url"] = pic_url
            db.supabase.table("clients").update({"custom_data": cd}).eq("id", client_id).execute()
            return {"profile_pic_url": pic_url}
        
        return {"profile_pic_url": None}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[Profile Pic] Error: {e}")
        return {"profile_pic_url": None}

class ReminderSettingsRequest(BaseModel):
    reminder_enabled: bool
    reminder_hours: int
    reminder_template: str

@app.get('/admin/api/settings/reminder')
async def get_reminder_settings_endpoint(username: str = Depends(verify_jwt)):
    import database as db
    return db.get_reminder_settings()

@app.post('/admin/api/settings/reminder')
async def save_reminder_settings_endpoint(payload: ReminderSettingsRequest, _admin = Depends(require_admin)):
    import database as db
    success = db.update_reminder_settings(payload.reminder_enabled, payload.reminder_hours, payload.reminder_template)
    if success:
        return {'ok': True, 'message': 'Emlékeztető beállítások mentve.'}
    raise HTTPException(status_code=500, detail='Adatbázis hiba mentéskor')


# ═══════════════════════════════════════════════════════════════════════════════
# ESEMÉNYVEZÉRELT AUTOMATIZÁCIÓK API
# ═══════════════════════════════════════════════════════════════════════════════

@app.get('/admin/api/outbound_automations')
async def get_outbound_automations_endpoint(username: str = Depends(verify_jwt)):
    import database as db
    return db.get_outbound_automations()

@app.put('/admin/api/outbound_automations/{automation_id}')
async def update_outbound_automation_endpoint(automation_id: int, request: Request, _admin = Depends(require_admin)):
    import database as db
    data = await request.json()
    success = db.update_outbound_automation(automation_id, data)
    if success:
        return {'ok': True, 'message': 'Automatizáció frissítve.'}
    raise HTTPException(status_code=500, detail='Hiba az automatizáció mentésekor.')


# ═══════════════════════════════════════════════════════════════════════════════
# KIMENŐ KOMMUNIKÁCIÓ – KAMPÁNY API
# ═══════════════════════════════════════════════════════════════════════════════

class CampaignCreateRequest(BaseModel):
    name: str
    channels: list[str] = ["email"]
    client_ids: list[int]
    ai_instructions: str = ""
    subject: str = ""

@app.get("/admin/api/campaigns")
def get_campaigns_api(username: str = Depends(verify_jwt)):
    campaigns = db.get_campaigns()
    return {"campaigns": campaigns}

@app.post("/admin/api/campaigns")
def create_campaign_api(req: CampaignCreateRequest, _auth = Depends(require_admin_or_manager)):
    instructions = req.ai_instructions
    if req.subject:
        instructions = f"SUBJECT:{req.subject}|{instructions}"
    
    clean_channels = []
    for ch in req.channels:
        ch_clean = ch.lower().strip()
        if ch_clean in {"e-mail", "email"}:
            ch_clean = "email"
        clean_channels.append(ch_clean)
        
    campaign_id = db.create_campaign(
        name=req.name,
        channels=clean_channels,
        client_ids=req.client_ids,
        ai_instructions=instructions
    )
    if campaign_id:
        return {"status": "success", "id": campaign_id}
    raise HTTPException(status_code=500, detail="Kampány létrehozása sikertelen")

@app.post("/admin/api/campaigns/{campaign_id}/start")
async def start_campaign_api(campaign_id: int, _auth = Depends(require_admin_or_manager)):
    campaign = db.get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Kampány nem található")
    channels = campaign.get("channels", [campaign.get("channel", "email")])
    supported = {"email", "messenger", "telefon"}
    
    active_channels = []
    for ch in channels:
        ch_clean = ch.lower().strip()
        if ch_clean in {"e-mail", "email"}:
            ch_clean = "email"
        if ch_clean in supported:
            active_channels.append(ch_clean)
            
    if not active_channels:
        raise HTTPException(status_code=400, detail="Jelenleg csak email, messenger és telefon kampányok támogatottak")

    db.update_campaign_status(campaign_id, "Aktív")

    # Text-based channels (email, messenger) → draft generation
    text_channels = [ch for ch in active_channels if ch in {"email", "messenger"}]
    if text_channels:
        task = asyncio.create_task(_run_campaign(campaign, text_channels))
        background_tasks.add(task)
        task.add_done_callback(background_tasks.discard)

    # Phone channel → outbound SIP calls with campaign script
    if "telefon" in active_channels:
        phone_task = asyncio.create_task(_run_phone_campaign(campaign))
        background_tasks.add(phone_task)
        phone_task.add_done_callback(background_tasks.discard)

    channel_names = {"email": "Email", "messenger": "Messenger", "telefon": "Telefon (AI hívás)"}
    ch_str = ", ".join(channel_names.get(c, c) for c in active_channels)
    msg_parts = []
    if text_channels:
        msg_parts.append("emailek küldése megkezdődött")
    if "telefon" in active_channels:
        msg_parts.append("AI telefonhívások indulnak")
    return {"status": "success", "message": f"Kampány elindítva ({ch_str}) — {', '.join(msg_parts)}."}

@app.post("/admin/api/campaigns/{campaign_id}/stop")
def stop_campaign_api(campaign_id: int, _auth = Depends(require_admin_or_manager)):
    campaign = db.get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Kampány nem található")
    db.update_campaign_status(campaign_id, "Megállítva")
    return {"status": "success"}

@app.post("/admin/api/campaigns/{campaign_id}/close")
def close_campaign_api(campaign_id: int, _auth = Depends(require_admin_or_manager)):
    campaign = db.get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Kampány nem található")
    db.update_campaign_status(campaign_id, "Befejezett")
    return {"status": "success"}

@app.delete("/admin/api/campaigns/{campaign_id}")
def delete_campaign_api(campaign_id: int, _auth = Depends(require_admin_or_manager)):
    success = db.delete_campaign(campaign_id)
    if success:
        return {"status": "success"}
    raise HTTPException(status_code=500, detail="Törlés sikertelen")

@app.put("/admin/api/campaigns/{campaign_id}")
async def update_campaign_content_api(campaign_id: int, request: Request, _auth = Depends(require_admin_or_manager)):
    """EAISY-241 §1.6.2 — Kampány üzenet + subject szerkesztése (csak Tervezet/Ütemezett)."""
    data = await request.json()
    ai_instructions = (data.get("ai_instructions") or "").strip()
    subject = (data.get("subject") or "").strip()
    if not ai_instructions:
        raise HTTPException(status_code=400, detail="Az üzenet tartalma kötelező")

    campaign = db.get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Kampány nem található")
    # Csak nem-indult kampányok szerkeszthetők
    if campaign.get("status") not in ("Vázlat", "Ütemezett", "Tervezet", "Megállítva"):
        raise HTTPException(status_code=409, detail="Csak tervezet/ütemezett/megállított kampány szerkeszthető")

    success = db.update_campaign_content(campaign_id, ai_instructions, subject)
    if success:
        return {"status": "success"}
    raise HTTPException(status_code=500, detail="Mentés sikertelen")

@app.post("/admin/api/campaigns/{campaign_id}/schedule")
async def schedule_campaign_api(campaign_id: int, request: Request, _auth = Depends(require_admin_or_manager)):
    """Kampány ütemezése jövőbeli időpontra (campaigns tábla)."""
    data = await request.json()
    scheduled_at = data.get("scheduled_at")
    if not scheduled_at:
        raise HTTPException(status_code=400, detail="Hiányzó scheduled_at mező")
    
    campaign = db.get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Kampány nem található")
    
    # Update status to Ütemezett and store scheduled_at in ai_instructions as prefix
    db.update_campaign_status(campaign_id, "Ütemezett")
    existing_instructions = campaign.get("ai_instructions", "") or ""
    # Remove any existing SCHED: prefix
    if existing_instructions.startswith("SCHED:"):
        pipe_idx = existing_instructions.find("|")
        if pipe_idx >= 0:
            existing_instructions = existing_instructions[pipe_idx + 1:]
    new_instructions = f"SCHED:{scheduled_at}|{existing_instructions}"
    try:
        db.supabase.table("campaigns").update({"ai_instructions": new_instructions}).eq("id", campaign_id).execute()
        logger.info(f"[Schedule] Kampány #{campaign_id} ütemezve: {scheduled_at}")
    except Exception as e:
        logger.error(f"Schedule update error: {e}")
    
    return {"status": "success", "scheduled_at": scheduled_at}


@app.post("/admin/api/campaigns/generate_message")
async def generate_campaign_message(request: Request, _auth = Depends(require_admin_or_manager)):
    """AI kampány varázsló — üzenet generálás Gemini-vel."""
    data = await request.json()
    brief = data.get("brief", "")
    style = data.get("style", "barátságos")
    channel = data.get("channel", "email").lower().strip()
    if channel in {"e-mail", "email"}:
        channel = "email"
    
    if not brief:
        raise HTTPException(status_code=400, detail="A kampány brief megadása kötelező.")
    
    max_lengths = {"email": "1500 karakter", "whatsapp": "500 karakter", "sms": "160 karakter", "instagram": "300 karakter", "telefon": "500 karakter (beszélt szöveg)"}
    max_len = max_lengths.get(channel, "1000 karakter")
    
    style_instructions = {
        "hivatalos": "Hivatalos, professzionális hangvétel. Magázódás, formális stílus.",
        "barátságos": "Barátságos, közvetlen hangvétel. Tegezés, meleg tónus.",
        "akciós": "Figyelemfelkeltő, akciós hangvétel. Sürgős, limitált ajánlat érzés. Emojik használata.",
        "személyes": "Személyes, intim hangvétel. Mintha egy barát írna. Közvetlen, egyedi."
    }
    style_desc = style_instructions.get(style, style_instructions["barátságos"])
    
    google_key = os.getenv("GOOGLE_API_KEY")
    if not google_key:
        raise HTTPException(status_code=500, detail="Nincs GOOGLE_API_KEY beállítva.")
    
    try:
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=google_key)
        
        if channel == "telefon":
            prompt = f"""Írj egy telefonos hívás scriptet az alábbi paraméterek alapján:

KAMPÁNY BRIEF: {brief}

STÍLUS: {style_desc}

SZABÁLYOK:
- Ez egy AI telefonhívás scriptje — az AI agent ezt fogja ELMONDANI az ügyfélnek
- Természetes, beszélgetős stílusú legyen (NE legyen olvasott szöveg érzése)
- Kezdd köszönéssel és bemutatkozással
- Mondd el az ajánlatot/üzenetet röviden és érthetően (max 2-3 mondat egyszerre)
- Legyen benne reakció lehetőség (pl. "Mit gondol erről?" / "Érdekli?")
- Ha az ügyfél nem érdeklődik, udvariasan búcsúzz el
- Maximum {max_len}
- Magyarul

A válaszod KIZÁRÓLAG egyetlen valid JSON objektum legyen az alábbi formában, minden egyéb szöveg vagy markdown blokk nélkül:
{{
    "subject": "A hívás rövid témája (pl. Időpont egyeztetés)",
    "message": "A hívás script teljes szövege"
}}"""
        else:
            prompt = f"""Írj egy kampány üzenetet az alábbi paraméterek alapján:

KAMPÁNY BRIEF: {brief}

STÍLUS: {style_desc}

CSATORNA: {channel} (maximum {max_len})

SZABÁLYOK:
- Az üzenet legyen célratörő és hatásos
- NE használj HTML tageket, csak sima szöveget
- Listákhoz kötőjelet használj
- Az üzenet végén legyen egy call-to-action
- Ha {channel} == "sms", nagyon rövid legyen (max 160 karakter)
- Nem kell aláírás sor

A válaszod KIZÁRÓLAG egyetlen valid JSON objektum legyen az alábbi formában, minden egyéb szöveg vagy markdown blokk nélkül:
{{
    "subject": "Az üzenet tárgysora (pl. Különleges ajánlat Önnek!)",
    "message": "A kampányüzenet teljes szövege"
}}"""

        response = await client.aio.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0.7,
                response_mime_type="application/json"
            )
        )
        generated = response.text.strip()
        
        # Clean markdown code blocks if any
        if generated.startswith("```json"):
            generated = generated[7:]
        if generated.startswith("```"):
            generated = generated[3:]
        if generated.endswith("```"):
            generated = generated[:-3]
        generated = generated.strip()
        
        import json
        subject = ""
        message = generated
        try:
            data_json = json.loads(generated)
            subject = data_json.get("subject", "")
            message = data_json.get("message", "")
        except Exception as e:
            logger.error(f"Error parsing campaign message generation JSON: {e}")
            
        return {"subject": subject, "message": message, "style": style, "channel": channel}
    except Exception as e:
        logger.error(f"Campaign message generation error: {e}")
        raise HTTPException(status_code=500, detail=f"AI generálási hiba: {str(e)}")

@app.get("/admin/api/campaigns/{campaign_id}/clients")
def get_campaign_clients_api(campaign_id: int, username: str = Depends(verify_jwt)):
    campaign = db.get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Kampány nem található")
    client_ids = campaign.get("client_ids", [])
    clients = db.get_clients_by_ids(client_ids)
    result = []
    for c in clients:
        custom = c.get("custom_data", {})
        if isinstance(custom, str):
            try:
                custom = json.loads(custom)
            except:
                custom = {}
        result.append({
            "id": c.get("id"),
            "name": custom.get("name") or c.get("name", "Névtelen"),
            "email": custom.get("email") or c.get("email", "-"),
            "phone": custom.get("phone") or c.get("phone", "-"),
            "status": c.get("status", "-")
        })
    return {"clients": result, "campaign_name": campaign.get("name", "")}


async def _run_campaign(campaign: dict, active_channels: list[str]):
    """Háttérfolyamat: végigmegy a kampány ügyfelein és közvetlenül elküldi az emaileket."""
    import httpx
    import base64 as b64module

    campaign_id = campaign["id"]
    campaign_name = campaign["name"]
    ai_instructions = campaign.get("ai_instructions", "") or ""
    subject = campaign_name
    changed = True
    while changed:
        changed = False
        if ai_instructions.startswith("SCHED:"):
            pipe_idx = ai_instructions.find("|")
            if pipe_idx >= 0:
                ai_instructions = ai_instructions[pipe_idx + 1:]
                changed = True
        if ai_instructions.startswith("MODE:"):
            colon_idx = ai_instructions.find(":", 5)
            if colon_idx >= 0:
                ai_instructions = ai_instructions[colon_idx + 1:]
                changed = True
        if ai_instructions.startswith("SUBJECT:"):
            pipe_idx = ai_instructions.find("|")
            if pipe_idx >= 0:
                subject = ai_instructions[8:pipe_idx]
                ai_instructions = ai_instructions[pipe_idx + 1:]
                changed = True
    ai_instructions = ai_instructions.strip()
    client_ids = campaign.get("client_ids", [])

    if not ai_instructions:
        print(f"[Campaign] Nincs kampány tartalom, megszakítva: {campaign_name}")
        db.update_campaign_status(campaign_id, "Megállítva")
        return

    brevo_key = os.getenv("BREVO_API_KEY", "")
    api_key = brevo_key
    if brevo_key and not brevo_key.startswith("xkeysib-"):
        try:
            decoded = b64module.b64decode(brevo_key).decode()
            parsed = json.loads(decoded)
            api_key = parsed.get("api_key", brevo_key)
        except:
            pass

    if not api_key:
        print(f"[Campaign] BREVO_API_KEY hiányzik, kampány megszakítva: {campaign_name}")
        db.update_campaign_status(campaign_id, "Megállítva")
        return

    clients = db.get_clients_by_ids(client_ids)
    if not clients:
        print(f"[Campaign] Nem találhatók ügyfelek, kampány lezárva: {campaign_name}")
        db.update_campaign_status(campaign_id, "Befejezett", processed_count=0)
        return

    processed = 0
    failed = 0

    async with httpx.AsyncClient() as http_client:
        for client in clients:
            current = db.get_campaign(campaign_id)
            if not current or current.get("status") == "Megállítva":
                print(f"[Campaign] Kampány megállítva: {campaign_name}")
                return

            custom_data = client.get("custom_data", {})
            if isinstance(custom_data, str):
                try:
                    custom_data = json.loads(custom_data)
                except:
                    custom_data = {}

            client_email = custom_data.get("email") or client.get("email", "")
            client_name = custom_data.get("name") or client.get("name", "Névtelen")

            if not client_email or client_email == "-":
                print(f"[Campaign] Nincs email: {client_name}, kihagyva")
                continue

            # Personalize content
            body = ai_instructions.replace("{név}", client_name).replace("{name}", client_name)
            html_body = f'<div style="font-family: Arial, sans-serif;">{body.replace(chr(10), "<br>")}</div>'

            try:
                bi = db.get_business_info()
                _sender = {"name": bi.get("sender_name") or bi.get("practice_name", "Virtuális Asszisztens"), "email": bi.get("sender_email") or os.getenv("BREVO_SENDER_EMAIL", "noreply@example.com")}
                resp = await http_client.post(
                    "https://api.brevo.com/v3/smtp/email",
                    headers={"api-key": api_key, "Content-Type": "application/json"},
                    json={
                        "sender": _sender,
                        "to": [{"email": client_email, "name": client_name}],
                        "subject": subject,
                        "htmlContent": html_body,
                    },
                    timeout=20,
                )
                resp.raise_for_status()
                processed += 1
                print(f"[Campaign] Email elküldve: {client_name} <{client_email}>")
            except Exception as e:
                failed += 1
                print(f"[Campaign] Email küldési hiba ({client_name}): {e}")

            db.update_campaign_status(campaign_id, "Aktív", processed_count=processed)
            await asyncio.sleep(0.5)

    db.update_campaign_status(campaign_id, "Befejezett", processed_count=processed)
    print(f"[Campaign] Kampány befejezve: {campaign_name} – {processed} email elküldve, {failed} hiba")


async def _run_phone_campaign(campaign: dict):
    """Háttérfolyamat: végigmegy a kampány ügyfelein és AI telefonhívást indít mindegyiknek."""
    from livekit import api as lk_api_module

    campaign_id = campaign["id"]
    campaign_name = campaign["name"]
    ai_instructions = campaign.get("ai_instructions", "")
    client_ids = campaign.get("client_ids", [])

    lk_url    = os.getenv("LIVEKIT_URL")
    lk_key    = os.getenv("LIVEKIT_API_KEY")
    lk_secret = os.getenv("LIVEKIT_API_SECRET")
    trunk_id  = os.getenv("SIP_OUTBOUND_TRUNK_ID", "ST_jgPctgJYZcAf")  # Telnyx HD Voice outbound

    if not all([lk_url, lk_key, lk_secret]):
        print(f"[PhoneCampaign] LiveKit credentials hiányzik, kampány megszakítva: {campaign_name}")
        db.update_campaign_status(campaign_id, "Megállítva")
        return

    clients = db.get_clients_by_ids(client_ids)
    if not clients:
        print(f"[PhoneCampaign] Nem találhatók ügyfelek, kampány lezárva: {campaign_name}")
        db.update_campaign_status(campaign_id, "Befejezett", processed_count=0)
        return

    processed = 0
    failed = 0

    for client in clients:
        # Check if campaign was stopped
        current = db.get_campaign(campaign_id)
        if not current or current.get("status") == "Megállítva":
            print(f"[PhoneCampaign] Kampány megállítva: {campaign_name}")
            return

        custom_data = client.get("custom_data", {})
        if isinstance(custom_data, str):
            try:
                custom_data = json.loads(custom_data)
            except:
                custom_data = {}

        client_name = custom_data.get("name") or client.get("name", "Névtelen")
        client_phone = custom_data.get("phone") or client.get("phone", "")

        if not client_phone or client_phone == "-":
            print(f"[PhoneCampaign] Nincs telefonszám: {client_name}, kihagyva")
            continue

        # Normalize phone number
        phone = client_phone.strip()
        if not phone.startswith("+"):
            phone = "+" + phone

        # Build campaign metadata for the agent
        campaign_metadata = json.dumps({
            "type": "campaign_call",
            "campaign_id": campaign_id,
            "campaign_name": campaign_name,
            "client_name": client_name,
            "script": ai_instructions,
            "tenant_id": db.get_current_tenant() or "",
        })

        room_name = f"call-out-camp-{campaign_id}-{uuid.uuid4().hex[:6]}"

        try:
            lk = lk_api_module.LiveKitAPI(url=lk_url, api_key=lk_key, api_secret=lk_secret)

            # 1. Create room with campaign metadata
            await lk.room.create_room(
                lk_api_module.CreateRoomRequest(
                    name=room_name,
                    empty_timeout=120,
                    metadata=campaign_metadata,
                )
            )

            # 2. SIP call — wait for answer
            participant = await lk.sip.create_sip_participant(
                lk_api_module.CreateSIPParticipantRequest(
                    sip_trunk_id=trunk_id,
                    sip_call_to=phone,
                    room_name=room_name,
                    participant_identity="phone-caller",
                    participant_name=phone,
                    wait_until_answered=True,
                    krisp_enabled=True,
                )
            )

            # 3. Dispatch agent with campaign metadata
            await lk.agent_dispatch.create_dispatch(
                lk_api_module.CreateAgentDispatchRequest(
                    agent_name="dobozos-ai",
                    room=room_name,
                    metadata=campaign_metadata,
                )
            )

            await lk.aclose()

            # Log the interaction
            session_id = f"campaign_phone_{campaign_id}_{client['id']}"
            db.create_session(session_id=session_id, room_name=f"Kampány hívás: {campaign_name}", participant=client_name)
            db.log_interaction(
                type="telefon",
                topic=f"Kampány: {campaign_name}",
                summary=f"Kimenő AI telefonhívás – {client_name} ({phone})",
                result="Hívás indítva",
                tool_name="phone_campaign_worker",
                session_id=session_id,
                direction="outbound",
                funnel_stage="relevant",
                alert_tags=[],
                handover_reason=None,
            )

            processed += 1
            db.update_campaign_status(campaign_id, "Aktív", processed_count=processed)
            print(f"[PhoneCampaign] Hivas inditva ({processed}/{len(clients)}): {client_name} -> {phone}")

            # Wait between calls (15 sec) to avoid overwhelming + let calls finish
            await asyncio.sleep(15)

        except Exception as e:
            failed += 1
            print(f"[PhoneCampaign] Hivas sikertelen ({client_name} -> {phone}): {e}")
            await asyncio.sleep(3)

    db.update_campaign_status(campaign_id, "Befejezett", processed_count=processed)
    print(f"[PhoneCampaign] Kampany befejezve: {campaign_name} - {processed} hivas, {failed} sikertelen")


@app.get('/api/public/cancel')
async def public_cancel_appointment(token: str):
    import jwt as pyjwt
    import database as db
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        event_id = payload.get("event_id")
        if not event_id:
            raise ValueError("No event_id in token")
        
        # Mark client as cancelled if found
        event = db.get_calendar_event(event_id)
        if event:
            client = None
            email = event.get("attendee_email")
            if email and email != "-":
                client = db.find_client_by_contact(email=email)
            
            # Fallback to search by name
            if not client:
                name = event.get("attendee")
                if name and name != "-":
                    res = db.supabase.table("clients").select("*").ilike("name", f"%{name}%").order("id", desc=True).limit(1).execute()
                    if res.data:
                        client = res.data[0]

            # If still no client found, CREATE ONE to ensure Kanban and Toast work!
            if not client:
                name = event.get("attendee") or "Ismeretlen"
                email = event.get("attendee_email") or ""
                if name != "-" or email != "-":
                    new_client_id = db.add_client({
                        "name": name,
                        "email": email,
                        "phone": "",
                        "forras_csatorna": "Rendszer (Lemondás)"
                    }, status="lemondott")
                    if new_client_id:
                        client = {"id": new_client_id, "custom_data": {}}

            if client:
                custom_data = client.get("custom_data")
                if isinstance(custom_data, str):
                    try:
                        custom_data = json.loads(custom_data)
                    except:
                        custom_data = {}
                if not isinstance(custom_data, dict):
                    custom_data = {}
                
                custom_data["cancelled_viewed"] = False
                # Automatikus 'törölt időpont' tag hozzáadása
                existing_tags = custom_data.get("tags", [])
                if not isinstance(existing_tags, list): existing_tags = []
                if "törölt időpont" not in existing_tags:
                    existing_tags.append("törölt időpont")
                    custom_data["tags"] = existing_tags
                db.edit_client_details(client["id"], custom_data)
                db.update_client_status(client["id"], "lemondott")
                # Reset automation sent log so cancelled_no_rebook can fire again
                db.clear_automation_sent(client["id"])

        # Delete from calendar
        success = db.delete_calendar_event(event_id)
        
        if success:
            html = """
            <html>
            <head><title>Sikeres lemondás</title><meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body { font-family: 'Segoe UI', Arial, sans-serif; background: #f9fafb; text-align: center; padding: 50px 20px; color: #333; }
                .box { background: white; max-width: 500px; margin: 0 auto; padding: 40px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); }
                h1 { color: #10b981; margin-bottom: 20px; }
                p { font-size: 16px; line-height: 1.5; color: #6b7280; }
                .icon { font-size: 48px; margin-bottom: 20px; }
            </style>
            </head>
            <body>
                <div class="box">
                    <div class="icon">✅</div>
                    <h1>Időpont sikeresen lemondva!</h1>
                    <p>Köszönjük, hogy jelezte felénk. Az időpont törlésre került a naptárunkból.</p>
                </div>
            </body>
            </html>
            """
        else:
            html = """
            <html>
            <head><title>Hiba a lemondásnál</title><meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body { font-family: 'Segoe UI', Arial, sans-serif; background: #f9fafb; text-align: center; padding: 50px 20px; color: #333; }
                .box { background: white; max-width: 500px; margin: 0 auto; padding: 40px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); }
                h1 { color: #ef4444; margin-bottom: 20px; }
                p { font-size: 16px; line-height: 1.5; color: #6b7280; }
            </style>
            </head>
            <body>
                <div class="box">
                    <h1>Sikertelen lemondás</h1>
                    <p>Ezt az időpontot már korábban lemondták, vagy a hivatkozás érvénytelen.</p>
                </div>
            </body>
            </html>
            """
        return HTMLResponse(content=html, status_code=200)
    except Exception as e:
        html = f"""
        <html>
        <head><title>Érvénytelen link</title><meta charset="utf-8">
        <style>body {{ font-family: sans-serif; text-align: center; padding: 50px; color: #333; }}</style>
        </head>
        <body>
            <h1>Érvénytelen vagy lejárt link</h1>
            <p>Kérjük, vegye fel a kapcsolatot ügyfélszolgálatunkkal.</p>
        </body>
        </html>
        """
        return HTMLResponse(content=html, status_code=400)


def find_emails_in_html(html_text):
    import re
    email_pattern = r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+'
    found = re.findall(email_pattern, html_text)
    cleaned = []
    for email in found:
        email_lower = email.lower()
        if any(email_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg', '.css', '.js']):
            continue
        if email not in cleaned:
            cleaned.append(email)
    return cleaned


SOCIAL_PLATFORMS_30 = [
    'facebook.com', 'fb.com', 'instagram.com', 'linkedin.com', 'youtube.com', 'youtu.be',
    'tiktok.com', 'twitter.com', 'x.com', 'pinterest.com', 'pinterest.hu', 'snapchat.com',
    'reddit.com', 'tumblr.com', 'flickr.com', 'vimeo.com', 'twitch.tv', 'discord.gg',
    'discord.com', 'telegram.org', 't.me', 'whatsapp.com', 'wa.me', 'messenger.com',
    'skype.com', 'threads.net', 'medium.com', 'github.com', 'behance.net', 'dribbble.com',
    'viber.com', 'line.me', 'wechat.com', 'weibo.com', 'quora.com', 'patreon.com',
    'soundcloud.com'
]


def extract_social_links(html_content, base_url):
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin, urlparse
    soup = BeautifulSoup(html_content, "lxml")
    socials = {}
    
    for a in soup.find_all('a', href=True):
        href = a['href'].strip()
        if not href or href.startswith('#') or href.startswith('javascript:') or href.startswith('mailto:') or href.startswith('tel:'):
            continue
        full_url = urljoin(base_url, href)
        parsed_url = urlparse(full_url)
        domain = parsed_url.netloc.lower()
        
        for platform in SOCIAL_PLATFORMS_30:
            if platform in domain:
                key = platform.split('.')[0]
                if key == 'youtu':
                    key = 'youtube'
                elif key == 'fb':
                    key = 'facebook'
                elif key == 't':
                    key = 'telegram'
                elif key == 'wa':
                    key = 'whatsapp'
                
                if key not in socials:
                    socials[key] = full_url
                break
    return socials


def find_contact_links(base_url, html_content):
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin, urlparse
    
    soup = BeautifulSoup(html_content, "lxml")
    contact_links = []
    seen_urls = set()
    
    keywords = [
        # Hungarian
        'kapcsolat', 'elerhetoseg', 'elerhetosegek', 'impresszum', 'ceginfo', 'céginfó', 'cegadatok', 
        'cégadatok', 'rolunk', 'rólunk', 'cegunk', 'cégünk', 'bemutatkozas', 'bemutatkozás', 'adatok', 
        'szervezet', 'kapcsolati', 'elerhetosegi', 'terkep', 'térkép', 'ugyfelszolgalat', 'ügyfélszolgálat',
        # English
        'contact', 'contacts', 'about', 'about-us', 'aboutus', 'company', 'info', 'support', 'legal', 
        'terms', 'privacy', 'imprint',
        # German
        'kontakt', 'impressum', 'uber-uns', 'ueber-uns', 'firma', 'unternehmen',
        # Romanian
        'despre', 'despre-noi', 'companie', 'date-firma',
        # Slovak / Czech
        'o-nas', 'o-nás', 'firma', 'informacie', 'informace',
        # Polish
        'o-firmie',
        # French
        'a-propos', 'apropos', 'societe', 'société', 'entreprise', 'mentions-legales',
        # Spanish
        'contacto', 'contactenos', 'contáctenos', 'sobre-nosotros', 'quienes-somos', 'quiénes-somos', 'empresa', 'aviso-legal',
        # Italian
        'contatti', 'chi-siamo', 'azienda', 'societa', 'società', 'note-legali'
    ]
    
    parsed_base = urlparse(base_url)
    base_domain = parsed_base.netloc.lower()
    
    for a in soup.find_all('a', href=True):
        href = a['href'].strip()
        text = a.get_text(strip=True).lower()
        
        if not href or href.startswith('#') or href.startswith('javascript:'):
            continue
            
        full_url = urljoin(base_url, href)
        parsed_full = urlparse(full_url)
        
        if parsed_full.netloc.lower() != base_domain:
            continue
            
        url_lower = full_url.lower()
        matches = False
        for kw in keywords:
            if kw in url_lower or kw in text:
                matches = True
                break
                
        if matches and full_url not in seen_urls and full_url != base_url:
            seen_urls.add(full_url)
            contact_links.append(full_url)
            
    return contact_links[:5]


class ZomboScrapeRequest(BaseModel):
    url: str
    limit: int = 10

@app.post("/admin/api/zombo/scrape")
@app.post("/marketing/api/zombo/scrape")
async def api_zombo_scrape(req: ZomboScrapeRequest, username: str = Depends(verify_jwt)):
    """Streaming NDJSON endpoint for website SEO audit and visual tone analysis using multi-agent pipeline."""
    from fastapi.responses import StreamingResponse
    import httpx
    from bs4 import BeautifulSoup
    import re
    import colorsys
    from collections import Counter
    import json as _json
    import os
    import asyncio

    target_url = req.url.strip()
    if not target_url.startswith('http://') and not target_url.startswith('https://'):
        target_url = 'https://' + target_url

    # Strip query parameters and fragments to get the clean audit URL
    from urllib.parse import urlparse, urlunparse
    try:
        parsed = urlparse(target_url)
        target_url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, '', '', ''))
    except Exception as parse_err:
        print(f"[Zombo Scrape] Error parsing/cleaning URL: {parse_err}")

    async def generate():
        try:
            # ── Step 1: Fetch URL ──
            firecrawl_key = os.getenv("FIRECRAWL_API_KEY")
            crawled_pages = []
            combined_content = ""
            page_limit = getattr(req, "limit", 10)
            if page_limit < 1 or page_limit > 10:
                page_limit = 10 # clamp to 1-10
            
            html = ""
            ttfb_ms = 0.0
            
            if firecrawl_key:
                yield _json.dumps({"step": "progress", "message": f"Firecrawl: Honlap feltérképezése elindítva (limit: {page_limit} oldal)..."}) + "\n"
                headers = {
                    "Authorization": f"Bearer {firecrawl_key}",
                    "Content-Type": "application/json"
                }
                crawl_payload = {
                    "url": target_url,
                    "excludePaths": [
                        "/kosar", "/kosár", "/pénztár", "/penztar", "/checkout", "/cart",
                        "/tag/*", "/category/*", "*/feed*", "*/wp-json*", "*/search*"
                    ],
                    "maxDepth": 2,
                    "limit": page_limit,
                    "allowBackwardLinks": False,
                    "scrapeOptions": {
                        "formats": ["markdown", "html"]
                    }
                }
                
                try:
                    start_time = asyncio.get_event_loop().time()
                    async with httpx.AsyncClient(timeout=30.0) as crawl_client:
                        resp = await crawl_client.post("https://api.firecrawl.dev/v1/crawl", headers=headers, json=crawl_payload)
                        if resp.status_code == 200:
                            crawl_data = resp.json()
                            job_id = crawl_data.get("id")
                            if job_id:
                                # Polling loop
                                poll_url = f"https://api.firecrawl.dev/v1/crawl/{job_id}"
                                # Max poll time: 120s (24 iterations * 5 seconds)
                                for attempt in range(24):
                                    await asyncio.sleep(5.0)
                                    poll_resp = await crawl_client.get(poll_url, headers=headers)
                                    if poll_resp.status_code == 200:
                                        poll_data = poll_resp.json()
                                        status = poll_data.get("status")
                                        completed_count = poll_data.get("completed", 0)
                                        
                                        yield _json.dumps({"step": "progress", "message": f"Feltérképezés folyamatban: {completed_count} oldal beolvasva..."}) + "\n"
                                        
                                        if status == "completed":
                                            crawled_pages = poll_data.get("data", [])
                                            break
                                        elif status == "failed":
                                            print(f"[Firecrawl] Crawl job failed: {poll_data.get('error')}")
                                            break
                                    else:
                                        print(f"[Firecrawl] Polling error: {poll_resp.status_code}")
                                        break
                        else:
                            print(f"[Firecrawl] Crawl request failed status: {resp.status_code}")
                except Exception as ex:
                    print(f"[Firecrawl] Crawl exception: {ex}")
            
            # Extract HTML and calculate ttfb from crawled pages or fallback
            if crawled_pages:
                main_page = crawled_pages[0]
                html = main_page.get("html") or ""
                ttfb_ms = 150.0  # mock excellent ttfb from cache
                
                # Combine crawled pages markdown
                for p in crawled_pages:
                    p_url = p.get("metadata", {}).get("sourceURL") or p.get("url") or "Aloldal"
                    markdown = p.get("markdown") or ""
                    if len(markdown) > 6000:
                        markdown = markdown[:6000] + "\n[Tartalom csonkolva...]"
                    combined_content += f"\n\n--- PAGE: {p_url} ---\n{markdown}\n"
            
            if not html:
                # Fallback to standard fetch
                yield _json.dumps({"step": "progress", "message": "Kapcsolódás közvetlen letöltéssel (Firecrawl fallback)..."}) + "\n"
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                }
                start_time = asyncio.get_event_loop().time()
                async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
                    resp = await client.get(target_url, headers=headers)
                    resp.raise_for_status()
                    html = resp.text
                ttfb_ms = (asyncio.get_event_loop().time() - start_time) * 1000
            # If combined_content was empty, initialize it with homepage html
            allowed_tags = {'meta', 'title', 'link', 'a', 'img', 'h1', 'h2', 'h3', 'p', 'section', 'article', 'nav', 'header', 'footer', 'main', 'body', 'html', 'head'}
            if not combined_content:
                scraper_soup = BeautifulSoup(html, "lxml")
                for tag in scraper_soup(["script", "style", "svg", "noscript", "iframe", "form", "button", "input", "select", "option", "textarea"]):
                    tag.decompose()
                for tag in list(scraper_soup.find_all(True)):
                    if tag.name not in allowed_tags:
                        tag.unwrap()
                    else:
                        tag.attrs = {k: v for k, v in tag.attrs.items() if k in {'href', 'src', 'alt', 'name', 'content'}}
                for tag in list(scraper_soup.find_all(['p', 'a', 'section', 'article', 'nav'])):
                    if not tag.get_text(strip=True) and not tag.find_all(True) and not tag.get('href') and not tag.get('src'):
                        tag.decompose()
                combined_content = f"--- HOMEPAGE: {target_url} ---\n{str(scraper_soup)[:15000]}\n"

            # Search and fetch contact subpages separately
            contact_urls = []
            accumulated_socials = {}
            homepage_emails = []
            
            try:
                contact_urls = find_contact_links(target_url, html)
                accumulated_socials.update(extract_social_links(html, target_url))
                homepage_emails = find_emails_in_html(html)
            except Exception as e_find:
                print(f"[Zombo Scrape] Error scanning homepage contact/social details: {e_find}")

            if contact_urls:
                yield _json.dumps({"step": "progress", "message": f"Elérhetőségek aloldalainak letöltése ({len(contact_urls)} oldal)..."}) + "\n"
                headers_fetch = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                }
                
                email_subpages_content = []
                normal_subpages_content = []
                
                async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
                    for c_url in contact_urls:
                        try:
                            c_resp = await client.get(c_url, headers=headers_fetch)
                            if c_resp.status_code == 200:
                                c_html_raw = c_resp.text
                                c_emails = find_emails_in_html(c_html_raw)
                                
                                try:
                                    sub_socials = extract_social_links(c_html_raw, c_url)
                                    for k, v in sub_socials.items():
                                        if k not in accumulated_socials:
                                            accumulated_socials[k] = v
                                except Exception:
                                    pass
                                    
                                c_soup = BeautifulSoup(c_html_raw, "lxml")
                                for tag in c_soup(["script", "style", "svg", "noscript", "iframe", "form", "button", "input", "select", "option", "textarea"]):
                                    tag.decompose()
                                for tag in list(c_soup.find_all(True)):
                                    if tag.name not in allowed_tags:
                                        tag.unwrap()
                                    else:
                                        tag.attrs = {k: v for k, v in tag.attrs.items() if k in {'href', 'src', 'alt', 'name', 'content'}}
                                for tag in list(c_soup.find_all(['p', 'a', 'section', 'article', 'nav'])):
                                    if not tag.get_text(strip=True) and not tag.find_all(True) and not tag.get('href') and not tag.get('src'):
                                        tag.decompose()
                                c_html = str(c_soup)[:12000]
                                
                                if c_emails:
                                    email_subpages_content.append(f"--- CONTACT/COMPANY SUBPAGE (CONTAINS EMAILS: {', '.join(c_emails)}): {c_url} ---\n{c_html}\n")
                                else:
                                    normal_subpages_content.append(f"--- CONTACT/COMPANY SUBPAGE: {c_url} ---\n{c_html}\n")
                        except Exception as e_c:
                            print(f"[Zombo Scrape] Error fetching contact subpage {c_url}: {e_c}")
                
                subpages_combined = "\n\n".join(email_subpages_content + normal_subpages_content)
                if subpages_combined:
                    combined_content = f"{subpages_combined}\n\n" + combined_content

            # ── Step 2: Parse raw metrics & colors in Python ──
            yield _json.dumps({"step": "progress", "message": "HTML struktúra elemzése..."}) + "\n"
            soup = BeautifulSoup(html, "lxml")
            
            # Simple python counting
            images = soup.find_all('img')
            total_images = len(images)
            missing_alt = sum(1 for img in images if not img.get('alt'))

            links = soup.find_all('a', href=True)
            total_links = len(links)
            internal_links = 0
            external_links = 0
            for link in links:
                href = link['href']
                if href.startswith('http://') or href.startswith('https://'):
                    external_links += 1
                else:
                    internal_links += 1

            # Schema markup detection
            has_schema = bool(soup.find_all('script', type='application/ld+json'))

            # Language detection
            html_tag = soup.find('html')
            lang_val = html_tag.get('lang') if html_tag else None
            has_lang = bool(lang_val)

            # Robots.txt and sitemap.xml detection
            from urllib.parse import urlparse
            has_robots = False
            has_sitemap = False
            try:
                parsed_target = urlparse(target_url)
                root_url = f"{parsed_target.scheme}://{parsed_target.netloc}"
                async with httpx.AsyncClient(timeout=5.0) as check_client:
                    robots_resp = await check_client.get(f"{root_url}/robots.txt")
                    if robots_resp.status_code == 200:
                        has_robots = True
                        if "sitemap" in robots_resp.text.lower():
                            has_sitemap = True
            except Exception:
                pass

            if not has_sitemap:
                try:
                    parsed_target = urlparse(target_url)
                    root_url = f"{parsed_target.scheme}://{parsed_target.netloc}"
                    async with httpx.AsyncClient(timeout=5.0) as check_client:
                        sitemap_resp = await check_client.get(f"{root_url}/sitemap.xml")
                        if sitemap_resp.status_code == 200:
                            has_sitemap = True
                except Exception:
                    pass

            # Word count
            text_soup = BeautifulSoup(html, "lxml")
            for tag in text_soup(["script", "style", "nav", "footer", "header"]):
                tag.decompose()
            clean_text = text_soup.get_text(separator=" ")
            clean_text = re.sub(r'\s+', ' ', clean_text).strip()
            word_count = len(clean_text.split())

            # Color analysis
            styles_text = []
            for tag in soup.find_all(style=True):
                styles_text.append(tag['style'])
            for tag in soup.find_all('style'):
                styles_text.append(tag.string or '')
            full_style_content = " ".join(styles_text)

            def hex_to_rgb(hex_str):
                hex_str = hex_str.lstrip('#')
                if len(hex_str) == 3:
                    hex_str = ''.join([c*2 for c in hex_str])
                if len(hex_str) == 6:
                    try:
                        return int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
                    except ValueError:
                        return None
                return None

            def parse_rgb(rgb_str):
                nums = re.findall(r'\d+', rgb_str)
                if len(nums) >= 3:
                    return int(nums[0]), int(nums[1]), int(nums[2])
                return None

            def get_hue_and_warmth(r, g, b):
                h, l, s = colorsys.rgb_to_hls(r/255.0, g/255.0, b/255.0)
                hue_deg = h * 360.0
                if s < 0.1 or l < 0.1 or l > 0.9:
                    return hue_deg, 'neutral'
                elif (0 <= hue_deg <= 60) or (300 <= hue_deg <= 360):
                    return hue_deg, 'warm'
                elif (120 <= hue_deg <= 240):
                    return hue_deg, 'cool'
                else:
                    return hue_deg, 'neutral'

            standardized_colors = []
            for h in re.findall(r'#(?:[0-9a-fA-F]{3}){1,2}\b', full_style_content):
                h = h.lower()
                if len(h) == 4:
                    h = '#' + ''.join([c*2 for c in h[1:]])
                standardized_colors.append(h)

            for rgb in re.findall(r'rgba?\(\s*\d+\s*,\s*\d+\s*,\s*\d+\s*(?:,\s*[\d\.]+)?\s*\)', full_style_content):
                rgb_val = parse_rgb(rgb)
                if rgb_val:
                    r, g, b = rgb_val
                    hex_val = f"#{r:02x}{g:02x}{b:02x}"
                    standardized_colors.append(hex_val)

            color_counts = Counter(standardized_colors)
            total_cols = len(standardized_colors)
            top_colors = [item[0] for item in color_counts.most_common(5)]
            if not top_colors:
                top_colors = ["#3b82f6", "#10b981", "#ef4444", "#f59e0b", "#6b7280"]

            top_colors_detail = []
            if total_cols > 0:
                for col, count in color_counts.most_common(5):
                    top_colors_detail.append({
                        "hex": col,
                        "pct": round((count / total_cols) * 100, 1)
                    })
            else:
                for col in top_colors:
                    top_colors_detail.append({
                        "hex": col,
                        "pct": 20.0
                    })

            warm_count = 0
            cool_count = 0
            neutral_count = 0
            for col in standardized_colors:
                rgb = hex_to_rgb(col)
                if rgb:
                    _, warmth = get_hue_and_warmth(*rgb)
                    if warmth == 'warm':
                        warm_count += 1
                    elif warmth == 'cool':
                        cool_count += 1
                    else:
                        neutral_count += 1

            warm_pct = round((warm_count / total_cols) * 100, 1) if total_cols > 0 else 40.0
            cool_pct = round((cool_count / total_cols) * 100, 1) if total_cols > 0 else 40.0
            neutral_pct = round((neutral_count / total_cols) * 100, 1) if total_cols > 0 else 20.0

            if warm_pct > cool_pct + 10:
                visual_tone = "Meleg és Barátságos"
            elif cool_pct > warm_pct + 10:
                visual_tone = "Hideg és Professzionális"
            else:
                visual_tone = "Kiegyensúlyozott / Neutrális"

            # ── Init Gemini Client ──
            google_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
            gemini_client = None
            if google_key:
                from google import genai
                from google.genai import types
                gemini_client = genai.Client(api_key=google_key)

            # Default fallback analysis values
            ai_analysis = {
                "score": 60,
                "deductions_detail": [],
                "business_category": "Általános",
                "tone": "Informatív",
                "summary": "Nem sikerült részletes leírást generálni a tartalomról.",
                "seo_advice": "Javasoljuk kulcsszó-kutatás végzését és a címsorok optimalizálását.",
                "global_improvements": ["Optimalizáld a képek alt leíróit.", "Ügyelj a címsorok hierarchiájára."],
                "detected_posts": [],
                "word_style_analysis": "Nem áll rendelkezésre részletes szövegezési stíluselemzés.",
                "visual_style_description": "Nem áll rendelkezésre részletes vizuális elemzés."
            }
            scraped_data_json_str = "{}"

            # ── Agent 1: Scraper Agent ──
            if gemini_client:
                yield _json.dumps({"step": "progress", "message": "Scraper Agent: Tartalom kinyerése..."}) + "\n"
                
                # Clean HTML to save tokens and keep only semantic tags
                scraper_soup = BeautifulSoup(html, "lxml")
                for tag in scraper_soup(["script", "style", "svg", "noscript", "iframe", "form", "button", "input", "select", "option", "textarea"]):
                    tag.decompose()
                allowed_tags = {'meta', 'title', 'link', 'a', 'img', 'h1', 'h2', 'h3', 'p', 'section', 'article', 'nav', 'header', 'footer', 'main', 'body', 'html', 'head'}
                for tag in list(scraper_soup.find_all(True)):
                    if tag.name not in allowed_tags:
                        tag.unwrap()
                essential_attrs = {'href', 'src', 'alt', 'name', 'content'}
                for tag in scraper_soup.find_all(True):
                    tag.attrs = {k: v for k, v in tag.attrs.items() if k in essential_attrs}
                for tag in list(scraper_soup.find_all(['p', 'a', 'section', 'article', 'nav'])):
                    if not tag.get_text(strip=True) and not tag.find_all(True) and not tag.get('href') and not tag.get('src'):
                        tag.decompose()
                clean_html_for_scraper = str(scraper_soup)[:15000]

                scraper_prompt = f"""
                You are the Page Scraper Agent. Your task is to analyze the following webpage content (HTML or markdown of crawled pages) and extract structured details into a JSON object.
                Limit the number of extracted images and links to a maximum of 20 elements each (focusing on the most prominent ones like header, hero, navigation, or main articles).
                
                Also search through all content to identify products or services offered and contact information.

                The output JSON MUST follow exactly this format:
                {{
                  "metadata": {{
                    "title": "Page title tag value or empty string",
                    "description": "Meta description content or empty string",
                    "keywords": "Meta keywords content if any, or null",
                    "viewport": "Viewport meta tag value if any, or null",
                    "canonical": "Canonical link href value if any, or null",
                    "lang": "Html lang attribute value if any, or null",
                    "is_https": true/false
                  }},
                  "headings": {{
                    "h1": ["H1 text 1", "H1 text 2"],
                    "h2": ["H2 text 1", ...],
                    "h3": ["H3 text 1", ...]
                  }},
                  "images": [
                    {{"src": "resolved absolute image URL", "alt": "image alt text if any, or empty string"}}
                  ],
                  "links": [
                    {{"href": "resolved absolute link URL", "text": "link text label", "is_external": true/false}}
                  ],
                  "sections": [
                    {{"id": "section-id-or-selector", "title": "Logical section title", "type": "Hero | Navigation | Articles | Sidebar | Footer | Contact | Services"}}
                  ],
                  "products": [
                    {{
                      "name": "Product or service name",
                      "price": "Price with currency if found (e.g. '15 000 Ft' or '$99'), or 'N/A'",
                      "brand": "Brand or manufacturer if found, or 'N/A'",
                      "description": "Short description of product/service",
                      "page_url": "Absolute URL of the page containing this product/service"
                    }}
                  ]
                }}
                Ensure all relative URLs in images, links, and products are made absolute using the base URL: {target_url}
                
                Here is the webpage content:
                {combined_content if combined_content else clean_html_for_scraper}
                """
                try:
                    ai_resp = await gemini_client.aio.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=scraper_prompt,
                        config=types.GenerateContentConfig(response_mime_type="application/json")
                    )
                    if ai_resp and ai_resp.text:
                        scraped_data_json_str = ai_resp.text.strip()
                except Exception as e:
                    print(f"[Zombo Scrape] Scraper Agent error: {e}")

            # ── Phase 2: Parallel Specialist Evaluations (SEO, Marketing, Images) ──
            marketing_analysis = {
                "marketing_score": 60,
                "value_proposition_evaluation": "Nem sikerült kiértékelni az értékajánlatot.",
                "frameworks_analysis": {
                    "pas_alignment": "Nem áll rendelkezésre PAS elemzés.",
                    "aida_alignment": "Nem áll rendelkezésre AIDA elemzés."
                },
                "cta_evaluation": "Nem sikerült kiértékelni a CTA-kat.",
                "credibility_evaluation": "Nem sikerült kiértékelni a hitelességet.",
                "copy_recommendations": ["Frissítse a főcímet előny-orientáltra.", "Használjon egyértelmű CTA gombokat."]
            }
            images_analysis = []
            product_analysis = []
            contact_analysis = {
                "emails": [],
                "phone_numbers": [],
                "addresses": [],
                "company_name": None,
                "tax_number": None,
                "registration_number": None
            }
            for platform in SOCIAL_PLATFORMS_30:
                key = platform.split('.')[0]
                if key == 'youtu': key = 'youtube'
                elif key == 'fb': key = 'facebook'
                elif key == 't': key = 'telegram'
                elif key == 'wa': key = 'whatsapp'
                contact_analysis[key] = None

            if gemini_client and scraped_data_json_str != "{}":
                yield _json.dumps({"step": "progress", "message": "Specialista ágensek indítása párhuzamosan..."}) + "\n"

                # 1. SEO Agent Task
                async def run_seo_agent():
                    seo_prompt = f"""
                    You are the SEO Specialist Agent, an expert search engine optimization auditor.
                    Your task is to analyze the following webpage metadata JSON extracted by the Scraper Agent.
                    Perform a comprehensive SEO audit based on the scraped elements: Title tag, Description tag, H1-H3 headings count and hierarchy, images alt attributes, links (broken/external/internal), viewport configuration, and HTTPS security.
                    
                    Calculate the final SEO audit score by starting from 0 points and adding/subtracting points for each criterion below:
                    1. Meta Title (Max +15 points):
                       - Good: +15 points if title length is 40-70 characters and matches the content.
                       - Bad: -5 points if title length is non-optimal (too short/long), or -20 points if missing entirely.
                    2. Meta Description (Max +15 points):
                       - Good: +15 points if description is 110-160 characters.
                       - Bad: -5 points if description length is non-optimal, or -15 points if missing entirely.
                    3. H1 Headings (Max +15 points):
                       - Good: +15 points if there is exactly 1 H1 heading.
                       - Bad: -15 points if missing H1, or -10 points if there are multiple H1s.
                    4. Heading Hierarchy H2-H3 (Max +10 points):
                       - Good: +10 points if headings have a logical structure with H2 and H3 tags present.
                       - Bad: -5 points if hierarchy is broken (e.g. H3 before H2, or missing H2/H3 entirely).
                    5. HTTPS Security (Max +15 points):
                       - Good: +15 points if connection is HTTPS.
                       - Bad: -15 points if connection is HTTP.
                    6. Mobile Friendly Viewport (Max +15 points):
                       - Good: +15 points if viewport tag is configured.
                       - Bad: -15 points if viewport tag is missing.
                    7. Canonical Link (Max +5 points):
                       - Good: +5 points if canonical tag is present.
                       - Bad: -5 points if canonical tag is missing.
                    8. Content Word Count (Max +10 points):
                       - Good: +10 points if word count >= 600 words, or +5 points if word count >= 300 words.
                       - Bad: -10 points if word count < 300 words (thin content).
                    9. Image Alt Attributes (Max +10 points):
                       - Good: +10 points if all images have alt attributes.
                       - Bad: -5 to -15 points (proportionally based on missing alt ratio) if images are missing alt attributes.
                    10. Server Response Time / TTFB (Max +10 points):
                        - Good: +10 points if TTFB < 300ms, or +5 points if TTFB < 600ms.
                        - Bad: -10 points if TTFB > 600ms, or -15 points if TTFB > 1000ms.
                    11. Robots.txt and Sitemap.xml (Max +5 points):
                        - Good: +5 points if both robots.txt and sitemap.xml are present on the domain.
                        - Bad: -5 points if either is missing.
                    12. Structured Data / Schema Markup (Max +5 points):
                        - Good: +5 points if JSON-LD structured data/schema markup is found on the page.
                        - Bad: -5 points if missing.
                    13. Language Declaration (Max +5 points):
                        - Good: +5 points if the <html> tag has a valid 'lang' attribute.
                        - Bad: -5 points if missing or empty.
                    
                    The final score should be exactly: sum(points of all evaluated criteria). Minimum score is 0.
                    
                    IMPORTANT: For each item in "deductions_detail", the "reason" field MUST explicitly cite the exact text value being evaluated in quotes (e.g. the actual meta title text, the actual meta description text, the actual H1 text content, the actual sitemap status, or sitemap/robots URLs, etc.) so that the user can immediately see the raw data that was processed.
                    Example: "A meta cím hossza nem optimális (19 karakter): 'Apple (Magyarország)'."
                    
                    Provide your response strictly as a JSON response in Hungarian using exactly the keys in the example:
                    {{
                      "score": 55,
                      "deductions_detail": [
                        {{
                          "criterion": "Meta Cím hossza",
                          "status": "good",
                          "points": 15,
                          "reason": "A title tag hossza megfelelő (55 karakter).",
                          "recommendation": "Nem szükséges beavatkozás."
                        }},
                        {{
                          "criterion": "Helyettesítő szöveg (alt)",
                          "status": "bad",
                          "points": -10,
                          "reason": "90 képnél hiányzik a helyettesítő (alt) szöveg.",
                          "recommendation": "Adjon meg rövid, kifejező alt attribútumot minden képnél az akadálymentesség javításához."
                        }}
                      ],
                      "business_category": "A 1-3 word business category/profile (e.g., Online Hírportál, Webáruház)",
                      "tone": "A 1-3 word tone description (e.g., Kritikus, Tényfeltáró, Professzionális)",
                      "summary": "A 2-3 sentence summary of what the business/page is about.",
                      "seo_advice": "A detailed paragraph of advice specifically for keyword optimization and content improvement based on the text.",
                      "global_improvements": [
                        "Összefoglaló tanács 1 a weboldal keresőoptimalizálásának globális javításához",
                        "Összefoglaló tanács 2...",
                        "Összefoglaló tanács 3..."
                      ],
                      "detected_posts": [
                        {{
                          "title": "Title of the post or service section found",
                          "placement": "Where it is located (e.g., Főhír, Oldalsáv, Kiemelt)",
                          "inferred_popularity": "Estimated popularity level (e.g., Kiemelkedő, Magas, Átlagos)",
                          "words": ["keyword1", "keyword2"],
                          "style": "Tone of this headline (e.g., Kritikus, Szenzációs, Elemző)",
                          "category": "Topic (e.g., Belföld, Politika, Gazdaság)"
                        }}
                      ],
                      "word_style_analysis": "A 1-2 sentence analysis of what tone and writing style is inferred from the overall word usage of this page.",
                      "visual_style_description": "A 1-2 sentence analysis of the overall visual style, layout, and brand mood."
                    }}
                    
                    Here is the Scraper Agent's JSON:
                    {scraped_data_json_str}
                    
                    Additional Raw Metrics for your context:
                    - Total Raw Images found: {total_images}
                    - Images Missing Alt text: {missing_alt}
                    - Total Links: {total_links} (Internal: {internal_links}, External: {external_links})
                    - Server Response Time (TTFB): {int(ttfb_ms)}ms
                    - Word Count: {word_count}
                    - JSON-LD Structured Data / Schema Markup Present: {has_schema}
                    - Robots.txt Present: {has_robots}
                    - Sitemap.xml Present: {has_sitemap}
                    - Language Attribute Declared: {has_lang} (Value: {lang_val if lang_val else 'None'})
                    """
                    try:
                        ai_resp = await gemini_client.aio.models.generate_content(
                            model='gemini-2.5-flash',
                            contents=seo_prompt,
                            config=types.GenerateContentConfig(response_mime_type="application/json")
                        )
                        if ai_resp and ai_resp.text:
                            parsed = _json.loads(ai_resp.text)
                            for k in ai_analysis.keys():
                                if k in parsed:
                                    ai_analysis[k] = parsed[k]
                    except Exception as e:
                        print(f"[Zombo Scrape] SEO Specialist Agent error: {e}")

                # 2. Copywriting & Marketing Agent Task
                async def run_marketing_agent():
                    marketing_prompt = f"""
                    You are the Copywriting & Marketing Specialist Agent, an expert in web copywriting, brand messaging, and conversion rate optimization (CRO).
                    Your task is to analyze the following webpage metadata JSON extracted by the Scraper Agent.
                    Perform a comprehensive copywriting and marketing audit based on the content, headings, products/services descriptions, and overall structure.
                    
                    Calculate the final Marketing/Copy score by starting from 0 points and adding/subtracting points for each criterion below:
                    1. Value Proposition & Hook (Max 20 points):
                       - Good: +20 points if the main heading (H1) and introductory paragraphs present a clear, unique, and benefit-driven value proposition.
                       - Bad: -10 points if the value proposition is missing, generic (e.g. 'Üdvözöljük a honlapunkon'), or confusing.
                    2. Copywriting Frameworks - PAS & AIDA (Max 25 points):
                       - Good: +25 points if the page copy is clearly structured around AIDA (Attention, Interest, Desire, Action) or PAS (Problem, Agitate, Solve) frameworks.
                       - Bad: -10 points if the copy lacks structure, is purely descriptive without addressing customer pain points or benefits.
                    3. Call to Action - CTA (Max 20 points):
                       - Good: +20 points if buttons and links use persuasive, action-oriented, and specific text (e.g. 'Kérem az ingyenes konzultációt') instead of passive text (e.g. 'Küldés', 'Kattintson ide').
                       - Bad: -10 points if CTAs are weak, missing, or badly placed.
                    4. Social Proof & Credibility (Max 20 points):
                       - Good: +20 points if the page prominently features trust indicators such as client reviews, testimonials, case studies, partner logos, certifications, or satisfaction guarantees.
                       - Bad: -10 points if no trust factors or credibility signals are found.
                    5. Readability & Tone Consistency (Max 15 points):
                       - Good: +15 points if the text is easy to read, uses paragraphs, has appropriate length, and maintains a consistent professional/friendly tone suited to the target audience.
                       - Bad: -5 points if it is flooded with expert jargon, or has wall-of-text paragraphs.
                    
                    The final score should be exactly: sum(points of all evaluated criteria). Minimum score is 0.
                    
                    IMPORTANT: In your evaluations ("value_proposition_evaluation", "frameworks_analysis", "cta_evaluation", "credibility_evaluation"), you MUST always quote the exact texts, headlines, or button labels you analyzed (e.g. the actual H1 text or the exact CTA button text like "Bővebben", "Vásárlás") so the user can verify the raw copy being audited.
                    
                    Provide your response strictly as a JSON response in Hungarian using exactly this format:
                    {{
                      "marketing_score": 75,
                      "value_proposition_evaluation": "Rövid, 1-2 mondatos szöveges értékelés a fő értékajánlatról és a hook hatásosságáról.",
                      "frameworks_analysis": {{
                        "pas_alignment": "Hogyan érvényesül vagy miért hiányzik a PAS (Probléma, fokozás, megoldás) szerkezet a szövegben.",
                        "aida_alignment": "Hogyan érvényesül vagy miért hiányzik az AIDA (Figyelem, érdeklődés, vágy, cselekvés) folyamat."
                      }},
                      "cta_evaluation": "Értékelés a cselekvésre ösztönző gombok és linkek meggyőző erejéről és elhelyezéséről.",
                      "credibility_evaluation": "Értékelés az oldalon lévő társadalmi bizonyítékokról (vélemények, partnerek, garanciák) és hitelességről.",
                      "copy_recommendations": [
                        "Konkrét, gyakorlati javaslat 1 a szöveg meggyőző erejének növelésére",
                        "Konkrét, gyakorlati javaslat 2 a CTA-k javítására",
                        "Konkrét, gyakorlati javaslat 3..."
                      ]
                    }}
                    
                    Here is the Scraper Agent's JSON:
                    {scraped_data_json_str}
                    """
                    try:
                        ai_resp = await gemini_client.aio.models.generate_content(
                            model='gemini-2.5-flash',
                            contents=marketing_prompt,
                            config=types.GenerateContentConfig(response_mime_type="application/json")
                        )
                        if ai_resp and ai_resp.text:
                            parsed = _json.loads(ai_resp.text)
                            for k in marketing_analysis.keys():
                                if k in parsed:
                                    marketing_analysis[k] = parsed[k]
                    except Exception as e:
                        print(f"[Zombo Scrape] Copywriting & Marketing Agent error: {e}")

                # 3. Image Analysis Task
                async def run_images_analysis():
                    nonlocal images_analysis
                    img_tags = soup.find_all('img')
                    from urllib.parse import urljoin
                    img_urls = []
                    for img in img_tags:
                        src = img.get('src')
                        if src and not src.startswith('data:') and len(src) > 5:
                            abs_url = urljoin(target_url, src)
                            img_urls.append((abs_url, img.get('alt', '').strip()))

                    seen = set()
                    valid_images = []
                    for url, alt in img_urls:
                        lower_url = url.lower()
                        if any(p in lower_url for p in ['logo', 'icon', 'sprite', 'avatar', 'loader', 'spacer', 'pixel', 'facebook', 'instagram', 'twitter', 'linkedin', 'youtube', 'tiktok']):
                            continue
                        if url not in seen:
                            seen.add(url)
                            valid_images.append((url, alt))
                            if len(valid_images) == 3:
                                break
                    if len(valid_images) < 3:
                        for url, alt in img_urls:
                            if url not in seen:
                                seen.add(url)
                                valid_images.append((url, alt))
                                if len(valid_images) == 3:
                                    break

                    if valid_images:
                        from PIL import Image
                        import io

                        async def analyze_single_image(client, url, alt):
                            try:
                                async with httpx.AsyncClient(timeout=3.0) as img_client:
                                    resp = await img_client.get(url)
                                    if resp.status_code == 200:
                                        img_bytes = resp.content
                                        img_pil = Image.open(io.BytesIO(img_bytes))
                                        img_pil.thumbnail((512, 512))
                                        
                                        prompt = """
                                        Írd le 1 mondatban magyarul, hogy mi látható ezen a képen és milyen a stílusa (pl. fotó, grafika, modern, minimalista).
                                        Valamint határozd meg a képen leginkább domináló 3 színt és add vissza őket hex kód formájában (pl. '#ff0000').
                                        
                                        A választ JSON formátumban add meg a következő kulcsokkal:
                                        {
                                          "description": "A kép leírása magyarul",
                                          "dominant_colors": ["#hex1", "#hex2", "#hex3"]
                                        }
                                        """
                                        ai_resp = await client.aio.models.generate_content(
                                            model='gemini-2.5-flash',
                                            contents=[img_pil, prompt],
                                            config=types.GenerateContentConfig(response_mime_type="application/json")
                                        )
                                        desc_text = ai_resp.text.strip() if ai_resp and ai_resp.text else ""
                                        if desc_text:
                                            try:
                                                img_data = _json.loads(desc_text)
                                                return {
                                                    "src": url,
                                                    "alt": alt,
                                                    "description": img_data.get("description", ""),
                                                    "dominant_colors": img_data.get("dominant_colors", []),
                                                    "status": "success"
                                                }
                                            except Exception as pe:
                                                print(f"[Zombo Scrape] JSON parse error for image: {pe}")
                                                # If it wasn't JSON but returned text, use it as description
                                                return {
                                                    "src": url,
                                                    "alt": alt,
                                                    "description": desc_text,
                                                    "dominant_colors": [],
                                                    "status": "success"
                                                }
                            except Exception as e:
                                print(f"[Zombo Scrape] Multimodal analysis error for {url}: {e}")

                            # Fallback descriptive analysis using Gemini metadata prompt
                            desc = f"A(z) '{url.split('/')[-1]}' fájlnevű kép."
                            if alt:
                                desc += f" Helyettesítő leírása: '{alt}'."
                            try:
                                prompt = f"Következtess a kép tartalmára az URL és az alt szöveg alapján. Írd le 1 mondatban magyarul, hogy valószínűleg mit ábrázol és mi a szerepe a weboldalon. URL: {url}, ALT: {alt}"
                                ai_resp = await client.aio.models.generate_content(
                                    model='gemini-2.5-flash',
                                    contents=prompt
                                )
                                if ai_resp and ai_resp.text:
                                    desc = ai_resp.text.strip()
                            except Exception:
                                pass
                            return {
                                "src": url,
                                "alt": alt,
                                "description": desc,
                                "dominant_colors": [],
                                "status": "fallback"
                            }

                        tasks = [analyze_single_image(gemini_client, url, alt) for url, alt in valid_images]
                        images_analysis = await asyncio.gather(*tasks)

                # 4. Contact & Company Specialist Agent Task
                async def run_contact_agent():
                    nonlocal contact_analysis
                    contact_prompt = f"""
                    You are the Contact & Company Specialist Agent.
                    Your sole task is to analyze the crawled content of the website (homepage and contact/company subpages) and extract ALL contact and corporate identifier details.
                    
                    Identify:
                    1. Email addresses (e.g. info@domain.hu, sales@domain.hu).
                    2. Phone numbers (e.g. +36 20 123 4567, 06-1-1234567).
                    3. Postal / physical addresses (e.g. 1051 Budapest, Fő utca 1.).
                    4. Hungarian corporate details:
                       - Company Name (Hivatalos Cégnév, e.g. Bégé Design Kft., Teszt Bt.)
                       - Tax Number (Adószám, e.g. 12345678-1-12)
                       - Company Registration Number (Cégjegyzékszám, e.g. 01-09-123456)
                    5. Social Media links:
                       Extract any links pointing to social, video-sharing, messaging, or community platforms (e.g. Facebook, Instagram, LinkedIn, YouTube, TikTok, Twitter/X, Pinterest, Threads, WhatsApp, Telegram, Viber, GitHub, Twitch, Vimeo, etc.).
                    
                    Here are the social media links programmatically extracted from the HTML links as hints:
                    {_json.dumps(accumulated_socials, ensure_ascii=False)}
                    
                    Be extremely thorough. Look in the headers, footers, about sections, and contact sections of the provided HTML.
                    If a field is not found, return null or an empty list.
                    
                    Provide your response strictly as a JSON response in Hungarian using this format:
                    {{
                      "emails": ["email1", "email2"],
                      "phone_numbers": ["phone1", "phone2"],
                      "addresses": ["address1", "address2"],
                      "company_name": "Official company name or null",
                      "tax_number": "Tax number or null",
                      "registration_number": "Registration number or null",
                      "facebook": "Facebook link or null",
                      "instagram": "Instagram link or null",
                      "linkedin": "LinkedIn link or null",
                      "youtube": "YouTube link or null",
                      "tiktok": "TikTok link or null",
                      "pinterest": "Pinterest link or null",
                      "twitter": "Twitter/X link or null",
                      "github": "GitHub link or null",
                      "viber": "Viber link or null",
                      "whatsapp": "WhatsApp link or null",
                      "telegram": "Telegram link or null"
                      // ... you can add any other of the 30 social keys as needed
                    }}
                    
                    Here is the compiled webpage content (homepage and contact subpages):
                    {combined_content}
                    """
                    try:
                        ai_resp = await gemini_client.aio.models.generate_content(
                            model='gemini-2.5-flash',
                            contents=contact_prompt,
                            config=types.GenerateContentConfig(response_mime_type="application/json")
                        )
                        if ai_resp and ai_resp.text:
                            parsed = _json.loads(ai_resp.text)
                            for k in contact_analysis.keys():
                                if k in parsed:
                                    contact_analysis[k] = parsed[k]
                            
                            # Merge programmatically extracted socials as fallback
                            for k, v in accumulated_socials.items():
                                if k in contact_analysis and not contact_analysis[k]:
                                    contact_analysis[k] = v
                    except Exception as e:
                        print(f"[Zombo Scrape] Contact & Company Specialist Agent error: {e}")

                # 5. Product Specialist Agent Task
                async def run_product_agent():
                    nonlocal product_analysis
                    product_prompt = f"""
                    You are the Product & Service Specialist Agent.
                    Your sole task is to analyze the crawled content of the website (homepage and subpages) and extract a list of products or services offered.
                    
                    Extract up to a maximum of 10 products or services. Do not exceed 10.
                    For each product/service, extract:
                    1. name: Product or service name
                    2. price: Price with currency if found (e.g. '15 000 Ft' or '$99'), or 'N/A'
                    3. brand: Brand or manufacturer if found, or 'N/A'
                    4. description: Short description of product/service (max 2 sentences)
                    5. page_url: Absolute URL of the page containing this product/service
                    
                    Provide your response strictly as a JSON array of objects using this format:
                    [
                      {{
                        "name": "Product or service name",
                        "price": "Price with currency or 'N/A'",
                        "brand": "Brand or manufacturer or 'N/A'",
                        "description": "Short description of product/service",
                        "page_url": "Page URL"
                      }}
                    ]
                    
                    Here is the compiled webpage content:
                    {combined_content if combined_content else clean_html_for_scraper}
                    """
                    try:
                        ai_resp = await gemini_client.aio.models.generate_content(
                            model='gemini-2.5-flash',
                            contents=product_prompt,
                            config=types.GenerateContentConfig(response_mime_type="application/json")
                        )
                        if ai_resp and ai_resp.text:
                            parsed = _json.loads(ai_resp.text)
                            if isinstance(parsed, list):
                                product_analysis = parsed[:10]
                    except Exception as e:
                        print(f"[Zombo Scrape] Product Specialist Agent error: {e}")

                # Execute parallel specialist evaluations
                await asyncio.gather(
                    run_seo_agent(),
                    run_marketing_agent(),
                    run_images_analysis(),
                    run_contact_agent(),
                    run_product_agent()
                )

            # Extract final SEO scores from Agent
            score = ai_analysis["score"]
            deductions_detail = ai_analysis["deductions_detail"]
            deductions = [d["reason"] for d in deductions_detail if d.get("status") == "bad"]

            title_text = soup.title.string.strip() if soup.title and soup.title.string else ""
            desc_meta = soup.find('meta', attrs={'name': 'description'})
            if not desc_meta:
                desc_meta = soup.find('meta', attrs={'property': 'og:description'})
            desc_text = desc_meta['content'].strip() if desc_meta and desc_meta.get('content') else ""
            h1_count = len(soup.find_all('h1'))
            h2_count = len(soup.find_all('h2'))
            h3_count = len(soup.find_all('h3'))
            has_viewport = bool(soup.find('meta', attrs={'name': 'viewport'}))
            is_https = target_url.startswith('https://')

            # ── Phase 3: Sequential Brand DNA Synthesis (3 calls) ──

            # --- Default Brand DNA structure ---
            brand_personality = {
                "brand_archetype": "Általános",
                "brand_archetype_reasoning": "Nem áll rendelkezésre részletes elemzés.",
                "brand_voice": ["Informatív", "Egyszerű"],
                "alignment_score": 70,
                "alignment_reasoning": "Nem áll rendelkezésre részletes elemzés.",
                "target_audience": "Általános közönség",
                "personality_summary": "Nem áll rendelkezésre márkaszemélyiség összefoglaló.",
                "brand_coordinates": {
                    "tone": {
                        "formal_vs_casual": 50,
                        "rational_vs_emotional": 50,
                        "modern_vs_traditional": 50,
                        "simple_vs_technical": 50,
                        "authority_vs_peer": 50
                    },
                    "business": {
                        "price_segment_score": 50,
                        "price_segment_label": "Közepes",
                        "b2b_vs_b2c": 50,
                        "product_vs_service": 50
                    },
                    "visual": {
                        "minimalist_vs_decorative": 50,
                        "warmth_vs_coolness": 50,
                        "vibrancy": 50,
                        "visual_style_tags": []
                    },
                    "content": {
                        "primary_industry": "Általános",
                        "key_content_themes": [],
                        "humor_level": 0,
                        "storytelling_level": 0,
                        "educational_level": 50,
                        "promotional_level": 50
                    },
                    "engagement": {
                        "cta_aggressiveness": 50,
                        "emoji_usage": 20,
                        "hashtag_density": 30,
                        "post_length_preference": 50,
                        "interaction_asking": 30
                    }
                },
                "addressing": {
                    "mode": "vegyes",
                    "confidence": 0,
                    "evidence": []
                },
                "cta_library": {
                    "primary_ctas": [],
                    "secondary_ctas": [],
                    "slogans": [],
                    "tagline": ""
                },
                "brand_dont": {
                    "tone_restrictions": [],
                    "content_restrictions": [],
                    "visual_restrictions": []
                },
                "hashtag_strategy": {
                    "brand_hashtags": [],
                    "industry_hashtags": [],
                    "campaign_hashtags": [],
                    "max_per_platform": {"instagram": 15, "facebook": 5, "linkedin": 3}
                },
                "content_pillars": [],
                "platform_rules": {
                    "instagram": {"active": False, "tone_modifier": "", "emoji_allowed": True, "max_hashtags": 15, "preferred_format": "image_with_text", "optimal_post_length": "150-300 karakter"},
                    "facebook": {"active": False, "tone_modifier": "", "emoji_allowed": True, "max_hashtags": 5, "preferred_format": "image_with_text", "optimal_post_length": "200-500 karakter"},
                    "tiktok": {"active": False, "tone_modifier": "", "emoji_allowed": True, "max_hashtags": 5, "preferred_format": "short_video_script", "optimal_post_length": "50-150 karakter"},
                    "youtube": {"active": False, "tone_modifier": "", "emoji_allowed": False, "max_hashtags": 3, "preferred_format": "tutorial_description", "optimal_post_length": "500-1500 karakter"},
                    "linkedin": {"active": False, "tone_modifier": "", "emoji_allowed": False, "max_hashtags": 3, "preferred_format": "text_post", "optimal_post_length": "300-800 karakter"}
                },
                "visual_recipe": {
                    "color_palette": {"primary": "#808080", "secondary": "#606060", "accent": "#a0a0a0", "background": "#ffffff", "text_color": "#000000"},
                    "photography_style": "product_studio",
                    "lighting": "bright_even",
                    "composition": "centered_clean",
                    "background_type": "solid_white",
                    "mood": "clean_professional",
                    "image_prompt_prefix": "",
                    "image_prompt_suffix": "",
                    "negative_prompt": "text, watermark, logo, blurry, low quality, distorted"
                }
            }

            if gemini_client and scraped_data_json_str != "{}":
                # Prepare product summary for brand analysis
                _products_for_brand = []
                if 'product_analysis' in locals() and product_analysis:
                    _products_for_brand = product_analysis[:10]
                elif scraped_data_json_str and scraped_data_json_str != "{}":
                    try:
                        _ps = _json.loads(scraped_data_json_str)
                        if isinstance(_ps, dict) and "products" in _ps:
                            _products_for_brand = _ps["products"][:10]
                        elif isinstance(_ps, list):
                            for _pg in _ps:
                                if isinstance(_pg, dict) and "products" in _pg:
                                    _products_for_brand.extend(_pg["products"])
                            _products_for_brand = _products_for_brand[:10]
                    except Exception:
                        pass
                _products_summary = _json.dumps(_products_for_brand, ensure_ascii=False) if _products_for_brand else "Nem találtunk termékeket."

                # Detect active social platforms from contact_analysis
                _active_platforms = []
                _social_keys_map = {"facebook": "facebook", "instagram": "instagram", "tiktok": "tiktok", "youtube": "youtube", "linkedin": "linkedin"}
                for plat_key, plat_name in _social_keys_map.items():
                    if contact_analysis.get(plat_key):
                        _active_platforms.append(plat_name)
                        brand_personality["platform_rules"][plat_name]["active"] = True

                # ╔══════════════════════════════════════════════╗
                # ║  CALL 1: Brand Personality + Coordinates     ║
                # ╚══════════════════════════════════════════════╝
                yield _json.dumps({"step": "progress", "message": "Brand DNA Agent (1/3): Márkaszemélyiség és koordináták..."}) + "\n"

                brand_personality_prompt = f"""
                You are the Brand Personality & Visual Sync Agent, an expert brand strategist and brand DNA mapper.
                Your task is to synthesize the technical SEO layout, copywriting analysis, visual elements, and product catalog of this website to construct a unified Brand Personality profile AND a precise Brand DNA coordinate map.
                
                Here are your inputs:
                1. SEO Specialist Audit:
                   - Score: {ai_analysis['score']}/135
                   - Deductions: {_json.dumps(ai_analysis['deductions_detail'], ensure_ascii=False)}
                2. Copywriting & Marketing Specialist Audit:
                   - Marketing Score: {marketing_analysis['marketing_score']}/100
                   - Value Proposition: {marketing_analysis['value_proposition_evaluation']}
                   - Copywriting Frameworks: {_json.dumps(marketing_analysis['frameworks_analysis'], ensure_ascii=False)}
                   - Copy Recommendations: {_json.dumps(marketing_analysis['copy_recommendations'], ensure_ascii=False)}
                3. Visual & Color details:
                   - Color Palette Balance: Warm: {warm_pct}%, Cool: {cool_pct}%, Neutral: {neutral_pct}%
                   - Color Details: {_json.dumps(top_colors_detail, ensure_ascii=False)}
                   - Images Visual Analysis: {_json.dumps(images_analysis, ensure_ascii=False)}
                4. Product Catalog (up to 10 products):
                   {_products_summary}
                
                PART A — Evaluate the Brand Personality:
                1. Dominant Brand Archetype (e.g., Creator (Alkotó), Sage (Bölcs), Hero (Hős), Rebel (Lázadó), Lover (Szerető), Jester (Mókamester), Everyman (Átlagember), Caregiver (Gondoskodó), Ruler (Uralkodó), Magician (Varázsló), Innocent (Ártatlan), Explorer (Felfedező)). Explain why.
                2. Brand Voice Adjectives (3-5 words in Hungarian, e.g. ['közvetlen', 'szakmai', 'modern']).
                3. Visual-Textual Alignment Score (0-100 points): How well do the visual design/images/colors match the copywriting's marketing message and tone? Explain why.
                4. Target Audience Profile: Who is the ideal customer persona target?
                5. Brand Personality Summary (1 paragraph in Hungarian): Summarize the overall character, mood, and style of the brand/page.
                
                PART B — Brand DNA Coordinates (all numeric values are integers 0-100):

                Tone:
                - formal_vs_casual: 0=very formal/corporate, 100=very casual/friendly
                - rational_vs_emotional: 0=purely data-driven/rational, 100=purely emotional/feeling-based
                - modern_vs_traditional: 0=cutting-edge/trendy, 100=classic/traditional
                - simple_vs_technical: 0=extremely simple/everyday language, 100=highly technical/jargon-heavy
                - authority_vs_peer: 0=authoritative expert voice, 100=peer-to-peer friendly voice

                Business:
                - price_segment_score: 0=budget/discount, 50=mid-range, 100=luxury/premium
                - price_segment_label: One of "Budget", "Értékalapú", "Közepes", "Prémium", "Luxus"
                - b2b_vs_b2c: 0=purely B2B, 100=purely B2C
                - product_vs_service: 0=purely product-based, 100=purely service-based

                Visual:
                - minimalist_vs_decorative: 0=extremely minimalist, 100=very decorative/ornate
                - warmth_vs_coolness: 0=very warm tones, 100=very cool tones
                - vibrancy: 0=muted/desaturated, 100=vivid/saturated
                - visual_style_tags: Array of 3-6 short English tags for AI image generation prompts

                Content:
                - primary_industry: The main industry/niche in Hungarian
                - key_content_themes: Array of 3-5 main content themes in Hungarian
                - humor_level, storytelling_level, educational_level, promotional_level (0-100 each)

                Engagement:
                - cta_aggressiveness: 0=no CTAs, 100=every sentence is a CTA
                - emoji_usage: 0=never, 100=heavy emoji use
                - hashtag_density: 0=no hashtags, 100=10+ hashtags typical
                - post_length_preference: 0=ultra-short 1 sentence, 100=long essay
                - interaction_asking: 0=never asks questions, 100=always asks audience questions

                PART C — Addressing Mode:
                Analyze the website's text to determine how they address the visitor:
                - mode: "te" (informal you/tegezés), "ön" (formal you/magázás), "vegyes" (mixed), "személytelen" (impersonal)
                - confidence: 0-100 how sure you are
                - evidence: Array of 3-5 quoted phrases from the website proving the addressing mode

                Respond strictly in JSON in Hungarian:
                {{
                  "brand_archetype": "...",
                  "brand_archetype_reasoning": "...",
                  "brand_voice": ["..."],
                  "alignment_score": 85,
                  "alignment_reasoning": "...",
                  "target_audience": "...",
                  "personality_summary": "...",
                  "brand_coordinates": {{
                    "tone": {{"formal_vs_casual": 40, "rational_vs_emotional": 30, "modern_vs_traditional": 35, "simple_vs_technical": 55, "authority_vs_peer": 60}},
                    "business": {{"price_segment_score": 45, "price_segment_label": "Közepes", "b2b_vs_b2c": 80, "product_vs_service": 20}},
                    "visual": {{"minimalist_vs_decorative": 30, "warmth_vs_coolness": 55, "vibrancy": 40, "visual_style_tags": ["clean", "product-focused", "bright-lighting"]}},
                    "content": {{"primary_industry": "...", "key_content_themes": ["..."], "humor_level": 10, "storytelling_level": 20, "educational_level": 60, "promotional_level": 70}},
                    "engagement": {{"cta_aggressiveness": 70, "emoji_usage": 20, "hashtag_density": 30, "post_length_preference": 50, "interaction_asking": 30}}
                  }},
                  "addressing": {{"mode": "te", "confidence": 90, "evidence": ["Rendelj festéket", "Adj új színt"]}}
                }}
                """
                try:
                    ai_resp = await gemini_client.aio.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=brand_personality_prompt,
                        config=types.GenerateContentConfig(response_mime_type="application/json")
                    )
                    if ai_resp and ai_resp.text:
                        parsed = _json.loads(ai_resp.text)
                        for k in list(brand_personality.keys()):
                            if k in parsed:
                                if isinstance(brand_personality[k], dict) and isinstance(parsed[k], dict):
                                    brand_personality[k].update(parsed[k])
                                else:
                                    brand_personality[k] = parsed[k]
                except Exception as e:
                    print(f"[Zombo Scrape] Brand Personality Agent (Call 1) error: {e}")

                # ╔══════════════════════════════════════════════╗
                # ║  CALL 2: Brand DNA Extractor Agent           ║
                # ╚══════════════════════════════════════════════╝
                yield _json.dumps({"step": "progress", "message": "Brand DNA Agent (2/3): CTA-k, tiltólista, hashtagek, content pillérek..."}) + "\n"

                _active_platforms_str = ", ".join(_active_platforms) if _active_platforms else "Nem találtunk közösségi média jelenlétét."
                _cta_eval_str = marketing_analysis.get('cta_evaluation', 'Nem áll rendelkezésre.')
                _detected_posts_str = _json.dumps(ai_analysis.get('detected_posts', []), ensure_ascii=False)
                _addressing_mode = brand_personality["addressing"]["mode"]

                brand_dna_extractor_prompt = f"""
                You are the Brand DNA Extractor Agent, an expert in brand strategy, social media marketing, and content planning.
                
                Your task is to extract detailed, actionable brand DNA data from a website analysis. This data will be directly used by AI to generate social media posts, images, and marketing content.
                
                Inputs:
                1. Brand Archetype: {brand_personality['brand_archetype']}
                2. Brand Voice: {_json.dumps(brand_personality['brand_voice'], ensure_ascii=False)}
                3. Target Audience: {brand_personality['target_audience']}
                4. Content Tone: {ai_analysis.get('tone', 'Nem ismert')}
                5. Business Category: {ai_analysis.get('business_category', 'Nem ismert')}
                6. CTA Evaluation from Marketing Audit: {_cta_eval_str}
                7. Detected Content/Posts: {_detected_posts_str}
                8. Active Social Platforms: {_active_platforms_str}
                9. Addressing Mode: {_addressing_mode}
                10. Website Language: {lang_val if 'lang_val' in dir() else 'hu'}
                11. Website URL: {target_url}
                12. Brand Coordinates: {_json.dumps(brand_personality['brand_coordinates'], ensure_ascii=False)}
                
                Extract the following:

                A) cta_library: Extract CTA phrases actually used on the website.
                   - primary_ctas: 3-5 main action CTAs (e.g. "Vásárolok", "Rendelj most!")
                   - secondary_ctas: 2-4 softer CTAs (e.g. "Tudj meg többet", "Nézd meg")
                   - slogans: 1-3 brand slogans or taglines found on the site
                   - tagline: The single most representative brand tagline

                B) brand_dont: Things the brand should NEVER do in generated content. Analyze what the brand avoids based on its tone, style, and industry.
                   - tone_restrictions: 3-5 tone rules (e.g. "Ne használj szlenget")
                   - content_restrictions: 3-5 content rules (e.g. "Ne írj más nyelven, csak magyarul")
                   - visual_restrictions: 2-4 visual rules (e.g. "Ne használj szöveg-overlay-t a képeken")

                C) hashtag_strategy:
                   - brand_hashtags: 3-5 brand-specific hashtags (from brand name, URL)
                   - industry_hashtags: 5-10 industry/niche hashtags in Hungarian
                   - campaign_hashtags: 0-3 seasonal/campaign hashtags (if applicable)
                   - max_per_platform: {{"instagram": 15, "facebook": 5, "linkedin": 3}}

                D) content_pillars: 3-5 content pillars with ratios (must sum to 100).
                   Each pillar: name, ratio (%), description, example_title
                   Base these on the detected posts and business category.

                E) platform_rules: For each active platform ({_active_platforms_str}), define:
                   - tone_modifier: How should the tone shift for this platform? (1 sentence in Hungarian)
                   - emoji_allowed: boolean
                   - max_hashtags: integer
                   - preferred_format: one of "image_with_text", "carousel", "short_video_script", "tutorial_description", "text_post", "story"
                   - optimal_post_length: string like "150-300 karakter"

                Respond strictly in JSON in Hungarian:
                {{
                  "cta_library": {{
                    "primary_ctas": ["..."],
                    "secondary_ctas": ["..."],
                    "slogans": ["..."],
                    "tagline": "..."
                  }},
                  "brand_dont": {{
                    "tone_restrictions": ["..."],
                    "content_restrictions": ["..."],
                    "visual_restrictions": ["..."]
                  }},
                  "hashtag_strategy": {{
                    "brand_hashtags": ["#..."],
                    "industry_hashtags": ["#..."],
                    "campaign_hashtags": ["#..."],
                    "max_per_platform": {{"instagram": 15, "facebook": 5, "linkedin": 3}}
                  }},
                  "content_pillars": [
                    {{"name": "...", "ratio": 40, "description": "...", "example_title": "..."}},
                    {{"name": "...", "ratio": 30, "description": "...", "example_title": "..."}},
                    {{"name": "...", "ratio": 20, "description": "...", "example_title": "..."}},
                    {{"name": "...", "ratio": 10, "description": "...", "example_title": "..."}}
                  ],
                  "platform_rules": {{
                    "instagram": {{"tone_modifier": "...", "emoji_allowed": true, "max_hashtags": 15, "preferred_format": "carousel", "optimal_post_length": "150-300 karakter"}},
                    "facebook": {{"tone_modifier": "...", "emoji_allowed": true, "max_hashtags": 5, "preferred_format": "image_with_text", "optimal_post_length": "200-500 karakter"}}
                  }}
                }}
                """
                try:
                    ai_resp2 = await gemini_client.aio.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=brand_dna_extractor_prompt,
                        config=types.GenerateContentConfig(response_mime_type="application/json")
                    )
                    if ai_resp2 and ai_resp2.text:
                        parsed2 = _json.loads(ai_resp2.text)
                        for k in ["cta_library", "brand_dont", "hashtag_strategy", "content_pillars"]:
                            if k in parsed2:
                                if isinstance(brand_personality[k], dict) and isinstance(parsed2[k], dict):
                                    brand_personality[k].update(parsed2[k])
                                else:
                                    brand_personality[k] = parsed2[k]
                        # Merge platform_rules while preserving 'active' flags
                        if "platform_rules" in parsed2 and isinstance(parsed2["platform_rules"], dict):
                            for plat_name, plat_data in parsed2["platform_rules"].items():
                                if plat_name in brand_personality["platform_rules"]:
                                    was_active = brand_personality["platform_rules"][plat_name].get("active", False)
                                    brand_personality["platform_rules"][plat_name].update(plat_data)
                                    brand_personality["platform_rules"][plat_name]["active"] = was_active
                                else:
                                    brand_personality["platform_rules"][plat_name] = plat_data
                except Exception as e:
                    print(f"[Zombo Scrape] Brand DNA Extractor (Call 2) error: {e}")

                # ╔══════════════════════════════════════════════╗
                # ║  CALL 3: Visual Recipe Agent                 ║
                # ╚══════════════════════════════════════════════╝
                yield _json.dumps({"step": "progress", "message": "Brand DNA Agent (3/3): Vizuális recept és képgenerálási szabályok..."}) + "\n"

                _top_colors_for_recipe = _json.dumps(top_colors_detail, ensure_ascii=False)
                _images_for_recipe = _json.dumps(images_analysis[:5] if images_analysis else [], ensure_ascii=False)
                _visual_coords = _json.dumps(brand_personality["brand_coordinates"].get("visual", {}), ensure_ascii=False)

                visual_recipe_prompt = f"""
                You are the Visual Recipe Agent, an expert in visual branding, photography direction, and AI image generation prompting.

                Your task is to create a precise visual recipe for AI image generation (Imagen 4 / Stable Diffusion / DALL-E) based on the website's existing visual identity.

                Inputs:
                1. Top Colors (with hex codes): {_top_colors_for_recipe}
                2. Image Analysis (up to 5 images analyzed): {_images_for_recipe}
                3. Visual Coordinates: {_visual_coords}
                4. Brand Archetype: {brand_personality['brand_archetype']}
                5. Business Category: {ai_analysis.get('business_category', 'Nem ismert')}
                6. Visual Style Description: {ai_analysis.get('visual_style_description', 'Nem ismert')}

                Create:

                A) color_palette: Extract the 5 most important brand colors from the website.
                   - primary: Main brand color hex
                   - secondary: Second brand color hex
                   - accent: Accent/highlight color hex
                   - background: Primary background color hex
                   - text_color: Primary text color hex

                B) Photography direction:
                   - photography_style: One of "product_studio", "lifestyle", "flat_lay", "editorial", "abstract", "environmental", "macro_detail"
                   - lighting: One of "bright_even", "dramatic_shadow", "natural_daylight", "moody_low_key", "soft_diffused", "high_contrast"
                   - composition: One of "centered_clean", "rule_of_thirds", "symmetrical", "dynamic_diagonal", "overhead_flat", "close_up"
                   - background_type: One of "solid_white", "solid_color", "gradient", "contextual", "blurred_environment", "textured"
                   - mood: 2-3 word mood description in English (e.g. "clean professional minimalist")

                C) AI Image Prompt components:
                   - image_prompt_prefix: A reusable prompt PREFIX for image generation (in English, ~20-40 words). This should capture the photography style, lighting, composition, and mood.
                   - image_prompt_suffix: A reusable prompt SUFFIX (in English, ~10-20 words) for color palette and finishing style.
                   - negative_prompt: What to AVOID in generated images (in English, comma-separated). Be thorough and specific.

                Respond strictly in JSON:
                {{
                  "color_palette": {{
                    "primary": "#hex",
                    "secondary": "#hex",
                    "accent": "#hex",
                    "background": "#hex",
                    "text_color": "#hex"
                  }},
                  "photography_style": "product_studio",
                  "lighting": "bright_even",
                  "composition": "centered_clean",
                  "background_type": "solid_white",
                  "mood": "clean professional minimalist",
                  "image_prompt_prefix": "Professional product photography, clean white background, bright even lighting, commercial quality, high detail, studio shot",
                  "image_prompt_suffix": "neutral color palette, realistic, sharp focus, 8k quality",
                  "negative_prompt": "text, watermark, logo, blurry, oversaturated, cartoon, illustration, people, hands, low quality, distorted, artifacts, noise"
                }}
                """
                try:
                    ai_resp3 = await gemini_client.aio.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=visual_recipe_prompt,
                        config=types.GenerateContentConfig(response_mime_type="application/json")
                    )
                    if ai_resp3 and ai_resp3.text:
                        parsed3 = _json.loads(ai_resp3.text)
                        if isinstance(parsed3, dict):
                            brand_personality["visual_recipe"].update(parsed3)
                except Exception as e:
                    print(f"[Zombo Scrape] Visual Recipe Agent (Call 3) error: {e}")

            # ── Final: Send complete result ──
            yield _json.dumps({"step": "progress", "message": "Eredmények összeállítása..."}) + "\n"

            extracted_products = []
            extracted_contacts = contact_analysis
            try:
                if scraped_data_json_str and scraped_data_json_str != "{}":
                    parsed_scraper = _json.loads(scraped_data_json_str)
                    if isinstance(parsed_scraper, list):
                        for page in parsed_scraper:
                            if isinstance(page, dict) and "products" in page:
                                for prod in page["products"]:
                                    if prod not in extracted_products:
                                        extracted_products.append(prod)
                    elif isinstance(parsed_scraper, dict):
                        extracted_products = parsed_scraper.get("products", [])
            except Exception as pe:
                print(f"[Zombo Scrape] Error parsing products from Scraper Agent: {pe}")

            if 'product_analysis' in locals() and product_analysis:
                for prod in product_analysis:
                    name_exists = any(p.get("name", "").lower() == prod.get("name", "").lower() for p in extracted_products)
                    if not name_exists:
                        extracted_products.append(prod)
            
            extracted_products = extracted_products[:10]

            image_colors = []
            seen_img_cols = set()
            if images_analysis:
                for img in images_analysis:
                    if img and "dominant_colors" in img:
                        for col in img["dominant_colors"]:
                            c_hex = col.strip().lower()
                            if re.match(r'^#[0-9a-f]{6}$', c_hex):
                                if c_hex not in seen_img_cols:
                                    seen_img_cols.add(c_hex)
                                    image_colors.append(c_hex)

            result = {
                "status": "success",
                "url": target_url,
                "scraper_json": scraped_data_json_str,
                "products": extracted_products,
                "contacts": extracted_contacts,
                "seo": {
                    "score": score,
                    "title": title_text,
                    "description": desc_text,
                    "h1_count": h1_count,
                    "h2_count": h2_count,
                    "h3_count": h3_count,
                    "total_images": total_images,
                    "missing_alt": missing_alt,
                    "total_links": total_links,
                    "internal_links": internal_links,
                    "external_links": external_links,
                    "has_viewport": has_viewport,
                    "is_https": is_https,
                    "has_schema": has_schema,
                    "has_robots": has_robots,
                    "has_sitemap": has_sitemap,
                    "has_lang": has_lang,
                    "lang_val": lang_val,
                    "deductions": deductions,
                    "deductions_detail": deductions_detail
                },
                "visuals": {
                    "top_colors": top_colors,
                    "top_colors_detail": top_colors_detail,
                    "image_colors": image_colors,
                    "warm_pct": warm_pct,
                    "cool_pct": cool_pct,
                    "neutral_pct": neutral_pct,
                    "visual_tone": visual_tone,
                    "visual_style_description": ai_analysis["visual_style_description"]
                },
                "content": {
                    "word_count": word_count,
                    "business_category": ai_analysis["business_category"],
                    "tone": ai_analysis["tone"],
                    "summary": ai_analysis["summary"],
                    "seo_advice": ai_analysis["seo_advice"],
                    "global_improvements": ai_analysis["global_improvements"],
                    "word_style_analysis": ai_analysis["word_style_analysis"],
                    "detected_posts": ai_analysis["detected_posts"],
                    "images_analysis": images_analysis
                },
                "marketing_audit": marketing_analysis,
                "brand_personality": brand_personality
            }

            # Save latest result to local file
            try:
                import json as _json_save
                with open("latest_audit_result.json", "w", encoding="utf-8") as f_save:
                    _json_save.dump(result, f_save, indent=2, ensure_ascii=False)
                print("[Zombo Scrape] Saved latest audit result to latest_audit_result.json")
            except Exception as se:
                print(f"[Zombo Scrape] Error saving latest audit result: {se}")

            # Append to Zombo Audit History
            try:
                import json as _json_history
                history_file = "zombo_audit_history.json"
                history_data = []
                if os.path.exists(history_file):
                    try:
                        with open(history_file, "r", encoding="utf-8") as f_hist:
                            history_data = _json_history.load(f_hist)
                            if not isinstance(history_data, list):
                                history_data = []
                    except Exception:
                        history_data = []
                
                new_entry = {
                    "timestamp": datetime.utcnow().isoformat() + "Z",
                    "url": target_url,
                    "crawled_subpages": contact_urls if 'contact_urls' in locals() else [],
                    "scraper_json_str": scraped_data_json_str if 'scraped_data_json_str' in locals() else "",
                    "agents_input_content": combined_content if 'combined_content' in locals() else "",
                    "result": result
                }
                history_data.append(new_entry)
                
                with open(history_file, "w", encoding="utf-8") as f_hist_write:
                    _json_history.dump(history_data, f_hist_write, indent=2, ensure_ascii=False)
                print(f"[Zombo Scrape] Appended audit result to {history_file}")
            except Exception as he:
                print(f"[Zombo Scrape] Error saving history: {he}")

            yield _json.dumps({"step": "complete", "data": result}) + "\n"

        except Exception as e:
            yield _json.dumps({"step": "error", "message": f"Hiba a weboldal beolvasása közben: {str(e)}"}) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson")


# ─── Brand DNA Content Generation Endpoints ───────────────────────────

@app.post("/marketing/api/zombo/generate-post")
async def zombo_generate_post(request: Request):
    """Generate a social media post using Brand DNA from latest audit."""
    import json as _json
    from google import genai
    from google.genai import types

    body = await request.json()
    user_prompt = body.get("prompt", "").strip()
    platform = body.get("platform", "instagram")

    if not user_prompt:
        return JSONResponse({"error": "Kérlek adj meg egy prompt-ot."}, status_code=400)

    # Load Brand DNA
    try:
        with open("latest_audit_result.json", "r", encoding="utf-8") as f:
            audit = _json.load(f)
        bp = audit.get("brand_personality", {})
    except Exception:
        return JSONResponse({"error": "Nincs elérhető Brand DNA. Futtass először egy elemzést."}, status_code=400)

    google_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not google_key:
        return JSONResponse({"error": "GOOGLE_API_KEY hiányzik."}, status_code=500)

    client = genai.Client(api_key=google_key)

    # Build context from Brand DNA
    coords = bp.get("brand_coordinates", {})
    addressing = bp.get("addressing", {})
    cta_lib = bp.get("cta_library", {})
    dont = bp.get("brand_dont", {})
    ht = bp.get("hashtag_strategy", {})
    pillars = bp.get("content_pillars", [])
    plat_rules = bp.get("platform_rules", {}).get(platform, {})

    system_prompt = f"""Te egy márkahű social media szövegíró AI vagy. A feladatod, hogy a megadott Brand DNA koordináták alapján tökéletesen illeszkedő social media posztot írj.

MÁRKA PROFIL:
- Archetípus: {bp.get('brand_archetype', 'Általános')}
- Hangvétel: {', '.join(bp.get('brand_voice', []))}
- Célközönség: {bp.get('target_audience', 'Általános')}
- Megszólítás: {addressing.get('mode', 'vegyes')} ({"tegezés" if addressing.get('mode') == 'te' else "magázás" if addressing.get('mode') == 'ön' else "vegyes"})

HANGVÉTEL KOORDINÁTÁK (0-100):
- Formális↔Közvetlen: {coords.get('tone', {}).get('formal_vs_casual', 50)}
- Racionális↔Érzelmi: {coords.get('tone', {}).get('rational_vs_emotional', 50)}
- Modern↔Hagyományos: {coords.get('tone', {}).get('modern_vs_traditional', 50)}
- Egyszerű↔Szakmai: {coords.get('tone', {}).get('simple_vs_technical', 50)}

PLATFORM: {platform.upper()}
- Hangvétel módosító: {plat_rules.get('tone_modifier', 'nincs')}
- Emoji megengedett: {'igen' if plat_rules.get('emoji_allowed', True) else 'nem'}
- Max hashtag: {plat_rules.get('max_hashtags', 5)}
- Optimális hossz: {plat_rules.get('optimal_post_length', '150-300 karakter')}

CTA KÖNYVTÁR (használd ezeket):
- Elsődleges: {', '.join(cta_lib.get('primary_ctas', []))}
- Szlogenek: {', '.join(cta_lib.get('slogans', []))}

HASHTAG STRATÉGIA:
- Márka: {', '.join(ht.get('brand_hashtags', []))}
- Iparági: {', '.join(ht.get('industry_hashtags', []))}

TILTÓLISTA — ezeket SOHA ne csináld:
{chr(10).join('- ' + r for r in dont.get('tone_restrictions', []))}
{chr(10).join('- ' + r for r in dont.get('content_restrictions', []))}

CONTENT PILLÉREK:
{chr(10).join(f"- {p.get('name', '')}: {p.get('description', '')}" for p in pillars)}

FELADAT: Írj EGY darab {platform} posztot a felhasználó kérése alapján. Csak a posztot írd meg, semmi mást. A poszt legyen {plat_rules.get('optimal_post_length', '150-300 karakter')} hosszú. Válaszd ki a megfelelő hashtageket a stratégiából (max {plat_rules.get('max_hashtags', 5)} db). Használj CTA-t a könyvtárból ha releváns.
"""

    try:
        response = await client.aio.models.generate_content(
            model='gemini-2.5-flash',
            contents=f"{system_prompt}\n\nFelhasználó kérése: {user_prompt}",
        )
        generated_text = response.text if response and response.text else "Nem sikerült posztot generálni."
        return JSONResponse({"post": generated_text, "platform": platform})
    except Exception as e:
        return JSONResponse({"error": f"Generálási hiba: {str(e)}"}, status_code=500)


@app.post("/marketing/api/zombo/generate-image")
async def zombo_generate_image(request: Request):
    """Generate an image using Imagen and Brand DNA visual recipe."""
    import json as _json
    from google import genai
    from google.genai import types
    import base64
    import hashlib
    from pathlib import Path

    body = await request.json()
    user_prompt = body.get("prompt", "").strip()

    if not user_prompt:
        return JSONResponse({"error": "Kérlek adj meg egy prompt-ot a képgeneráláshoz."}, status_code=400)

    # Load Brand DNA visual recipe
    try:
        with open("latest_audit_result.json", "r", encoding="utf-8") as f:
            audit = _json.load(f)
        bp = audit.get("brand_personality", {})
        vr = bp.get("visual_recipe", {})
    except Exception:
        return JSONResponse({"error": "Nincs elérhető Brand DNA. Futtass először egy elemzést."}, status_code=400)

    google_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not google_key:
        return JSONResponse({"error": "GOOGLE_API_KEY hiányzik."}, status_code=500)

    client = genai.Client(api_key=google_key)

    # Build full imagen prompt from visual recipe + user input
    prefix = vr.get("image_prompt_prefix", "Professional product photography, clean background, high detail")
    suffix = vr.get("image_prompt_suffix", "realistic, sharp focus, 8k quality")
    negative = vr.get("negative_prompt", "text, watermark, blurry, low quality")

    full_prompt = f"{prefix}, {user_prompt}, {suffix}. Avoid: {negative}"

    try:
        response = await client.aio.models.generate_images(
            model='imagen-4.0-fast-generate-001',
            prompt=full_prompt,
            config=types.GenerateImagesConfig(
                number_of_images=1,
                aspect_ratio="1:1",
            )
        )

        if response and response.generated_images and len(response.generated_images) > 0:
            img_data = response.generated_images[0].image.image_bytes
            # Save image to generated-images folder
            img_dir = Path("generated-images")
            img_dir.mkdir(exist_ok=True)
            img_hash = hashlib.md5(img_data).hexdigest()[:12]
            img_filename = f"brand_{img_hash}.png"
            img_path = img_dir / img_filename
            with open(img_path, "wb") as f:
                f.write(img_data)

            img_b64 = base64.b64encode(img_data).decode("utf-8")
            return JSONResponse({
                "image_base64": img_b64,
                "image_url": f"/generated-images/{img_filename}",
                "prompt_used": full_prompt,
                "negative_prompt": negative
            })
        else:
            return JSONResponse({"error": "Az Imagen nem generált képet."}, status_code=500)

    except Exception as e:
        return JSONResponse({"error": f"Képgenerálási hiba: {str(e)}"}, status_code=500)


# Serve generated images
@app.get("/generated-images/{filename}")
async def serve_generated_image(filename: str):
    from pathlib import Path
    img_path = Path("generated-images") / filename
    if img_path.exists():
        return FileResponse(img_path, media_type="image/png")
    return JSONResponse({"error": "Kép nem található."}, status_code=404)


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN SPA CATCH-ALL — must be LAST to avoid intercepting /admin/api/* routes
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/admin")
@app.get("/admin/{path:path}")
async def admin_spa(path: str = ""):
    # Először nézzük meg, hogy létezik-e a fájl (logo, favicon, icons, stb.)
    if path:
        file_path = FRONTEND_DIST / path
        if file_path.exists() and file_path.is_file() and ".." not in path:
            return FileResponse(file_path)
    # Ha nem fájl, a React SPA index.html-t küldjük (client-side routing)
    index = FRONTEND_DIST / "index.html"
    if index.exists():
        return FileResponse(index)
    return FileResponse(THIS_DIR / "admin.html")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8000")))
